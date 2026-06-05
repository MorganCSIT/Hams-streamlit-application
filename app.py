import csv
import json
import logging
import os
import re
import shutil
import subprocess
import time
import zipfile
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, time as dtime, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from zoneinfo import ZoneInfo

import math

import numpy as np
import pandas as pd
import requests
import streamlit as st


APP_ROOT = Path(__file__).resolve().parent
BASE_URL = "https://csv.webfleet.com/extern"
API_ACTION = "showTripReportExtern"
LANG = "en"
OUTPUTFORMAT = "csv"
EXCEL_MAX_ROWS = 1_048_576
MISSING_FILTER_LABEL = "(Manquant)"
DASHBOARD_COLUMNS = [
    "tripmode",
    "start_time",
    "end_time",
    "duration",
    "distance",
    "drivername",
    "driverno",
    "objectname",
]
MERGE_OUTPUT_FOLDER = "MergedReports"
RDA_OUTPUT_FOLDER = "RDAReports"
LTR_OUTPUT_FOLDER = "LTRReports"
LTR_NOTEBOOK_PATH = APP_ROOT / "Scripts" / "Check LTR.ipynb"
AUDIT_OUTPUT_FOLDER = "AuditReports"
AUDIT_WORK_END_BUFFER_MIN = 30
AUDIT_PRE_SHIFT_BUFFER_MIN = 30
AUDIT_INTERNAL_BLOCK_GAP_MIN = 180
AUDIT_INTERNAL_BUFFER_MIN = 30
AUDIT_INTERNAL_BUF_MIN_OVERLAP_MIN = 0
AUDIT_INTERNAL_BUF_MIN_OVERLAP_RATIO = 0.0
AUDIT_MAX_REASONABLE_SPEED_KMH = 160
AUDIT_KM_MIN_FOR_FLAG_CONTEXT = 0.1
AUDIT_FULL_DAY_MINUTES = 420
AUDIT_FULL_SPAN_MINUTES = 480
AUDIT_MIN_BLOCKS_FOR_SPAN = 2
AUDIT_BLOCKS_FULL = 4
AUDIT_GAP_MERGE_MIN = 10
AUDIT_PRESTATION_61010_CODE = "61010"
AUDIT_PRESTATION_61010_BUFFER_MIN = 2
AUDIT_ENABLE_61010_FEATURE = True
AUDIT_TZ_NAME = "Europe/Zurich"
AUDIT_CHECK_PRE_SHIFT = True
AUDIT_PLAN_COLOR_MAP = {
    "as": "#FFD166",
    "inf": "#F15BB5",
    "adv": "#9C6644",
    "annule": "#D4A373",
    "annul?": "#D4A373",
    "demande d'horaire specifique": "#E9C46A",
    "demande d'horaire sp?cifique": "#E9C46A",
    "facturation (-24h)": "#7F5539",
    "avertir": "#FF99C8",
    "a ete averti - adv": "#B08968",
    "a ?t? averti - adv": "#B08968",
    "geplant": "#A68A64",
    "a avertir - inf": "#E6BE8A",
    "? avertir - inf": "#E6BE8A",
    "a avertir - adv": "#C9ADA7",
    "? avertir - adv": "#C9ADA7",
    "a ete averti - as": "#F6D6AD",
    "a ?t? averti - as": "#F6D6AD",
}
AUDIT_PLANNING_DATE_TIE_BREAKER = "monthfirst"
AUDIT_PLANNING_DATE_MIN_YEAR = 2000
AUDIT_PLANNING_DATE_MAX_YEAR = 2100
TASKS = {
    "webfleet": "Téléchargement Webfleet",
    "merge": "Fusionner des fichiers",
    "rda": "Transferts RDA",
    "ltr": "Contrôles LTR",
    "audit": "Audit Webfleet-RDA",
}
RDA_OE_MAP = {
    "NE 301": "100000000000000301",
    "SARL 201": "100000000000000201",
    "SA 101": "100000000000000101",
}
RDA_TRANSFER_TARGET_OE = "100000000000000101"
RDA_ALLOWED_TARGET_CODES = {"11000", "11100", "11200", "14000", "14100", "14200"}
RDA_CODE_61010 = "61010"
RDA_WHITELIST_CODES = {"16011", "909", "16009", "195"}
RDA_DATE_COLS = ["Date Début", "Date", "Jour", "Date de prestation"]
RDA_START_COLS = ["Début", "Heure début", "Heure Début", "Von"]
RDA_END_COLS = ["Fin", "Heure fin", "Heure Fin", "Bis"]
RDA_CODE_COLS = ["N° Prestation", "No prestation", "Code prestation", "Prestation", "Code"]
RDA_DUREE_COLS = ["Durée", "Duree", "Durée (min)", "Durée (minutes)", "Dauer_verrechnet"]
RDA_CLIENT_COLS = ["N° du client", "No client", "ID client", "Client", "KD-Nr", "KD_Nr"]
RDA_COLLAB_COLS = ["No collaborateur", "Collaborateur", "ID collaborateur", "Mitarbeiter-ID"]


logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")


@dataclass
class DownloadConfig:
    start_date: date
    end_date: date
    account: str
    username: str
    password: str
    api_key: str
    output_root: Path
    max_chunk_days: int
    request_interval_secs: int
    max_retries: int
    timeout_secs: int

    @property
    def start_str(self) -> str:
        return date_str(self.start_date)

    @property
    def end_str(self) -> str:
        return date_str(self.end_date)

    @property
    def run_folder(self) -> Path:
        return self.output_root / f"webfleet_year_download_{self.start_str}_to_{self.end_str}"

    @property
    def cache_folder(self) -> Path:
        return self.run_folder / "checkpoint_chunks"

    @property
    def output_csv(self) -> Path:
        return self.run_folder / f"webfleet_ALL_TRIPS_{self.start_str}_to_{self.end_str}.csv"

    @property
    def output_xlsx(self) -> Path:
        return self.run_folder / f"webfleet_ALL_TRIPS_{self.start_str}_to_{self.end_str}.xlsx"

    @property
    def manifest_csv(self) -> Path:
        return self.run_folder / "download_manifest.csv"


class WebfleetDownloader:
    def __init__(self, config: DownloadConfig, status_box, progress_bar):
        self.config = config
        self.status_box = status_box
        self.progress_bar = progress_bar
        self.last_request_time = 0.0
        self.session = requests.Session()
        self.session.auth = (config.username, config.password)
        self.session.headers.update(
            {
                "User-Agent": "Mozilla/5.0",
                "Accept": "text/csv, text/plain, */*",
            }
        )

    def chunk_cache_path(self, start_d: date, end_d: date) -> Path:
        return self.config.cache_folder / f"trips_{date_str(start_d)}_to_{date_str(end_d)}.csv"

    def chunk_empty_marker_path(self, start_d: date, end_d: date) -> Path:
        return self.config.cache_folder / f"EMPTY_{date_str(start_d)}_to_{date_str(end_d)}.txt"

    def build_params(self, start_d: date, end_d: date) -> dict:
        return {
            "account": self.config.account,
            "apikey": self.config.api_key,
            "action": API_ACTION,
            "outputformat": OUTPUTFORMAT,
            "lang": LANG,
            "useISO8601": "true",
            "rangefrom_string": f"{date_str(start_d)}T00:00:00",
            "rangeto_string": f"{date_str(end_d)}T23:59:59",
        }

    def wait_for_rate_limit(self) -> None:
        elapsed = time.monotonic() - self.last_request_time
        if elapsed < self.config.request_interval_secs:
            wait_s = self.config.request_interval_secs - elapsed
            self.status_box.info(f"Pause limite API : attente de {wait_s:.1f} secondes")
            time.sleep(wait_s)

    def fetch_period_once(self, start_d: date, end_d: date) -> pd.DataFrame:
        label = f"{date_str(start_d)} to {date_str(end_d)}"
        self.wait_for_rate_limit()
        self.status_box.info(f"Requête en cours : {label}")

        resp = self.session.get(
            BASE_URL,
            params=self.build_params(start_d, end_d),
            timeout=self.config.timeout_secs,
        )
        self.last_request_time = time.monotonic()

        if resp.status_code != 200:
            raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:500]}")

        text = resp.text
        if looks_like_webfleet_error(text):
            first_line = text.strip().splitlines()[0]
            if "document is empty" in first_line.lower():
                return pd.DataFrame()
            raise RuntimeError(f"Webfleet error for {label}: {first_line}")

        if not text.strip():
            return pd.DataFrame()

        return parse_csv_robust(text, self.config.cache_folder, label.replace(" ", "_"))

    def fetch_period_with_retries(self, start_d: date, end_d: date) -> pd.DataFrame:
        label = f"{date_str(start_d)} to {date_str(end_d)}"
        last_exc = None

        for attempt in range(1, self.config.max_retries + 1):
            try:
                return self.fetch_period_once(start_d, end_d)
            except Exception as exc:
                last_exc = exc
                self.status_box.warning(
                    f"{label} : tentative {attempt}/{self.config.max_retries} échouée : {exc}"
                )
                if attempt < self.config.max_retries:
                    time.sleep(min(10 * attempt, 60))

        raise RuntimeError(f"{label}: failed after {self.config.max_retries} attempts. Last error: {last_exc}")

    def save_chunk_atomic(self, df: pd.DataFrame, start_d: date, end_d: date) -> Path:
        path = self.chunk_cache_path(start_d, end_d)
        tmp_path = path.with_suffix(path.suffix + ".tmp")

        out = df.copy()
        out["__download_range_from"] = date_str(start_d)
        out["__download_range_to"] = date_str(end_d)
        out.to_csv(tmp_path, index=False, encoding="utf-8-sig")
        os.replace(tmp_path, path)
        return path

    def save_empty_marker(self, start_d: date, end_d: date) -> Path:
        path = self.chunk_empty_marker_path(start_d, end_d)
        path.write_text("empty\n", encoding="utf-8")
        return path

    def log_manifest(self, row: dict) -> None:
        row = dict(row)
        row["logged_at"] = datetime.now().isoformat(timespec="seconds")

        exists = self.config.manifest_csv.exists()
        pd.DataFrame([row]).to_csv(
            self.config.manifest_csv,
            mode="a",
            header=not exists,
            index=False,
            encoding="utf-8-sig",
        )

    def fetch_period_safe(self, start_d: date, end_d: date) -> list[Path]:
        cached_csv = self.chunk_cache_path(start_d, end_d)
        cached_empty = self.chunk_empty_marker_path(start_d, end_d)

        if cached_csv.exists():
            self.status_box.info(f"Déjà en cache : {cached_csv.name}")
            return [cached_csv]

        if cached_empty.exists():
            self.status_box.info(f"Déjà en cache comme vide : {cached_empty.name}")
            return []

        days_count = (end_d - start_d).days + 1
        label = f"{date_str(start_d)} to {date_str(end_d)}"

        try:
            df = self.fetch_period_with_retries(start_d, end_d)
            if df.empty:
                self.save_empty_marker(start_d, end_d)
                self.log_manifest(
                    {"from": date_str(start_d), "to": date_str(end_d), "status": "empty", "rows": 0, "file": "", "error": ""}
                )
                return []

            path = self.save_chunk_atomic(df, start_d, end_d)
            self.log_manifest(
                {
                    "from": date_str(start_d),
                    "to": date_str(end_d),
                    "status": "ok",
                    "rows": len(df),
                    "file": str(path),
                    "error": "",
                }
            )
            self.status_box.success(f"{path.name} enregistré avec {len(df):,} lignes")
            return [path]
        except Exception as exc:
            self.status_box.error(f"{label} : échec sur ce bloc : {exc}")
            if days_count <= 1:
                self.log_manifest(
                    {
                        "from": date_str(start_d),
                        "to": date_str(end_d),
                        "status": "failed",
                        "rows": 0,
                        "file": "",
                        "error": str(exc),
                    }
                )
                raise

            mid_d = start_d + timedelta(days=(days_count // 2) - 1)
            files = []
            files.extend(self.fetch_period_safe(start_d, mid_d))
            files.extend(self.fetch_period_safe(mid_d + timedelta(days=1), end_d))
            return files

    def run(self) -> dict:
        self.config.run_folder.mkdir(parents=True, exist_ok=True)
        self.config.cache_folder.mkdir(parents=True, exist_ok=True)

        period_list = list(make_periods(self.config.start_date, self.config.end_date, self.config.max_chunk_days))
        all_chunk_files = []
        failed = []

        for i, (start_d, end_d) in enumerate(period_list, start=1):
            self.progress_bar.progress((i - 1) / len(period_list), text=f"Téléchargement du bloc {i} sur {len(period_list)}")
            try:
                all_chunk_files.extend(self.fetch_period_safe(start_d, end_d))
            except Exception as exc:
                failed.append((date_str(start_d), date_str(end_d), str(exc)))

        self.progress_bar.progress(1.0, text="Assemblage des fichiers checkpoint")
        unique_chunk_files = dedupe_paths(all_chunk_files)
        total_rows, final_cols = combine_checkpoint_csvs(unique_chunk_files, self.config.output_csv)

        audit = audit_download(self.config)
        return {
            "failed": failed,
            "total_rows": total_rows,
            "columns": final_cols,
            "audit": audit,
            "chunk_count": len(unique_chunk_files),
        }


def date_str(value: date) -> str:
    return value.strftime("%Y-%m-%d")


def make_periods(start: date, end: date, max_days: int):
    cur = start
    while cur <= end:
        chunk_end = min(cur + timedelta(days=max_days - 1), end)
        yield cur, chunk_end
        cur = chunk_end + timedelta(days=1)


def looks_like_webfleet_error(text: str) -> bool:
    if not text or not text.strip():
        return False
    first_line = text.strip().splitlines()[0].strip()
    return bool(re.match(r"^9\d{3}\s*[,;]", first_line))


def parse_csv_robust(text: str, cache_folder: Path, label: str) -> pd.DataFrame:
    attempts = [
        dict(sep=None, engine="python"),
        dict(sep=";", engine="python", quoting=csv.QUOTE_MINIMAL),
        dict(sep=",", engine="python"),
        dict(sep=None, engine="python", on_bad_lines="skip"),
    ]

    last_exc = None
    for kwargs in attempts:
        try:
            return pd.read_csv(StringIO(text), **kwargs)
        except Exception as exc:
            last_exc = exc

    raw_path = cache_folder / f"RAW_PARSE_FAILED_{label}.csv"
    raw_path.write_text(text, encoding="utf-8")
    raise RuntimeError(f"Could not parse CSV for {label}. Raw file saved to {raw_path}. Last error: {last_exc}")


def collect_union_columns(csv_files: list[Path]) -> list[str]:
    cols = []
    for path in csv_files:
        header = pd.read_csv(path, nrows=0).columns.tolist()
        for col in header:
            if col not in cols:
                cols.append(col)
    return cols


def combine_checkpoint_csvs(csv_files: list[Path], output_csv: Path) -> tuple[int, list[str]]:
    if not csv_files:
        return 0, []

    csv_files = sorted(csv_files)
    all_cols = collect_union_columns(csv_files)
    seen_tripids = set()
    use_tripid_dedupe = "tripid" in all_cols
    first_write = True
    total_rows = 0
    tmp_output = output_csv.with_suffix(output_csv.suffix + ".tmp")

    if tmp_output.exists():
        tmp_output.unlink()

    for path in csv_files:
        for chunk in pd.read_csv(path, chunksize=50_000):
            chunk = chunk.reindex(columns=all_cols)
            if use_tripid_dedupe and "tripid" in chunk.columns:
                tripids = chunk["tripid"].astype(str)
                keep_mask = ~tripids.isin(seen_tripids)
                seen_tripids.update(tripids[keep_mask].tolist())
                chunk = chunk.loc[keep_mask]

            if chunk.empty:
                continue

            chunk.to_csv(
                tmp_output,
                mode="w" if first_write else "a",
                header=first_write,
                index=False,
                encoding="utf-8-sig",
            )
            first_write = False
            total_rows += len(chunk)

    if tmp_output.exists():
        os.replace(tmp_output, output_csv)
    return total_rows, all_cols


def write_xlsx_from_csv(input_csv: Path, output_xlsx: Path, total_rows: int) -> bool:
    if total_rows == 0 or total_rows > EXCEL_MAX_ROWS - 1:
        return False

    from openpyxl import Workbook

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("All Trips")
    first = True

    for chunk in pd.read_csv(input_csv, chunksize=50_000):
        if first:
            ws.append(list(chunk.columns))
            first = False
        for row in chunk.itertuples(index=False, name=None):
            ws.append(row)

    wb.save(output_xlsx)
    return True


def dedupe_paths(paths: list[Path]) -> list[Path]:
    seen = set()
    out = []
    for path in paths:
        if path not in seen:
            out.append(path)
            seen.add(path)
    return out


def count_csv_rows(path: Path) -> int:
    if not path.exists():
        return 0
    with path.open("r", encoding="utf-8-sig", errors="replace") as handle:
        return max(sum(1 for _ in handle) - 1, 0)


def parse_period_from_filename(path: Path):
    name = path.name
    match = re.match(r"trips_(\d{4}-\d{2}-\d{2})_to_(\d{4}-\d{2}-\d{2})\.csv$", name)
    if match:
        return date.fromisoformat(match.group(1)), date.fromisoformat(match.group(2)), "data"

    match = re.match(r"EMPTY_(\d{4}-\d{2}-\d{2})_to_(\d{4}-\d{2}-\d{2})\.txt$", name)
    if match:
        return date.fromisoformat(match.group(1)), date.fromisoformat(match.group(2)), "empty"

    return None


def days_in_period(start_d: date, end_d: date):
    cur = start_d
    while cur <= end_d:
        yield cur
        cur += timedelta(days=1)


def audit_download(config: DownloadConfig) -> dict:
    requested_days = set(days_in_period(config.start_date, config.end_date))
    covered_days = set()
    empty_days = set()

    chunk_files = sorted(config.cache_folder.glob("trips_*.csv"))
    empty_markers = sorted(config.cache_folder.glob("EMPTY_*.txt"))

    for path in chunk_files + empty_markers:
        parsed = parse_period_from_filename(path)
        if parsed is None:
            continue
        start_d, end_d, kind = parsed
        for day in days_in_period(start_d, end_d):
            covered_days.add(day)
            if kind == "empty":
                empty_days.add(day)

    manifest_failed_count = 0
    if config.manifest_csv.exists():
        manifest = pd.read_csv(config.manifest_csv)
        if "status" in manifest.columns:
            manifest_failed_count = int(manifest["status"].astype(str).str.lower().eq("failed").sum())

    raw_checkpoint_rows = sum(count_csv_rows(path) for path in chunk_files)
    final_rows = count_csv_rows(config.output_csv)

    tripid_summary = {}
    raw_tripids = []
    tripid_exists = False
    for path in chunk_files:
        try:
            header = pd.read_csv(path, nrows=0).columns.tolist()
        except Exception:
            continue
        if "tripid" in header:
            tripid_exists = True
            for chunk in pd.read_csv(path, usecols=["tripid"], chunksize=100_000):
                raw_tripids.extend(chunk["tripid"].dropna().astype(str).tolist())

    missing_tripids_count = 0
    duplicate_tripids_count = 0
    if tripid_exists:
        raw_counter = Counter(raw_tripids)
        duplicate_tripids_count = sum(1 for count in raw_counter.values() if count > 1)
        final_tripids = set()
        if config.output_csv.exists() and "tripid" in pd.read_csv(config.output_csv, nrows=0).columns.tolist():
            for chunk in pd.read_csv(config.output_csv, usecols=["tripid"], chunksize=100_000):
                final_tripids.update(chunk["tripid"].dropna().astype(str).tolist())
        missing_tripids_count = len(set(raw_counter.keys()) - final_tripids)
        tripid_summary = {
            "raw_tripids": len(raw_tripids),
            "unique_raw_tripids": len(raw_counter),
            "duplicate_tripids": duplicate_tripids_count,
            "final_tripids": len(final_tripids),
            "missing_tripids": missing_tripids_count,
        }

    missing_coverage = sorted(requested_days - covered_days)
    hard_fail = bool(missing_coverage or manifest_failed_count or missing_tripids_count)
    if not tripid_exists and raw_checkpoint_rows != final_rows:
        hard_fail = True

    return {
        "passed": not hard_fail,
        "checkpoint_files": len(chunk_files),
        "empty_markers": len(empty_markers),
        "missing_days": [date_str(day) for day in missing_coverage],
        "empty_days": [date_str(day) for day in sorted(empty_days & requested_days)],
        "manifest_failed_count": manifest_failed_count,
        "raw_checkpoint_rows": raw_checkpoint_rows,
        "final_rows": final_rows,
        "tripid_exists": tripid_exists,
        "tripid_summary": tripid_summary,
    }


def render_export_buttons(config: DownloadConfig) -> None:
    if not config.output_csv.exists():
        return

    st.subheader("Export")
    if config.output_csv.exists():
        with config.output_csv.open("rb") as handle:
            st.download_button(
                "Télécharger le CSV",
                handle,
                file_name=config.output_csv.name,
                mime="text/csv",
            )

    total_rows = count_csv_rows(config.output_csv)
    if st.button("Créer le fichier Excel", disabled=total_rows == 0 or total_rows > EXCEL_MAX_ROWS - 1):
        with st.spinner("Création du fichier Excel..."):
            made_xlsx = write_xlsx_from_csv(config.output_csv, config.output_xlsx, total_rows)
        if made_xlsx:
            st.session_state["excel_ready_path"] = str(config.output_xlsx)
            st.success(f"Le fichier Excel est prêt : {config.output_xlsx}")
        else:
            st.warning("Le fichier Excel n'a pas été créé car la sortie est vide ou trop grande pour une seule feuille.")

    excel_ready_path = Path(st.session_state.get("excel_ready_path", ""))
    if excel_ready_path == config.output_xlsx and config.output_xlsx.exists():
        with config.output_xlsx.open("rb") as handle:
            st.download_button(
                "Télécharger l'Excel",
                handle,
                file_name=config.output_xlsx.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


@st.cache_data(show_spinner=False)
def load_dashboard_data(path: str, modified_time: float) -> pd.DataFrame:
    _ = modified_time
    return pd.read_csv(path)


def find_column(columns: list[str], target: str) -> str | None:
    normalized_target = target.lower().replace("_", "")
    for column in columns:
        normalized_column = column.lower().replace("_", "")
        if normalized_column == normalized_target:
            return column
    return None


def coerce_datetime(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce")


def coerce_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def format_datetime_series(series: pd.Series) -> pd.Series:
    parsed = pd.to_datetime(series, errors="coerce")
    formatted = parsed.dt.strftime("%Y-%m-%d %H:%M:%S")
    return formatted.fillna(series.astype(str))


def format_duration_hh_mm_ss(series: pd.Series) -> pd.Series:
    seconds = coerce_numeric(series)

    def format_one(value):
        if pd.isna(value):
            return ""
        total_seconds = int(round(value))
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        remaining_seconds = total_seconds % 60
        return f"{hours}:{minutes:02d}:{remaining_seconds:02d}"

    return seconds.apply(format_one)


def missing_value_mask(series: pd.Series) -> pd.Series:
    return series.isna() | series.astype(str).str.strip().eq("")


def filter_options_with_missing(series: pd.Series) -> list[str]:
    present_values = sorted(series[~missing_value_mask(series)].astype(str).unique().tolist())
    if missing_value_mask(series).any():
        return [MISSING_FILTER_LABEL] + present_values
    return present_values


def apply_categorical_filter(df: pd.DataFrame, column: str, selected: list[str]) -> pd.DataFrame:
    if not selected:
        return df

    mask = pd.Series(False, index=df.index)
    if MISSING_FILTER_LABEL in selected:
        mask = mask | missing_value_mask(df[column])

    selected_present = [value for value in selected if value != MISSING_FILTER_LABEL]
    if selected_present:
        mask = mask | df[column].astype(str).isin(selected_present)

    return df[mask]


def render_dashboard(csv_path: Path) -> None:
    if not csv_path.exists():
        st.info("Téléchargez d'abord les données pour ouvrir le tableau de bord.")
        return

    try:
        df = load_dashboard_data(str(csv_path), csv_path.stat().st_mtime)
    except Exception as exc:
        st.error(f"Impossible de charger le CSV du tableau de bord : {exc}")
        return

    column_map = {name: find_column(df.columns.tolist(), name) for name in DASHBOARD_COLUMNS}
    available_dashboard_cols = [column_map[name] for name in DASHBOARD_COLUMNS if column_map[name]]
    missing_cols = [name for name in DASHBOARD_COLUMNS if not column_map[name]]

    if not available_dashboard_cols:
        st.warning("Aucune des colonnes attendues pour le tableau de bord n'a été trouvée dans le CSV de sortie.")
        st.dataframe(pd.DataFrame({"available_columns": df.columns.tolist()}), width="stretch")
        return

    st.subheader("Tableau de bord")
    st.caption(f"{len(df):,} trajets chargés depuis {csv_path.name}. Les indicateurs et le tableau se mettent à jour avec les filtres.")

    with st.expander("Colonnes", expanded=True):
        selected_display_cols = []
        toggle_cols = st.columns(4)
        for index, wanted_name in enumerate(DASHBOARD_COLUMNS):
            actual_col = column_map[wanted_name]
            if actual_col:
                enabled = toggle_cols[index % 4].toggle(actual_col, value=True, key=f"show_{wanted_name}")
                if enabled:
                    selected_display_cols.append(actual_col)
            else:
                toggle_cols[index % 4].caption(f"{wanted_name} : manquant")

    filters = st.container()
    filtered = df.copy()

    with filters:
        filter_cols = st.columns(4)

        if column_map["tripmode"]:
            values = filter_options_with_missing(filtered[column_map["tripmode"]])
            selected = filter_cols[0].multiselect("Mode de trajet", values, default=values)
            filtered = apply_categorical_filter(filtered, column_map["tripmode"], selected)

        if column_map["drivername"]:
            values = filter_options_with_missing(filtered[column_map["drivername"]])
            selected = filter_cols[1].multiselect("Conducteur", values)
            filtered = apply_categorical_filter(filtered, column_map["drivername"], selected)

        if column_map["driverno"]:
            values = filter_options_with_missing(filtered[column_map["driverno"]])
            selected = filter_cols[2].multiselect("N° conducteur", values)
            filtered = apply_categorical_filter(filtered, column_map["driverno"], selected)

        if column_map["objectname"]:
            values = filter_options_with_missing(filtered[column_map["objectname"]])
            selected = filter_cols[3].multiselect("Objet", values)
            filtered = apply_categorical_filter(filtered, column_map["objectname"], selected)

        date_filter_cols = st.columns(3)
        start_col = column_map["start_time"]
        end_col = column_map["end_time"]
        if start_col:
            start_times = coerce_datetime(filtered[start_col])
            valid_start_times = start_times.dropna()
            if not valid_start_times.empty:
                min_day = valid_start_times.min().date()
                max_day = valid_start_times.max().date()
                selected_days = date_filter_cols[0].date_input("Plage de dates de début", value=(min_day, max_day))
                if isinstance(selected_days, tuple) and len(selected_days) == 2:
                    start_day, end_day = selected_days
                    mask = start_times.dt.date.between(start_day, end_day)
                    filtered = filtered[mask.fillna(False)]

        if column_map["distance"]:
            distance_km = coerce_numeric(filtered[column_map["distance"]]) / 1000
            valid_distance = distance_km.dropna()
            if not valid_distance.empty:
                min_distance = float(valid_distance.min())
                max_distance = float(valid_distance.max())
                if min_distance < max_distance:
                    selected_distance = date_filter_cols[1].slider(
                        "Plage de distance (km)",
                        min_value=min_distance,
                        max_value=max_distance,
                        value=(min_distance, max_distance),
                    )
                    filtered = filtered[distance_km.between(*selected_distance).fillna(False)]
                else:
                    date_filter_cols[1].metric("Distance km", f"{min_distance:,.2f}")

        if column_map["duration"]:
            duration_hours = coerce_numeric(filtered[column_map["duration"]]) / 3600
            valid_duration = duration_hours.dropna()
            if not valid_duration.empty:
                min_duration = float(valid_duration.min())
                max_duration = float(valid_duration.max())
                if min_duration < max_duration:
                    selected_duration = date_filter_cols[2].slider(
                        "Plage de durée (heures)",
                        min_value=min_duration,
                        max_value=max_duration,
                        value=(min_duration, max_duration),
                    )
                    filtered = filtered[duration_hours.between(*selected_duration).fillna(False)]
                else:
                    date_filter_cols[2].metric("Durée en heures", f"{min_duration:,.2f}")

        search = st.text_input("Recherche")
        if search:
            search_cols = [col for col in [column_map["drivername"], column_map["driverno"], column_map["objectname"]] if col]
            if search_cols:
                mask = pd.Series(False, index=filtered.index)
                for col in search_cols:
                    mask = mask | filtered[col].astype(str).str.contains(search, case=False, na=False)
                filtered = filtered[mask]

    if missing_cols:
        st.caption(f"Colonnes manquantes dans ce CSV : {', '.join(missing_cols)}")

    st.subheader("Résumé filtré")
    filtered_metric_cols = st.columns(5)
    filtered_metric_cols[0].metric("Trajets", f"{len(filtered):,}")
    if column_map["distance"]:
        filtered_distance_km = coerce_numeric(filtered[column_map["distance"]]).sum(skipna=True) / 1000
        filtered_metric_cols[1].metric("Distance", f"{filtered_distance_km:,.1f} km")
    else:
        filtered_metric_cols[1].metric("Distance", "n/a")
    if column_map["duration"]:
        filtered_duration_seconds = coerce_numeric(filtered[column_map["duration"]]).sum(skipna=True)
        filtered_metric_cols[2].metric("Secondes", f"{filtered_duration_seconds:,.0f}")
        filtered_metric_cols[3].metric("Minutes", f"{filtered_duration_seconds / 60:,.2f}")
        filtered_metric_cols[4].metric("Heures", f"{filtered_duration_seconds / 3600:,.2f}")
        st.caption("La colonne durée du CSV est en secondes. Les minutes et heures sont calculées depuis ce total brut.")
    else:
        filtered_metric_cols[2].metric("Secondes", "n/a")
        filtered_metric_cols[3].metric("Minutes", "n/a")
        filtered_metric_cols[4].metric("Heures", "n/a")
    display_df = filtered.copy()
    for category_name in ["tripmode", "drivername", "driverno", "objectname"]:
        actual_col = column_map[category_name]
        if actual_col and actual_col in display_df.columns:
            values = display_df[actual_col]
            display_df[actual_col] = values.astype("string")
            display_df.loc[missing_value_mask(values), actual_col] = MISSING_FILTER_LABEL

    for time_col_name in ["start_time", "end_time"]:
        actual_col = column_map[time_col_name]
        if actual_col and actual_col in display_df.columns:
            display_df[actual_col] = format_datetime_series(display_df[actual_col])

    if column_map["duration"] and column_map["duration"] in display_df.columns:
        duration_col = column_map["duration"]
        display_df["duration_hh_mm_ss"] = format_duration_hh_mm_ss(display_df[duration_col])
        if duration_col in selected_display_cols:
            duration_index = selected_display_cols.index(duration_col)
            selected_display_cols = selected_display_cols.copy()
            selected_display_cols[duration_index] = "duration_hh_mm_ss"

    if column_map["distance"] and column_map["distance"] in display_df.columns:
        distance_col = column_map["distance"]
        display_df["distance_km"] = coerce_numeric(display_df[distance_col]) / 1000
        if distance_col in selected_display_cols:
            distance_index = selected_display_cols.index(distance_col)
            selected_display_cols = selected_display_cols.copy()
            selected_display_cols[distance_index] = "distance_km"
    if selected_display_cols:
        st.dataframe(display_df[selected_display_cols], width="stretch", hide_index=True)
    else:
        st.warning("Activez au moins une colonne pour afficher le tableau.")

    export_cols = selected_display_cols or available_dashboard_cols
    export_df = display_df
    st.download_button(
        "Télécharger le CSV filtré",
        export_df[export_cols].to_csv(index=False, encoding="utf-8-sig"),
        file_name=f"filtered_{csv_path.name}",
        mime="text/csv",
    )


def read_csv_flex(source) -> pd.DataFrame:
    try:
        return pd.read_csv(source, sep=None, engine="python")
    except Exception as first_exc:
        for sep in [",", ";", "\t", "|"]:
            try:
                if hasattr(source, "seek"):
                    source.seek(0)
                return pd.read_csv(source, sep=sep, engine="python")
            except Exception:
                pass
        for enc in ["utf-8", "utf-8-sig", "latin-1"]:
            try:
                if hasattr(source, "seek"):
                    source.seek(0)
                return pd.read_csv(source, sep=None, engine="python", encoding=enc)
            except Exception:
                pass
        raise RuntimeError(f"Could not read CSV. Last error: {first_exc}")


def read_excel_flex(source) -> pd.DataFrame:
    try:
        xls = pd.ExcelFile(source)
    except Exception as exc:
        try:
            if hasattr(source, "seek"):
                source.seek(0)
            xls = pd.ExcelFile(source, engine="openpyxl")
        except Exception as exc2:
            raise RuntimeError(f"Could not open Excel file. Errors: {exc} / {exc2}")

    chosen_sheet = None
    for sheet in xls.sheet_names:
        try:
            df_head = pd.read_excel(xls, sheet_name=sheet, nrows=5)
            if df_head.shape[1] > 0 and df_head.shape[0] > 0:
                chosen_sheet = sheet
                break
        except Exception:
            continue

    if chosen_sheet is None:
        chosen_sheet = xls.sheet_names[0]

    try:
        return pd.read_excel(xls, sheet_name=chosen_sheet)
    except Exception:
        return pd.read_excel(source, sheet_name=chosen_sheet, engine="openpyxl")


def read_any_flex(source, filename: str) -> pd.DataFrame:
    ext = Path(filename.lower()).suffix
    if ext == ".csv":
        return read_csv_flex(source)
    if ext in [".xlsx", ".xls"]:
        return read_excel_flex(source)
    raise ValueError(f"Unsupported file type for '{filename}'")


def clean_illegal_excel_chars(df: pd.DataFrame) -> pd.DataFrame:
    illegal_char_re = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
    out = df.copy()

    def clean_value(value):
        try:
            if pd.isna(value):
                return value
        except Exception:
            pass
        return illegal_char_re.sub("", str(value))

    for col in out.select_dtypes(include=["object"]).columns:
        out[col] = out[col].map(clean_value)
    return out


def merge_dataframes(named_frames: list[tuple[str, pd.DataFrame]], stop_on_sanity_mismatch: bool) -> dict:
    if len(named_frames) < 2:
        raise ValueError("Please provide at least 2 files to merge.")

    dfs = []
    row_counts = []
    schemas = []

    first_name, first_df = named_frames[0]
    first_df = first_df.loc[:, ~first_df.columns.astype(str).str.contains("^Unnamed")]
    if first_df.empty:
        raise AssertionError(f"First file appears empty: {first_name}")

    ref_cols = list(first_df.columns)
    dfs.append(first_df[ref_cols])
    row_counts.append(len(first_df))
    schemas.append({"file": first_name, "rows": len(first_df), "columns": len(ref_cols)})

    for filename, df in named_frames[1:]:
        df = df.loc[:, ~df.columns.astype(str).str.contains("^Unnamed")]
        if df.empty:
            raise AssertionError(f"File appears empty: {filename}")

        if set(df.columns) != set(ref_cols):
            missing = [col for col in ref_cols if col not in df.columns]
            extra = [col for col in df.columns if col not in ref_cols]
            raise AssertionError(
                f"Schema mismatch in {filename}.\n"
                f"Missing columns: {missing}\n"
                f"Extra columns: {extra}\n"
                "All files must have the same header names."
            )

        df = df.reindex(columns=ref_cols)
        dfs.append(df)
        row_counts.append(len(df))
        schemas.append({"file": filename, "rows": len(df), "columns": len(df.columns)})

    merged = pd.concat(dfs, ignore_index=True)
    sum_input_rows = sum(row_counts)
    merged_rows = len(merged)

    if merged_rows != sum_input_rows and stop_on_sanity_mismatch:
        raise AssertionError(
            f"Sanity check failed: merged={merged_rows} vs sum={sum_input_rows}."
        )

    return {
        "merged": clean_illegal_excel_chars(merged),
        "schemas": schemas,
        "row_counts": row_counts,
        "sum_input_rows": sum_input_rows,
        "merged_rows": merged_rows,
        "sanity_passed": merged_rows == sum_input_rows,
    }


def write_merged_output(merged: pd.DataFrame, output_dir: Path, output_basename: str, output_format: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_basename = re.sub(r"[^A-Za-z0-9_.-]+", "_", output_basename.strip()) or "merged"

    if output_format.lower() == "xlsx":
        out_path = output_dir / f"{safe_basename}_{timestamp}.xlsx"
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            merged.to_excel(writer, index=False, sheet_name="merged")
    else:
        out_path = output_dir / f"{safe_basename}_{timestamp}.csv"
        merged.to_csv(out_path, index=False, encoding="utf-8-sig", sep=";")

    return out_path


def render_download_for_path(path: Path, label: str) -> None:
    if not path.exists():
        return
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if path.suffix.lower() == ".xlsx" else "text/csv"
    with path.open("rb") as handle:
        st.download_button(label, handle, file_name=path.name, mime=mime)


def render_webfleet_task() -> None:
    st.title("Téléchargement des journaux Webfleet")

    default_output = Path.cwd() / "WebfleetReports"

    with st.expander("API Webfleet", expanded=True):
        api_cols = st.columns(4)
        account = api_cols[0].text_input("Compte")
        username = api_cols[1].text_input("Utilisateur")
        password = api_cols[2].text_input("Mot de passe", type="password")
        api_key = api_cols[3].text_input("Clé API", type="password")

    with st.expander("Sortie locale et paramètres avancés", expanded=True):
        output_root = Path(st.text_input("Dossier de sortie local", value=str(default_output))).expanduser()
        advanced_cols = st.columns(4)
        max_chunk_days = advanced_cols[0].number_input("Jours max par requête", min_value=1, max_value=31, value=7)
        request_interval_secs = advanced_cols[1].number_input("Secondes entre requêtes", min_value=0, max_value=600, value=61)
        max_retries = advanced_cols[2].number_input("Nombre max de tentatives", min_value=1, max_value=20, value=5)
        timeout_secs = advanced_cols[3].number_input("Timeout requête en secondes", min_value=10, max_value=1200, value=300)

    today = date.today()
    cols = st.columns(2)
    with cols[0]:
        start_date = st.date_input("Date de début", value=date(today.year, 1, 1))
    with cols[1]:
        end_date = st.date_input("Date de fin", value=today)

    if start_date > end_date:
        st.error("La date de début doit être antérieure ou égale à la date de fin.")
        return

    config = DownloadConfig(
        start_date=start_date,
        end_date=end_date,
        account=account.strip(),
        username=username.strip(),
        password=password,
        api_key=api_key,
        output_root=output_root,
        max_chunk_days=int(max_chunk_days),
        request_interval_secs=int(request_interval_secs),
        max_retries=int(max_retries),
        timeout_secs=int(timeout_secs),
    )

    st.caption(f"Dossier d'exécution local : {config.run_folder}")

    ready = all([config.account, config.username, config.password, config.api_key])
    if not ready:
        st.info("Saisissez vos identifiants Webfleet pour commencer.")

    download_tab, dashboard_tab = st.tabs(["Téléchargement", "Tableau de bord"])

    with download_tab:
        if st.button("Télécharger les trajets", type="primary", disabled=not ready):
            status_box = st.empty()
            progress_bar = st.progress(0.0, text="Démarrage")

            try:
                result = WebfleetDownloader(config, status_box, progress_bar).run()
                st.session_state["latest_output_csv"] = str(config.output_csv)
            except Exception as exc:
                st.exception(exc)
                return

            st.subheader("Résultat")
            metric_cols = st.columns(4)
            metric_cols[0].metric("Lignes", f"{result['total_rows']:,}")
            metric_cols[1].metric("Fichiers checkpoint", f"{result['chunk_count']:,}")
            metric_cols[2].metric("Colonnes", f"{len(result['columns']):,}")
            metric_cols[3].metric("Blocs échoués", f"{len(result['failed']):,}")

            if result["total_rows"] > 0:
                st.success(f"Trajets téléchargés et enregistrés localement pour le tableau de bord : {config.output_csv}")
                render_export_buttons(config)
            else:
                st.warning("Aucune ligne de trajet trouvée. Des marqueurs checkpoint vides peuvent quand même avoir été créés.")

            if result["failed"]:
                st.error("Certaines périodes ont échoué définitivement. Relancez avec les mêmes paramètres pour reprendre les blocs en cache.")
                st.dataframe(pd.DataFrame(result["failed"], columns=["from", "to", "error"]), width="stretch")

            audit = result["audit"]
            st.subheader("Audit")
            audit_cols = st.columns(4)
            audit_cols[0].metric("Audit", "Réussi" if audit["passed"] else "Échec")
            audit_cols[1].metric("Lignes checkpoint brutes", f"{audit['raw_checkpoint_rows']:,}")
            audit_cols[2].metric("Lignes finales", f"{audit['final_rows']:,}")
            audit_cols[3].metric("Marqueurs vides", f"{audit['empty_markers']:,}")

            if audit["missing_days"]:
                st.error("Couverture checkpoint manquante pour ces jours :")
                st.write(audit["missing_days"])

            if audit["manifest_failed_count"]:
                st.error(f"Le manifeste contient {audit['manifest_failed_count']} périodes en échec.")

            if audit["tripid_exists"]:
                st.write("Résumé des Trip ID")
                st.dataframe(pd.DataFrame([audit["tripid_summary"]]), width="stretch")

            if audit["empty_days"]:
                with st.expander("Jours marqués vides"):
                    st.write(audit["empty_days"])

        elif config.output_csv.exists():
            render_export_buttons(config)

    with dashboard_tab:
        dashboard_path = Path(st.session_state.get("latest_output_csv", config.output_csv))
        render_dashboard(dashboard_path)


def uploaded_merge_frames(uploaded_files) -> list[tuple[str, pd.DataFrame]]:
    frames = []
    for uploaded in uploaded_files:
        data = BytesIO(uploaded.getvalue())
        frames.append((uploaded.name, read_any_flex(data, uploaded.name)))
    return frames


def local_merge_frames(search_dir: Path, patterns: list[str], include_subfolders: bool) -> list[tuple[str, pd.DataFrame]]:
    files = []
    for pattern in patterns:
        iterator = search_dir.rglob(pattern) if include_subfolders else search_dir.glob(pattern)
        files.extend(path for path in iterator if path.is_file() and not path.name.startswith("."))
    paths = sorted(set(files))
    return [(str(path), read_any_flex(path, path.name)) for path in paths]


def render_merge_task() -> None:
    st.title("Fusion de fichiers")
    st.caption("Fusionne les fichiers CSV/XLSX/XLS avec les mêmes noms d'en-têtes. L'ordre des colonnes suit le premier fichier.")

    default_output = Path.cwd() / MERGE_OUTPUT_FOLDER
    upload_mode = "Téléverser des fichiers"
    folder_mode = "Dossier local"
    source_mode = st.radio("Source des fichiers", [upload_mode, folder_mode], horizontal=True)

    with st.expander("Paramètres de fusion", expanded=True):
        settings_cols = st.columns(4)
        min_required_files = settings_cols[0].number_input("Nombre minimum de fichiers", min_value=2, max_value=1000, value=2)
        stop_on_sanity_mismatch = settings_cols[1].toggle("Arrêter si le contrôle échoue", value=True)
        output_format = settings_cols[2].selectbox("Format de sortie", ["xlsx", "csv"], index=0)
        output_basename = settings_cols[3].text_input("Nom de base de sortie", value="merged_2025_herewego")
        output_dir = Path(st.text_input("Dossier de sortie local", value=str(default_output))).expanduser()

    uploaded_files = []
    search_dir = None
    include_subfolders = True
    file_patterns = ["*.csv", "*.xlsx", "*.xls"]

    if source_mode == upload_mode:
        uploaded_files = st.file_uploader(
            "Fichiers à fusionner",
            type=["csv", "xlsx", "xls"],
            accept_multiple_files=True,
        )
    else:
        folder_cols = st.columns([3, 1])
        search_dir = Path(folder_cols[0].text_input("Dossier de recherche", value=str(Path.cwd() / "to_merge"))).expanduser()
        include_subfolders = folder_cols[1].toggle("Inclure les sous-dossiers", value=True)
        patterns_raw = st.text_input("Modèles de fichiers", value="*.csv,*.xlsx,*.xls")
        file_patterns = [pattern.strip() for pattern in patterns_raw.split(",") if pattern.strip()]

    if st.button("Fusionner les fichiers", type="primary"):
        try:
            if source_mode == upload_mode:
                if len(uploaded_files) < int(min_required_files):
                    raise ValueError(f"{len(uploaded_files)} fichier(s) téléversé(s). Fournissez au moins {int(min_required_files)} fichiers.")
                named_frames = uploaded_merge_frames(uploaded_files)
            else:
                if search_dir is None or not search_dir.exists():
                    raise FileNotFoundError(f"Le dossier de recherche n'existe pas : {search_dir}")
                named_frames = local_merge_frames(search_dir, file_patterns, include_subfolders)
                if len(named_frames) < int(min_required_files):
                    raise ValueError(
                        f"{len(named_frames)} fichier(s) trouvé(s) dans '{search_dir}'. "
                        f"Fournissez au moins {int(min_required_files)} fichiers."
                    )

            result = merge_dataframes(named_frames, stop_on_sanity_mismatch=stop_on_sanity_mismatch)
            out_path = write_merged_output(result["merged"], output_dir, output_basename, output_format)
            st.session_state["latest_merge_output"] = str(out_path)
            st.session_state["latest_merge_result"] = {
                key: value
                for key, value in result.items()
                if key != "merged"
            }
        except Exception as exc:
            st.exception(exc)
            return

    latest_output = Path(st.session_state.get("latest_merge_output", ""))
    latest_result = st.session_state.get("latest_merge_result")
    if latest_result:
        st.subheader("Résumé de fusion")
        metric_cols = st.columns(4)
        metric_cols[0].metric("Fichiers fusionnés", f"{len(latest_result['schemas']):,}")
        metric_cols[1].metric("Lignes en entrée", f"{latest_result['sum_input_rows']:,}")
        metric_cols[2].metric("Lignes fusionnées", f"{latest_result['merged_rows']:,}")
        metric_cols[3].metric("Contrôle", "Réussi" if latest_result["sanity_passed"] else "Échec")
        st.dataframe(pd.DataFrame(latest_result["schemas"]), width="stretch", hide_index=True)

        if latest_output.exists():
            st.success(f"Sortie enregistrée dans : {latest_output}")
            render_download_for_path(latest_output, "Télécharger le fichier fusionné")

            preview = read_any_flex(latest_output, latest_output.name).head(10)
            st.subheader("Aperçu")
            st.dataframe(preview, width="stretch", hide_index=True)


def rda_pick_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    return next((col for col in candidates if col in df.columns), None)


def rda_norm_code(value) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if pd.notna(value) and abs(value - round(value)) < 1e-9:
            return str(int(round(value)))
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    if text.isdigit():
        return text
    runs = re.findall(r"\d+", text)
    return sorted(runs, key=len, reverse=True)[0] if runs else text


def rda_to_minutes(value) -> float:
    if pd.isna(value):
        return float("nan")
    if isinstance(value, pd.Timestamp):
        return int(value.hour) * 60 + int(value.minute)
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return int(value.hour) * 60 + int(value.minute)
    if isinstance(value, (float, int)) and not isinstance(value, bool):
        number = float(value)
        if 0 <= number < 1.0:
            return int(round(number * 24 * 60)) % 1440
        if 0 <= number < 24 * 60 and abs(number - round(number)) < 1e-9:
            return int(round(number))
        if 0 <= number < 24:
            return int(round(number * 60))
    text = str(value).strip()
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.notna(parsed):
        return int(parsed.hour) * 60 + int(parsed.minute)
    parts = text.split()[-1].split(":")
    if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
        return int(parts[0]) * 60 + int(parts[1])
    return float("nan")


def rda_duration_from_minutes(start_min, end_min) -> float:
    if pd.isna(start_min) or pd.isna(end_min):
        return float("nan")
    diff = int(end_min) - int(start_min)
    return diff if diff >= 0 else diff + 1440


def rda_end_abs_min(end_min, start_min) -> float:
    if pd.isna(end_min) or pd.isna(start_min):
        return float("nan")
    end_value = int(end_min)
    start_value = int(start_min)
    return end_value if end_value >= start_value else end_value + 1440


def rda_minutes_to_hhmmss(value) -> str:
    if pd.isna(value):
        return ""
    minutes = int(value) % 1440
    return f"{minutes // 60:02d}:{minutes % 60:02d}:00"


def rda_format_date(value) -> str:
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    return "" if pd.isna(parsed) else parsed.strftime("%d.%m.%Y")


def rda_format_time(value) -> str:
    minutes = rda_to_minutes(value)
    if pd.isna(minutes):
        return ""
    minutes = int(minutes) % 1440
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def rda_safe_folder_name(value) -> str:
    return "".join(char if char.isalnum() or char in " ._ +-" else "_" for char in str(value)).strip() or "unknown"


def rda_read_input(source, filename: str) -> pd.DataFrame:
    ext = Path(filename).suffix.lower()
    if ext == ".xlsx":
        return pd.read_excel(source)
    if ext == ".csv":
        return pd.read_csv(source)
    raise ValueError(f"Unsupported RDA file type: {ext}. Use .xlsx or .csv.")


def rda_detect_columns(df: pd.DataFrame) -> dict:
    cols = {
        "date": rda_pick_col(df, RDA_DATE_COLS),
        "start": rda_pick_col(df, RDA_START_COLS),
        "end": rda_pick_col(df, RDA_END_COLS),
        "code": rda_pick_col(df, RDA_CODE_COLS),
        "duration": rda_pick_col(df, RDA_DUREE_COLS),
        "client": rda_pick_col(df, RDA_CLIENT_COLS),
        "collab": rda_pick_col(df, RDA_COLLAB_COLS),
    }
    missing = [name for name, col in cols.items() if col is None]
    if missing:
        raise ValueError(f"Missing required RDA columns: {missing}")
    return cols


def rda_duration_sanity(df: pd.DataFrame, cols: dict) -> tuple[pd.DataFrame, dict]:
    check = df.copy()
    check["_start_min_calc"] = check[cols["start"]].apply(rda_to_minutes)
    check["_end_min_calc"] = check[cols["end"]].apply(rda_to_minutes)
    check["_duration_from_times"] = check.apply(
        lambda row: rda_duration_from_minutes(row["_start_min_calc"], row["_end_min_calc"]),
        axis=1,
    )
    check["_duration_col"] = pd.to_numeric(check[cols["duration"]], errors="coerce")
    check["_delta"] = check["_duration_col"] - check["_duration_from_times"]
    bad = check[check["_delta"].fillna(0) != 0].copy()
    summary = {
        "rows": len(check),
        "duration_column_total": int(check["_duration_col"].fillna(0).sum()),
        "duration_from_times_total": int(check["_duration_from_times"].fillna(0).sum()),
        "difference": int(check["_duration_col"].fillna(0).sum() - check["_duration_from_times"].fillna(0).sum()),
        "mismatch_rows": len(bad),
    }
    return bad, summary


def rda_normalize_duration_from_times(df: pd.DataFrame, cols: dict) -> pd.DataFrame:
    out = df.copy()
    starts = out[cols["start"]].apply(rda_to_minutes)
    ends = out[cols["end"]].apply(rda_to_minutes)
    calculated = [
        rda_duration_from_minutes(start, end)
        for start, end in zip(starts, ends)
    ]
    backup_col = f"{cols['duration']}_original"
    if backup_col not in out.columns:
        out[backup_col] = out[cols["duration"]]
    original = pd.to_numeric(out[cols["duration"]], errors="coerce")
    out[cols["duration"]] = pd.Series(calculated, index=out.index).where(pd.Series(calculated, index=out.index).notna(), original).astype("Int64")
    return out


def rda_adjust_61010(df: pd.DataFrame, cols: dict, apply_adjustment: bool = True, enable_whitelist_transfer: bool = True) -> dict:
    df = df.copy()
    df[cols["code"]] = df[cols["code"]].apply(rda_norm_code)
    whitelist_mask = df[cols["code"]].apply(rda_norm_code).isin(RDA_WHITELIST_CODES) if enable_whitelist_transfer else pd.Series(False, index=df.index)
    df_whitelist = df.loc[whitelist_mask].copy()
    df_reduction_input = df.loc[~whitelist_mask].copy()

    if df_reduction_input.empty:
        df_out = df_reduction_input.copy()
        df_out["Temps retiré"] = pd.Series(dtype="int64")
        df_out["Temps ajouté"] = pd.Series(dtype="int64")
        return {
            "df_out": df_out,
            "df_main_raw": df_reduction_input,
            "df_whitelist": df_whitelist,
            "allocation_recap": pd.DataFrame(),
            "guard_summary": {"sum_delta": 0, "bad_receivers": 0, "bad_givers": 0, "bad_duration_rows": 0},
            "bad_receivers": pd.DataFrame(),
            "bad_givers": pd.DataFrame(),
            "bad_duration": pd.DataFrame(),
        }

    if not apply_adjustment:
        df_out = df_reduction_input.copy()
        df_out["Temps retiré"] = 0
        df_out["Temps ajouté"] = 0
    else:
        work = df_reduction_input.copy()
        original_cols = work.columns.tolist()
        work["Temps retiré"] = 0
        work["Temps ajouté"] = 0
        work["_start_min"] = work[cols["start"]].apply(rda_to_minutes)
        work["_end_min"] = work[cols["end"]].apply(rda_to_minutes)
        work["_start_abs0"] = work["_start_min"].astype("Float64")
        work["_end_abs0"] = work.apply(lambda row: rda_end_abs_min(row["_end_min"], row["_start_min"]), axis=1).astype("Float64")
        work["_dur0"] = work.apply(lambda row: rda_duration_from_minutes(row["_start_min"], row["_end_min"]), axis=1).astype("Int64")
        work["_jour"] = pd.to_datetime(work[cols["date"]], dayfirst=True, errors="coerce").dt.date

        def process_block(block: pd.DataFrame) -> pd.DataFrame:
            block = block.sort_values(["_start_abs0", "_end_abs0"]).copy().reset_index(drop=False)
            idx_colname = "index"
            valid = block["_start_abs0"].notna() & block["_end_abs0"].notna() & block["_dur0"].notna()
            valid_idx = block.index[valid].tolist()
            if not valid_idx:
                return block.set_index(idx_colname).sort_index()

            dur0 = pd.to_numeric(block["_dur0"], errors="coerce").fillna(0).astype(int)
            day_start = int(np.nanmin(block.loc[valid_idx, "_start_abs0"].astype(float)))
            day_end = int(np.nanmax(block.loc[valid_idx, "_end_abs0"].astype(float)))
            day_span = day_end - day_start
            order = block.loc[valid_idx].sort_values(["_start_abs0", "_end_abs0"]).index.tolist()
            gaps0 = [
                max(0, int(block.loc[order[pos + 1], "_start_abs0"]) - int(block.loc[order[pos], "_end_abs0"]))
                for pos in range(len(order) - 1)
            ]
            total_gap0 = sum(gaps0)
            dur_new = block["_dur0"].copy()
            target_idx = [idx for idx in valid_idx if rda_norm_code(block.loc[idx, cols["code"]]) in RDA_ALLOWED_TARGET_CODES]

            def pick_nearest_target(source_idx):
                if not target_idx:
                    return None
                source_start = float(block.loc[source_idx, "_start_abs0"])
                return min(
                    target_idx,
                    key=lambda idx: (
                        abs(float(block.loc[idx, "_start_abs0"]) - source_start),
                        0 if float(block.loc[idx, "_start_abs0"]) >= source_start else 1,
                        float(block.loc[idx, "_start_abs0"]),
                    ),
                )

            for idx in valid_idx:
                if rda_norm_code(block.loc[idx, cols["code"]]) != RDA_CODE_61010:
                    continue
                duration = dur_new.loc[idx]
                if pd.isna(duration) or int(duration) <= 15:
                    continue
                target = pick_nearest_target(idx)
                if target is None:
                    continue
                surplus = int(duration) - 15
                dur_new.loc[idx] = 15
                dur_new.loc[target] = int(dur_new.loc[target]) + surplus

            dn = pd.to_numeric(dur_new, errors="coerce").fillna(0).astype(int)
            delta = dn - dur0
            block["Temps ajouté"] = np.where(delta > 0, delta, 0).astype(int)
            block["Temps retiré"] = np.where(delta < 0, -delta, 0).astype(int)

            work_new = int(dn.loc[valid_idx].sum())
            slack = day_span - work_new
            block["_start_abs"] = pd.NA
            block["_end_abs"] = pd.NA
            if slack >= 0 and len(order) > 1:
                scaled = [gap * slack / total_gap0 for gap in gaps0] if total_gap0 > 0 else [slack / (len(order) - 1)] * (len(order) - 1)
                gaps_int = [int(np.floor(value)) for value in scaled]
                for pos in range(int(slack - sum(gaps_int))):
                    gaps_int[pos % (len(order) - 1)] += 1
                cur = day_start
                for pos, idx in enumerate(order):
                    start_abs = cur
                    end_abs = start_abs + int(dn.loc[idx])
                    block.loc[idx, "_start_abs"] = start_abs
                    block.loc[idx, "_end_abs"] = end_abs
                    if pos < len(order) - 1:
                        cur = end_abs + gaps_int[pos]
                drift = int(block.loc[order[-1], "_end_abs"]) - day_end
                if drift:
                    block.loc[order[-1], "_end_abs"] = int(block.loc[order[-1], "_end_abs"]) - drift
            else:
                prev_end = None
                for idx in order:
                    start_abs = int(block.loc[idx, "_start_abs0"])
                    end_abs = start_abs + int(dn.loc[idx])
                    if prev_end is not None and start_abs < prev_end:
                        shift = prev_end - start_abs
                        start_abs += shift
                        end_abs += shift
                    block.loc[idx, "_start_abs"] = start_abs
                    block.loc[idx, "_end_abs"] = end_abs
                    prev_end = end_abs

            for idx in valid_idx:
                block.loc[idx, cols["start"]] = rda_minutes_to_hhmmss(block.loc[idx, "_start_abs"])
                block.loc[idx, cols["end"]] = rda_minutes_to_hhmmss(block.loc[idx, "_end_abs"])
            block[cols["duration"]] = pd.to_numeric(dur_new, errors="coerce").astype("Int64")
            return block.set_index(idx_colname).sort_index()

        processed = [
            process_block(group)
            for _, group in work.groupby(["_jour", cols["collab"]], dropna=False, sort=False)
        ]
        full = pd.concat(processed).sort_index()
        cleanup_cols = ["_start_min", "_end_min", "_start_abs0", "_end_abs0", "_dur0", "_jour", "_start_abs", "_end_abs"]
        full = full.drop(columns=[col for col in cleanup_cols if col in full.columns])
        final_cols = original_cols + [col for col in ["Temps retiré", "Temps ajouté"] if col not in original_cols]
        df_out = full.reindex(columns=final_cols)

    dur_in = pd.to_numeric(df_reduction_input.apply(lambda row: rda_duration_from_minutes(rda_to_minutes(row[cols["start"]]), rda_to_minutes(row[cols["end"]])), axis=1), errors="coerce").fillna(0).astype(int)
    dur_out = pd.to_numeric(df_out[cols["duration"]], errors="coerce").fillna(0).astype(int)
    delta = dur_out - dur_in
    codes = df_out[cols["code"]].apply(rda_norm_code)
    bad_receivers = df_out[(delta > 0) & (~codes.isin(RDA_ALLOWED_TARGET_CODES))].copy()
    bad_givers = df_out[(delta < 0) & (codes != RDA_CODE_61010)].copy()

    duration_check = df_out.copy()
    duration_check["_duration_from_times"] = duration_check.apply(
        lambda row: rda_duration_from_minutes(rda_to_minutes(row[cols["start"]]), rda_to_minutes(row[cols["end"]])),
        axis=1,
    )
    duration_check["_duration_col"] = pd.to_numeric(duration_check[cols["duration"]], errors="coerce")
    bad_duration = duration_check[
        duration_check["_duration_col"].fillna(-999999).astype(int)
        != duration_check["_duration_from_times"].fillna(-999999).astype(int)
    ].copy()
    allocation_recap = (
        df_out.assign(_code=df_out[cols["code"]].apply(rda_norm_code))
        .groupby("_code")[["Temps ajouté", "Temps retiré"]]
        .sum()
        .query("`Temps ajouté` > 0 or `Temps retiré` > 0")
        .sort_values(["Temps ajouté", "Temps retiré"], ascending=False)
        .reset_index()
    )
    return {
        "df_out": df_out,
        "df_main_raw": df_reduction_input,
        "df_whitelist": df_whitelist,
        "allocation_recap": allocation_recap,
        "guard_summary": {
            "sum_delta": int(delta.sum()),
            "bad_receivers": len(bad_receivers),
            "bad_givers": len(bad_givers),
            "bad_duration_rows": len(bad_duration),
        },
        "bad_receivers": bad_receivers,
        "bad_givers": bad_givers,
        "bad_duration": bad_duration,
    }


def rda_to_nexus_csv(df_src: pd.DataFrame, cols: dict, oe_value: str) -> pd.DataFrame:
    out = pd.DataFrame(
        {
            "Datum": df_src[cols["date"]].apply(rda_format_date),
            "Von": df_src[cols["start"]].apply(rda_format_time),
            "Bis": df_src[cols["end"]].apply(rda_format_time),
            "Leistungscode": df_src[cols["code"]].apply(rda_norm_code).astype(str),
            "Dauer_verrechnet": pd.to_numeric(df_src[cols["duration"]], errors="coerce").fillna(0).astype(int),
            "OE": oe_value,
            "KD-Nr": pd.to_numeric(df_src[cols["client"]], errors="coerce").fillna(0).astype(int),
            "Klient": 0,
            "Einsatzgrund": pd.to_numeric(df_src[cols["client"]], errors="coerce").fillna(0).apply(lambda value: 0 if int(value) == 0 else 2).astype(int),
            "Mitarbeiter-ID": pd.to_numeric(df_src[cols["collab"]], errors="coerce").fillna(0).astype(int),
        }
    )
    return out


def rda_write_batch(batch_path: Path, csv_name: str, oe_value: str, map_rel: str, nexus_client_dir: str = "..\\nx-spi-client") -> None:
    batch_content = (
        "@echo off\n"
        "chcp 65001\n"
        f"\"{nexus_client_dir}\\Asebis.Client.StarterCommand.exe\" "
        f"/u=nexus /p=fAvNCDnW3E /t=ImportLeistungen_CSV /o={oe_value} "
        f"/f=\"{csv_name}\" /map=\"{map_rel}\" /v\n"
        "Pause\n"
    )
    batch_path.write_text(batch_content, encoding="utf-8")


def rda_export_source_outputs(df_out: pd.DataFrame, cols: dict, source_root: Path, oe_value: str) -> dict:
    source_root.mkdir(parents=True, exist_ok=True)
    df_out.to_excel(source_root / "RDA_61010_adjusted_no_overlap.xlsx", index=False)

    codes = sorted([code for code in df_out[cols["code"]].apply(rda_norm_code).dropna().astype(str).unique().tolist() if code])
    pd.DataFrame({"Code_ext": codes, "Leistungstarif_nummer": codes}).to_csv(source_root / "HAS_map_main.csv", index=False, sep=";")

    duration_summary = df_out.groupby(cols["collab"])[cols["duration"]].sum().reset_index()
    duration_summary.columns = ["Collaborateur_ID", "Sum_Duree"]
    duration_summary.to_csv(source_root / "RDA_duree_check.csv", index=False, sep=";")

    runnable_batches = []
    folder_all = source_root / "01_All_Collabs_One_CSV"
    folder_all.mkdir(parents=True, exist_ok=True)
    all_total = int(pd.to_numeric(df_out[cols["duration"]], errors="coerce").fillna(0).sum())
    all_csv_name = f"RDA_AllCollabs+{all_total}.csv"
    rda_to_nexus_csv(df_out, cols, oe_value).to_csv(folder_all / all_csv_name, index=False, sep=";")
    all_batch = folder_all / "RDA_AllCollabs_batch.bat"
    rda_write_batch(all_batch, all_csv_name, oe_value, "..\\HAS_map_main.csv")
    runnable_batches.append(all_batch)

    folder_61010 = source_root / "02_Collabs_With_61010_One_CSV"
    folder_61010.mkdir(parents=True, exist_ok=True)
    mask_61010 = df_out[cols["code"]].apply(rda_norm_code).eq(RDA_CODE_61010)
    collabs_with_61010 = df_out.loc[mask_61010, cols["collab"]].unique().tolist()
    batch_61010 = None
    if collabs_with_61010:
        df_61010 = df_out[df_out[cols["collab"]].isin(collabs_with_61010)].copy()
        total_61010 = int(pd.to_numeric(df_61010[cols["duration"]], errors="coerce").fillna(0).sum())
        csv_61010_name = f"RDA_CollabsWith61010+{total_61010}.csv"
        rda_to_nexus_csv(df_61010, cols, oe_value).to_csv(folder_61010 / csv_61010_name, index=False, sep=";")
        batch_61010 = folder_61010 / "RDA_CollabsWith61010_batch.bat"
        rda_write_batch(batch_61010, csv_61010_name, oe_value, "..\\HAS_map_main.csv")
        runnable_batches.append(batch_61010)

    folder_per = source_root / "03_Per_Collab_Separate"
    folder_per.mkdir(parents=True, exist_ok=True)
    collab_name_col = next((col for col in ["Collaborateur", "Nom Collaborateur"] if col in df_out.columns and col != cols["collab"]), None)
    for collab_id, group in df_out.groupby(cols["collab"]):
        total = int(pd.to_numeric(group[cols["duration"]], errors="coerce").fillna(0).sum())
        safe_id = rda_safe_folder_name(collab_id)
        if collab_name_col and not group[collab_name_col].empty:
            identifier = f"{safe_id}-{rda_safe_folder_name(group[collab_name_col].iloc[0])}"
        else:
            identifier = safe_id
        collab_folder = folder_per / f"RDA-{identifier}+{total}"
        collab_folder.mkdir(parents=True, exist_ok=True)
        csv_name = f"{identifier}+{total}.csv"
        rda_to_nexus_csv(group, cols, oe_value).to_csv(collab_folder / csv_name, index=False, sep=";")
        rda_write_batch(collab_folder / f"{identifier}_batch.bat", csv_name, oe_value, "..\\..\\HAS_map_main.csv")

    return {
        "source_root": source_root,
        "runnable_batches": runnable_batches,
        "batch_61010": batch_61010,
        "collabs_with_61010": collabs_with_61010,
    }


def rda_norm_col(value) -> str:
    import unicodedata

    text = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def rda_numeric_id_series(series: pd.Series) -> pd.Series:
    extracted = series.astype(str).str.extract(r"(\d+)")[0]
    return pd.to_numeric(extracted, errors="coerce").astype("Int64")


def rda_auto_mapping_col(df_map: pd.DataFrame, explicit_name: str, role: str, side: str) -> str:
    if explicit_name and explicit_name.strip():
        lookup = {rda_norm_col(col): col for col in df_map.columns}
        wanted = rda_norm_col(explicit_name)
        if wanted not in lookup:
            raise ValueError(f"Configured mapping column '{explicit_name}' not found. Available: {list(df_map.columns)}")
        return lookup[wanted]

    role_tokens = {
        "client": ["client", "klient", "kd", "customer"],
        "collab": ["collab", "collaborateur", "collaborator", "mitarbeiter", "employee"],
    }[role]
    side_tokens = ["sa", "target", "to", "new", "101"] if side == "target" else ["sarl", "source", "from", "old", "201"]
    scored = []
    for col in df_map.columns:
        normalized = rda_norm_col(col)
        if not any(token in normalized for token in role_tokens):
            continue
        score = sum(token in normalized for token in side_tokens)
        if "101" in normalized and side == "target":
            score += 2
        if "201" in normalized and side == "source":
            score += 2
        id_tokens = [" no ", " num ", " numero ", " number ", " nr ", " id "]
        padded = f" {normalized} "
        if any(token in padded for token in id_tokens) or normalized.startswith(("no ", "num ", "id ")):
            score += 4
        if any(token in padded for token in [" name ", " nom ", " display ", " label ", " libelle "]):
            score -= 3
        scored.append((score, col, normalized))
    if not scored:
        raise ValueError(f"Could not auto-detect {side} {role} mapping column. Available: {list(df_map.columns)}")
    scored.sort(key=lambda item: (-item[0], item[2]))
    best_score = scored[0][0]
    best = [col for score, col, _ in scored if score == best_score]
    if len(best) > 1:
        raise ValueError(f"Ambiguous {side} {role} mapping columns: {best}. Use explicit overrides.")
    return best[0]


def rda_export_whitelist_outputs(df_whitelist: pd.DataFrame, cols: dict, sa_root: Path, mapping_source, mapping_filename: str | None, overrides: dict) -> dict:
    transfer_folder = sa_root / "02_Whitelisted_Ready_For_101"
    transfer_folder.mkdir(parents=True, exist_ok=True)
    if df_whitelist.empty:
        return {
            "sa_root": sa_root,
            "transfer_folder": transfer_folder,
            "whitelist_transfer_csv_df": pd.DataFrame(),
            "df_whitelist_mapped": df_whitelist.copy(),
            "mapping_summary": pd.DataFrame([["Whitelisted rows", 0]], columns=["Metric", "Value"]),
            "unmapped_clients": pd.DataFrame(),
            "unmapped_collabs": pd.DataFrame(),
            "runnable_batches": [],
        }
    if mapping_source is None:
        raise ValueError("A mapping workbook is required when whitelist rows exist for the SA 101 transfer.")

    mapping_bytes = BytesIO(mapping_source.getvalue()) if hasattr(mapping_source, "getvalue") else mapping_source
    xls = pd.ExcelFile(mapping_bytes)
    clients_sheet = next((sheet for sheet in xls.sheet_names if "client" in rda_norm_col(sheet)), None)
    collabs_sheet = next((sheet for sheet in xls.sheet_names if "collab" in rda_norm_col(sheet) or "collabor" in rda_norm_col(sheet)), None)
    if not clients_sheet or not collabs_sheet:
        raise ValueError(f"Could not find both clients and collabs sheets in {mapping_filename}. Found: {xls.sheet_names}")
    clients_map_df = pd.read_excel(xls, sheet_name=clients_sheet)
    collabs_map_df = pd.read_excel(xls, sheet_name=collabs_sheet)
    src_client_col = rda_auto_mapping_col(clients_map_df, overrides.get("source_client", ""), "client", "source")
    tgt_client_col = rda_auto_mapping_col(clients_map_df, overrides.get("target_client", ""), "client", "target")
    src_collab_col = rda_auto_mapping_col(collabs_map_df, overrides.get("source_collab", ""), "collab", "source")
    tgt_collab_col = rda_auto_mapping_col(collabs_map_df, overrides.get("target_collab", ""), "collab", "target")
    client_map = {
        int(k): int(v)
        for k, v in zip(rda_numeric_id_series(clients_map_df[src_client_col]), rda_numeric_id_series(clients_map_df[tgt_client_col]))
        if pd.notna(k) and pd.notna(v)
    }
    collab_map = {
        int(k): int(v)
        for k, v in zip(rda_numeric_id_series(collabs_map_df[src_collab_col]), rda_numeric_id_series(collabs_map_df[tgt_collab_col]))
        if pd.notna(k) and pd.notna(v)
    }
    mapped = df_whitelist.copy()
    mapped["orig_client_no"] = rda_numeric_id_series(mapped[cols["client"]])
    mapped["orig_collab_no"] = rda_numeric_id_series(mapped[cols["collab"]])
    mapped["KD-Nr"] = mapped["orig_client_no"].map(client_map).fillna(0).astype(int)
    mapped["Mitarbeiter-ID"] = mapped["orig_collab_no"].map(collab_map).fillna(0).astype(int)
    mapped["Einsatzgrund"] = mapped["KD-Nr"].apply(lambda value: 2 if int(value) != 0 else 0).astype(int)
    transfer_csv = pd.DataFrame(
        {
            "Datum": mapped[cols["date"]].apply(rda_format_date),
            "Von": mapped[cols["start"]].apply(rda_format_time),
            "Bis": mapped[cols["end"]].apply(rda_format_time),
            "Leistungscode": mapped[cols["code"]].apply(rda_norm_code).astype(str),
            "Dauer_verrechnet": pd.to_numeric(mapped[cols["duration"]], errors="coerce").fillna(0).astype(int),
            "OE": RDA_TRANSFER_TARGET_OE,
            "KD-Nr": mapped["KD-Nr"].fillna(0).astype(int),
            "Klient": 0,
            "Einsatzgrund": mapped["Einsatzgrund"].fillna(0).astype(int),
            "Mitarbeiter-ID": mapped["Mitarbeiter-ID"].fillna(0).astype(int),
        }
    )
    codes = sorted([code for code in transfer_csv["Leistungscode"].dropna().astype(str).unique().tolist() if code])
    pd.DataFrame({"Code_ext": codes, "Leistungstarif_nummer": codes}).to_csv(sa_root / "HAS_map.csv", index=False, sep=";")
    total = int(transfer_csv["Dauer_verrechnet"].fillna(0).sum())
    csv_name = f"RDA_Whitelisted_Ready_For_101+{total}.csv"
    transfer_csv.to_csv(transfer_folder / csv_name, index=False, sep=";", encoding="utf-8")
    batch_path = transfer_folder / "RDA_Whitelisted_Ready_For_101_batch.bat"
    rda_write_batch(batch_path, csv_name, RDA_TRANSFER_TARGET_OE, "..\\HAS_map.csv")
    unmapped_clients_mask = mapped["orig_client_no"].notna() & (mapped["orig_client_no"] != 0) & (mapped["KD-Nr"] == 0)
    unmapped_collabs_mask = mapped["orig_collab_no"].notna() & (mapped["orig_collab_no"] != 0) & (mapped["Mitarbeiter-ID"] == 0)
    summary = pd.DataFrame(
        [
            ["Whitelisted rows", len(mapped)],
            ["Total whitelist minutes", total],
            ["Mapped client rows", int((mapped["KD-Nr"] != 0).sum())],
            ["Unmapped client rows", int(unmapped_clients_mask.sum())],
            ["Mapped collaborator rows", int((mapped["Mitarbeiter-ID"] != 0).sum())],
            ["Unmapped collaborator rows", int(unmapped_collabs_mask.sum())],
            ["Clients sheet", clients_sheet],
            ["Collabs sheet", collabs_sheet],
        ],
        columns=["Metric", "Value"],
    )
    with pd.ExcelWriter(transfer_folder / "RDA_Whitelisted_Ready_For_101_QA.xlsx", engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Summary")
        transfer_csv.to_excel(writer, index=False, sheet_name="Import_CSV")
        mapped.to_excel(writer, index=False, sheet_name="Mapped_Source_Rows")
    return {
        "sa_root": sa_root,
        "transfer_folder": transfer_folder,
        "whitelist_transfer_csv_df": transfer_csv,
        "df_whitelist_mapped": mapped,
        "mapping_summary": summary,
        "unmapped_clients": mapped.loc[unmapped_clients_mask].copy(),
        "unmapped_collabs": mapped.loc[unmapped_collabs_mask].copy(),
        "runnable_batches": [batch_path],
    }


def rda_audit_generated_csvs(root: Path) -> dict:
    folder_rows = []
    csv_rows = []
    bad_parts = []
    for folder in ["01_All_Collabs_One_CSV", "02_Collabs_With_61010_One_CSV", "03_Per_Collab_Separate", "02_Whitelisted_Ready_For_101"]:
        folder_path = root / folder
        csv_files = sorted(folder_path.rglob("*.csv")) if folder_path.exists() else []
        folder_count = folder_duration = folder_time = folder_bad = folder_row_count = 0
        for csv_path in csv_files:
            try:
                csv_df = pd.read_csv(csv_path, sep=";")
            except Exception:
                continue
            if not {"Von", "Bis", "Dauer_verrechnet"}.issubset(csv_df.columns):
                continue
            start = csv_df["Von"].apply(rda_to_minutes)
            end = csv_df["Bis"].apply(rda_to_minutes)
            calc = [rda_duration_from_minutes(s, e) for s, e in zip(start, end)]
            duration = pd.to_numeric(csv_df["Dauer_verrechnet"], errors="coerce")
            delta = duration - pd.Series(calc, index=csv_df.index)
            bad = csv_df[delta.fillna(0) != 0].copy()
            if not bad.empty:
                bad["_file"] = str(csv_path.relative_to(root))
                bad_parts.append(bad)
            row = {
                "Folder": folder,
                "CSV file": str(csv_path.relative_to(root)),
                "Rows": len(csv_df),
                "Sum Dauer_verrechnet": int(duration.fillna(0).sum()),
                "Sum calculated Von/Bis": int(pd.Series(calc).fillna(0).sum()),
                "Difference": int(duration.fillna(0).sum() - pd.Series(calc).fillna(0).sum()),
                "Mismatch rows": len(bad),
            }
            csv_rows.append(row)
            folder_count += 1
            folder_row_count += row["Rows"]
            folder_duration += row["Sum Dauer_verrechnet"]
            folder_time += row["Sum calculated Von/Bis"]
            folder_bad += row["Mismatch rows"]
        folder_rows.append(
            {
                "Folder": folder,
                "Folder exists": folder_path.exists(),
                "CSV files": folder_count,
                "Rows": folder_row_count,
                "Sum Dauer_verrechnet": folder_duration,
                "Sum calculated Von/Bis": folder_time,
                "Difference": folder_duration - folder_time,
                "Mismatch rows": folder_bad,
            }
        )
    return {
        "folder_summary": pd.DataFrame(folder_rows),
        "csv_summary": pd.DataFrame(csv_rows),
        "bad_rows": pd.concat(bad_parts, ignore_index=True) if bad_parts else pd.DataFrame(),
    }


def rda_make_zip(root: Path) -> Path:
    zip_path = root.with_suffix(".zip")
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in root.rglob("*"):
            if path.is_file():
                archive.write(path, path.relative_to(root.parent))
    return zip_path


def rda_validate_nexus_path(path_text: str) -> Path | None:
    if not path_text.strip():
        return None
    candidate = Path(path_text).expanduser()
    exe = candidate / "Asebis.Client.StarterCommand.exe"
    return exe if exe.exists() else None


def rda_find_nexus_client_dirs(search_root_text: str, max_results: int = 25) -> list[Path]:
    if not search_root_text.strip():
        return []
    search_root = Path(search_root_text).expanduser()
    if not search_root.exists() or not search_root.is_dir():
        raise FileNotFoundError(f"Search folder does not exist: {search_root}")

    matches = []
    for exe_path in search_root.rglob("Asebis.Client.StarterCommand.exe"):
        if exe_path.is_file():
            matches.append(exe_path.parent)
            if len(matches) >= max_results:
                break
    return matches


def rda_run_batch(batch_path: Path, nexus_client_dir: Path) -> dict:
    if not batch_path.exists():
        raise FileNotFoundError(f"Batch file not found: {batch_path}")
    nexus_exe = nexus_client_dir / "Asebis.Client.StarterCommand.exe"
    if not nexus_exe.exists():
        raise FileNotFoundError(f"Nexus client executable not found in: {nexus_client_dir}")
    original_text = batch_path.read_text(encoding="utf-8")
    run_text = re.sub(
        r'"[^"]*Asebis\.Client\.StarterCommand\.exe"',
        f'"{nexus_exe}"',
        original_text,
        count=1,
    )
    run_text = re.sub(r"(?im)^\s*Pause\s*$", "", run_text)
    run_batch_path = batch_path.with_name(f"__streamlit_run_{batch_path.name}")
    try:
        run_batch_path.write_text(run_text, encoding="utf-8")
        result = subprocess.run(
            ["cmd.exe", "/c", str(run_batch_path)],
            cwd=str(batch_path.parent),
            capture_output=True,
            text=True,
            timeout=3600,
            env=os.environ.copy(),
        )
        return {
            "batch": str(batch_path),
            "returncode": result.returncode,
            "stdout": result.stdout,
            "stderr": result.stderr,
        }
    finally:
        if run_batch_path.exists():
            run_batch_path.unlink()


def rda_batch_preview_rows(batch_paths: list[Path], nexus_exe: Path | None) -> pd.DataFrame:
    rows = []
    for batch_path in batch_paths:
        text = batch_path.read_text(encoding="utf-8") if batch_path.exists() else ""
        csv_match = re.search(r'/f="([^"]+)"', text)
        map_match = re.search(r'/map="([^"]+)"', text)
        csv_name = csv_match.group(1) if csv_match else ""
        map_rel = map_match.group(1) if map_match else ""
        csv_path = (batch_path.parent / csv_name).resolve() if csv_name else None
        map_path = (batch_path.parent / map_rel).resolve() if map_rel else None
        rows.append(
            {
                "Batch": str(batch_path),
                "Executable used by app": str(nexus_exe) if nexus_exe else "",
                "CSV": str(csv_path) if csv_path else "",
                "CSV exists": bool(csv_path and csv_path.exists()),
                "HAS map": str(map_path) if map_path else "",
                "HAS map exists": bool(map_path and map_path.exists()),
            }
        )
    return pd.DataFrame(rows)


def rda_blank_checks(df: pd.DataFrame, cols: dict) -> dict:
    blank_clients = df[df[cols["client"]].isna() | df[cols["client"]].astype(str).str.strip().isin(["", "0", "0.0"])]
    blank_collabs = df[df[cols["collab"]].isna() | df[cols["collab"]].astype(str).str.strip().isin(["", "0", "0.0"])]
    return {"blank_clients": blank_clients.copy(), "blank_collabs": blank_collabs.copy()}


def rda_check_overlaps(df: pd.DataFrame, cols: dict) -> pd.DataFrame:
    check = df.copy()
    check["_start_min_chk"] = check[cols["start"]].apply(rda_to_minutes)
    check["_end_min_chk"] = check[cols["end"]].apply(rda_to_minutes)
    check["_jour_chk"] = pd.to_datetime(check[cols["date"]], dayfirst=True, errors="coerce").dt.date

    overlaps = []
    for (jour, collab), group in check.groupby(["_jour_chk", cols["collab"]], dropna=False, sort=False):
        block = group.sort_values(["_start_min_chk", "_end_min_chk"]).copy().reset_index()
        for idx in range(len(block) - 1):
            start_1 = block.loc[idx, "_start_min_chk"]
            end_1 = block.loc[idx, "_end_min_chk"]
            start_2 = block.loc[idx + 1, "_start_min_chk"]
            end_2 = block.loc[idx + 1, "_end_min_chk"]
            if pd.isna(start_1) or pd.isna(end_1) or pd.isna(start_2) or pd.isna(end_2):
                continue
            if end_1 > start_2:
                overlaps.append(
                    {
                        "Date": jour,
                        "Collaborateur": collab,
                        "Row1_index_orig": block.loc[idx, "index"],
                        "Row1_Début": block.loc[idx, cols["start"]],
                        "Row1_Fin": block.loc[idx, cols["end"]],
                        "Row1_Code": block.loc[idx, cols["code"]],
                        "Row2_index_orig": block.loc[idx + 1, "index"],
                        "Row2_Début": block.loc[idx + 1, cols["start"]],
                        "Row2_Fin": block.loc[idx + 1, cols["end"]],
                        "Row2_Code": block.loc[idx + 1, cols["code"]],
                        "Overlap_minutes": end_1 - start_2,
                    }
                )

    return pd.DataFrame(
        overlaps,
        columns=[
            "Date",
            "Collaborateur",
            "Row1_index_orig",
            "Row1_Début",
            "Row1_Fin",
            "Row1_Code",
            "Row2_index_orig",
            "Row2_Début",
            "Row2_Fin",
            "Row2_Code",
            "Overlap_minutes",
        ],
    )


def rda_process(
    raw_source,
    raw_filename: str,
    mapping_source,
    mapping_filename: str | None,
    source_uo: str,
    output_name: str,
    mapping_overrides: dict,
    apply_15min_adjustment: bool = True,
    enable_whitelist_transfer: bool = True,
) -> dict:
    raw_df = rda_read_input(BytesIO(raw_source.getvalue()) if hasattr(raw_source, "getvalue") else raw_source, raw_filename)
    cols = rda_detect_columns(raw_df)
    raw_bad_duration, raw_duration_summary = rda_duration_sanity(raw_df, cols)
    normalized = rda_normalize_duration_from_times(raw_df, cols)
    adjusted = rda_adjust_61010(
        normalized,
        cols,
        apply_adjustment=apply_15min_adjustment,
        enable_whitelist_transfer=enable_whitelist_transfer,
    )

    date_series = pd.to_datetime(normalized[cols["date"]], dayfirst=True, errors="coerce").dropna()
    if date_series.empty:
        raise ValueError("Could not determine output month/year from RDA date column.")
    first_valid = date_series.iloc[0]
    safe_output_name = rda_safe_folder_name(output_name) if output_name.strip() else f"RDA-{first_valid.month:02d}{first_valid.year}"
    root = Path.cwd() / RDA_OUTPUT_FOLDER / safe_output_name
    if root.exists():
        root = root.with_name(f"{root.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    source_root = root / f"Source_{source_uo.replace(' ', '_')}"
    sa_root = root / "SA_101"
    source_export = rda_export_source_outputs(adjusted["df_out"], cols, source_root, RDA_OE_MAP[source_uo])
    if enable_whitelist_transfer:
        whitelist_export = rda_export_whitelist_outputs(adjusted["df_whitelist"], cols, sa_root, mapping_source, mapping_filename, mapping_overrides)
    else:
        sa_root.mkdir(parents=True, exist_ok=True)
        whitelist_export = {
            "sa_root": sa_root,
            "transfer_folder": sa_root,
            "whitelist_transfer_csv_df": pd.DataFrame(),
            "df_whitelist_mapped": pd.DataFrame(),
            "mapping_summary": pd.DataFrame([["Whitelist transfer", "Disabled"]], columns=["Metric", "Value"]),
            "unmapped_clients": pd.DataFrame(),
            "unmapped_collabs": pd.DataFrame(),
            "runnable_batches": [],
        }
    source_audit = rda_audit_generated_csvs(source_root)
    sa_audit = rda_audit_generated_csvs(sa_root)
    blank_checks = rda_blank_checks(normalized, cols)
    overlaps_df = rda_check_overlaps(adjusted["df_out"], cols)

    source_total = int(pd.to_numeric(adjusted["df_out"][cols["duration"]], errors="coerce").fillna(0).sum())
    whitelist_total = int(whitelist_export["whitelist_transfer_csv_df"]["Dauer_verrechnet"].fillna(0).sum()) if not whitelist_export["whitelist_transfer_csv_df"].empty else 0
    raw_total = int(pd.to_numeric(normalized[cols["duration"]], errors="coerce").fillna(0).sum())
    split_total_ok = source_total + whitelist_total == raw_total

    qa_checks = [
        {"Check": "Le delta d'allocation est égal à zéro", "Passed": adjusted["guard_summary"]["sum_delta"] == 0, "Detail": adjusted["guard_summary"]["sum_delta"]},
        {"Check": "Aucun code non autorisé n'a reçu de minutes", "Passed": adjusted["guard_summary"]["bad_receivers"] == 0, "Detail": adjusted["guard_summary"]["bad_receivers"]},
        {"Check": "Aucun code hors 61010 n'a donné de minutes", "Passed": adjusted["guard_summary"]["bad_givers"] == 0, "Detail": adjusted["guard_summary"]["bad_givers"]},
        {"Check": "La Durée ajustée correspond à Fin - Début", "Passed": adjusted["guard_summary"]["bad_duration_rows"] == 0, "Detail": adjusted["guard_summary"]["bad_duration_rows"]},
        {"Check": "Les durées CSV source correspondent à Von/Bis", "Passed": int(source_audit["folder_summary"]["Mismatch rows"].sum()) == 0, "Detail": int(source_audit["folder_summary"]["Mismatch rows"].sum())},
        {"Check": "Les totaux des dossiers CSV source correspondent", "Passed": int(source_audit["folder_summary"]["Difference"].abs().sum()) == 0, "Detail": int(source_audit["folder_summary"]["Difference"].abs().sum())},
        {"Check": "Les durées CSV SA101 correspondent à Von/Bis", "Passed": int(sa_audit["folder_summary"]["Mismatch rows"].sum()) == 0, "Detail": int(sa_audit["folder_summary"]["Mismatch rows"].sum())},
        {"Check": "Les totaux des dossiers CSV SA101 correspondent", "Passed": int(sa_audit["folder_summary"]["Difference"].abs().sum()) == 0, "Detail": int(sa_audit["folder_summary"]["Difference"].abs().sum())},
        {"Check": "Les minutes source + SA101 correspondent au total brut", "Passed": split_total_ok, "Detail": source_total + whitelist_total - raw_total},
        {"Check": "Aucun client brut vide ou zéro", "Passed": len(blank_checks["blank_clients"]) == 0, "Detail": len(blank_checks["blank_clients"])},
        {"Check": "Aucun collaborateur brut vide ou zéro", "Passed": len(blank_checks["blank_collabs"]) == 0, "Detail": len(blank_checks["blank_collabs"])},
        {"Check": "Aucun client SA101 non mappé", "Passed": len(whitelist_export["unmapped_clients"]) == 0, "Detail": len(whitelist_export["unmapped_clients"])},
        {"Check": "Aucun collaborateur SA101 non mappé", "Passed": len(whitelist_export["unmapped_collabs"]) == 0, "Detail": len(whitelist_export["unmapped_collabs"])},
        {"Check": "Aucun chevauchement même jour / même collaborateur", "Passed": len(overlaps_df) == 0, "Detail": len(overlaps_df)},
    ]
    qa_checks_df = pd.DataFrame(qa_checks)
    qa_passed = bool(qa_checks_df["Passed"].all())
    zip_path = rda_make_zip(root)
    runnable_batches = source_export["runnable_batches"] + whitelist_export["runnable_batches"]

    return {
        "root": root,
        "source_root": source_root,
        "sa_root": sa_root,
        "zip_path": zip_path,
        "columns": cols,
        "raw_df": raw_df,
        "raw_bad_duration": raw_bad_duration,
        "raw_duration_summary": raw_duration_summary,
        "adjusted": adjusted,
        "source_audit": source_audit,
        "sa_audit": sa_audit,
        "blank_checks": blank_checks,
        "overlaps": overlaps_df,
        "whitelist_export": whitelist_export,
        "split_summary": {
            "raw_total": raw_total,
            "source_total": source_total,
            "whitelist_total": whitelist_total,
            "combined_export_total": source_total + whitelist_total,
            "split_total_ok": split_total_ok,
        },
        "qa_checks": qa_checks_df,
        "qa_passed": qa_passed,
        "runnable_batches": runnable_batches,
        "all_collabs": sorted([str(value) for value in normalized[cols["collab"]].dropna().unique()]),
        "all_codes": sorted([str(value) for value in normalized[cols["code"]].dropna().apply(rda_norm_code).unique()]),
        "options": {
            "apply_15min_adjustment": apply_15min_adjustment,
            "enable_whitelist_transfer": enable_whitelist_transfer,
        },
    }


def render_rda_dataframe(title: str, df: pd.DataFrame, empty_message: str = "Aucune ligne trouvée.") -> None:
    with st.expander(f"{title} ({len(df):,})", expanded=not df.empty):
        if df.empty:
            st.success(empty_message)
        else:
            st.dataframe(df, width="stretch", hide_index=True)


def render_rda_summary_table(title: str, df: pd.DataFrame) -> None:
    st.subheader(title)
    if df.empty:
        st.info("Aucune ligne.")
    else:
        st.dataframe(df, width="stretch", hide_index=True)


def render_rda_batch_controls(result: dict, nexus_client_dir_text: str) -> None:
    st.subheader("Exécuter les fichiers batch")
    nexus_exe = rda_validate_nexus_path(nexus_client_dir_text)
    if not result["qa_passed"]:
        st.warning("Les boutons batch sont désactivés car les contrôles qualité ne sont pas validés.")
    if nexus_exe is None:
        st.warning("Les boutons batch sont désactivés tant que le dossier du client Nexus n'est pas valide.")

    runnable = result["runnable_batches"]
    if not runnable:
        st.info("Aucun fichier batch de package n'a été généré.")
        return

    st.write("Aperçu de l'exécution batch")
    st.dataframe(rda_batch_preview_rows(runnable, nexus_exe), width="stretch", hide_index=True)

    run_enabled = result["qa_passed"] and nexus_exe is not None
    for batch_path in runnable:
        label = batch_path.parent.name
        cols = st.columns([3, 1])
        cols[0].code(str(batch_path), language="text")
        if cols[1].button(f"Exécuter {label}", disabled=not run_enabled, key=f"run_{batch_path}"):
            try:
                run_result = rda_run_batch(batch_path, nexus_exe.parent)
                st.session_state.setdefault("rda_batch_runs", []).append(run_result)
            except Exception as exc:
                st.exception(exc)

    if st.session_state.get("rda_batch_runs"):
        st.subheader("Journal d'exécution batch")
        for item in st.session_state["rda_batch_runs"]:
            status = "OK" if item["returncode"] == 0 else f"Exit {item['returncode']}"
            with st.expander(f"{status}: {Path(item['batch']).name}", expanded=item["returncode"] != 0):
                st.code(item["batch"], language="text")
                if item["stdout"]:
                    st.text_area("stdout", item["stdout"], height=180)
                if item["stderr"]:
                    st.text_area("stderr", item["stderr"], height=120)


def render_rda_results(result: dict, nexus_client_dir_text: str) -> None:
    st.subheader("Statut des contrôles qualité")
    status_cols = st.columns(5)
    status_cols[0].metric("QA", "Réussi" if result["qa_passed"] else "Échec")
    status_cols[1].metric("Minutes brutes", f"{result['split_summary']['raw_total']:,}")
    status_cols[2].metric("Minutes source", f"{result['split_summary']['source_total']:,}")
    status_cols[3].metric("SA101 minutes", f"{result['split_summary']['whitelist_total']:,}")
    status_cols[4].metric("Combiné", f"{result['split_summary']['combined_export_total']:,}")

    st.caption(f"Dossier de sortie : {result['root']}")
    st.write(
        {
            "Ajustement 15 minutes 61010": "activé" if result["options"]["apply_15min_adjustment"] else "désactivé",
            "Transfert whitelist SA 101": "activé" if result["options"]["enable_whitelist_transfer"] else "désactivé",
        }
    )
    if result["zip_path"].exists():
        render_download_for_path(result["zip_path"], "Télécharger le zip du package de transfert RDA")

    st.subheader("Liste des contrôles qualité")
    qa_display = result["qa_checks"].copy()
    qa_display["Statut"] = qa_display["Passed"].map({True: "Réussi", False: "Échec"})
    qa_display = qa_display.rename(columns={"Check": "Contrôle", "Detail": "Détail"})
    st.dataframe(qa_display[["Statut", "Contrôle", "Détail"]], width="stretch", hide_index=True)

    st.subheader("Contrôles stricts d'allocation")
    st.dataframe(pd.DataFrame([result["adjusted"]["guard_summary"]]), width="stretch", hide_index=True)
    render_rda_dataframe("Récapitulatif d'allocation", result["adjusted"]["allocation_recap"], "Aucune réallocation 61010 nécessaire.")
    render_rda_dataframe("Récepteurs hors codes autorisés", result["adjusted"]["bad_receivers"])
    render_rda_dataframe("Donneurs hors 61010", result["adjusted"]["bad_givers"])
    render_rda_dataframe("Lignes où la Durée ajustée diffère de Fin - Début", result["adjusted"]["bad_duration"])

    st.subheader("Contrôle des durées brutes")
    st.dataframe(pd.DataFrame([result["raw_duration_summary"]]), width="stretch", hide_index=True)
    render_rda_dataframe("Lignes brutes où Durée diffère de Début/Fin avant normalisation", result["raw_bad_duration"])

    st.subheader("Contrôle des chevauchements")
    render_rda_dataframe(
        "Chevauchements après ajustement / séparation source",
        result["overlaps"],
        "Aucun chevauchement détecté dans le planning d'un même collaborateur pour une journée donnée.",
    )

    st.subheader("Audits des CSV générés")
    render_rda_summary_table("Résumé des dossiers UO source", result["source_audit"]["folder_summary"])
    render_rda_summary_table("Résumé des CSV UO source", result["source_audit"]["csv_summary"])
    render_rda_dataframe("Écarts de durée dans les CSV générés UO source", result["source_audit"]["bad_rows"])
    render_rda_summary_table("Résumé des dossiers SA 101", result["sa_audit"]["folder_summary"])
    render_rda_summary_table("Résumé des CSV SA 101", result["sa_audit"]["csv_summary"])
    render_rda_dataframe("Écarts de durée dans les CSV générés SA 101", result["sa_audit"]["bad_rows"])

    st.subheader("Contrôles de mapping et valeurs vides")
    st.dataframe(result["whitelist_export"]["mapping_summary"], width="stretch", hide_index=True)
    render_rda_dataframe("Clients SA101 non mappés", result["whitelist_export"]["unmapped_clients"])
    render_rda_dataframe("Collaborateurs SA101 non mappés", result["whitelist_export"]["unmapped_collabs"])
    render_rda_dataframe("Clients bruts vides ou zéro", result["blank_checks"]["blank_clients"])
    render_rda_dataframe("Collaborateurs bruts vides ou zéro", result["blank_checks"]["blank_collabs"])

    with st.expander("Tous les collaborateurs et codes prestation", expanded=False):
        st.write("Collaborateurs")
        st.write(result["all_collabs"])
        st.write("Codes prestation")
        st.write(result["all_codes"])

    render_rda_batch_controls(result, nexus_client_dir_text)


def render_rda_task() -> None:
    st.title("Transferts RDA")
    st.caption("Crée les packages de transfert UO source et SA 101 avec contrôles qualité avant l'exécution batch.")

    pending_nexus_dir = st.session_state.pop("rda_pending_nexus_client_dir", None)
    if pending_nexus_dir is not None:
        st.session_state["rda_nexus_client_dir"] = pending_nexus_dir

    raw_file = st.file_uploader("Fichier RDA brut", type=["xlsx", "csv"], key="rda_raw_file")
    source_uo = st.selectbox("UO source", list(RDA_OE_MAP.keys()), index=0)
    option_cols = st.columns(2)
    apply_15min_adjustment = option_cols[0].checkbox("Appliquer l'ajustement 15 minutes 61010", value=True)
    enable_whitelist_transfer = option_cols[1].checkbox("Créer le transfert whitelist SA 101", value=True)
    output_cols = st.columns(2)
    output_name = output_cols[0].text_input("Nom du dossier d'export", value="")
    if "rda_nexus_client_dir" not in st.session_state:
        st.session_state["rda_nexus_client_dir"] = ""
    nexus_client_dir = output_cols[1].text_input("Dossier du client Nexus", key="rda_nexus_client_dir")
    nexus_exe = rda_validate_nexus_path(nexus_client_dir)
    if nexus_client_dir.strip():
        if nexus_exe:
            st.success(f"Client Nexus trouvé : {nexus_exe}")
        else:
            st.warning("Ce dossier ne contient pas Asebis.Client.StarterCommand.exe.")

    with st.expander("Rechercher le dossier du client Nexus", expanded=False):
        default_search_root = str(Path("C:/").resolve()) if Path("C:/").exists() else str(Path.cwd())
        search_cols = st.columns([3, 1])
        search_root = search_cols[0].text_input("Dossier de départ de la recherche", value=default_search_root, key="rda_nexus_search_root")
        if search_cols[1].button("Rechercher", key="rda_nexus_search_button"):
            try:
                st.session_state["rda_nexus_search_results"] = [str(path) for path in rda_find_nexus_client_dirs(search_root)]
            except Exception as exc:
                st.session_state["rda_nexus_search_results"] = []
                st.exception(exc)

        results = st.session_state.get("rda_nexus_search_results", [])
        if results:
            selected = st.selectbox("Dossiers trouvés", results, key="rda_nexus_found_select")
            if st.button("Utiliser le dossier sélectionné", key="rda_nexus_use_selected"):
                st.session_state["rda_pending_nexus_client_dir"] = selected
                st.rerun()
        elif "rda_nexus_search_results" in st.session_state:
            st.info("Aucun dossier client Nexus trouvé dans ce dossier de recherche.")

    mapping_file = None
    source_client = ""
    target_client = ""
    source_collab = ""
    target_collab = ""
    if enable_whitelist_transfer:
        st.subheader("Mapping SA 101")
        mapping_file = st.file_uploader("Classeur de mapping pour le transfert whitelist", type=["xlsx", "xls"], key="rda_mapping_file")
        with st.expander("Forcer les colonnes de mapping", expanded=False):
            override_cols = st.columns(4)
            source_client = override_cols[0].text_input("Col. client source")
            target_client = override_cols[1].text_input("Col. client cible")
            source_collab = override_cols[2].text_input("Col. collaborateur source")
            target_collab = override_cols[3].text_input("Col. collaborateur cible")
    else:
        st.info("Le transfert whitelist SA 101 est désactivé. Les lignes de prestation whitelist restent dans la sortie UO source et aucun classeur de mapping n'est requis.")

    if st.button("Générer le package de transfert RDA", type="primary", disabled=raw_file is None):
        try:
            result = rda_process(
                raw_source=raw_file,
                raw_filename=raw_file.name,
                mapping_source=mapping_file,
                mapping_filename=mapping_file.name if mapping_file else None,
                source_uo=source_uo,
                output_name=output_name,
                mapping_overrides={
                    "source_client": source_client,
                    "target_client": target_client,
                    "source_collab": source_collab,
                    "target_collab": target_collab,
                },
                apply_15min_adjustment=apply_15min_adjustment,
                enable_whitelist_transfer=enable_whitelist_transfer,
            )
            st.session_state["latest_rda_result"] = result
            st.session_state["rda_batch_runs"] = []
        except Exception as exc:
            st.exception(exc)
            return

    result = st.session_state.get("latest_rda_result")
    if result:
        render_rda_results(result, nexus_client_dir)


def ltr_unique_output_root(output_name: str) -> Path:
    safe_name = rda_safe_folder_name(output_name) if output_name.strip() else f"LTR_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    root = APP_ROOT / LTR_OUTPUT_FOLDER / safe_name
    if root.exists():
        root = root.with_name(f"{root.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    root.mkdir(parents=True, exist_ok=True)
    return root


def ltr_save_upload(uploaded_file, target_dir: Path) -> Path:
    target_dir.mkdir(parents=True, exist_ok=True)
    path = target_dir / uploaded_file.name
    path.write_bytes(uploaded_file.getvalue())
    return path


class LtrPytzCompat:
    @staticmethod
    def timezone(name: str):
        return ZoneInfo(name)


@st.cache_resource(show_spinner=False)
def ltr_load_notebook_functions(notebook_mtime: float) -> dict:
    _ = notebook_mtime
    notebook_path = LTR_NOTEBOOK_PATH
    if not notebook_path.exists():
        raise FileNotFoundError(f"LTR notebook not found: {notebook_path}")

    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    env = {
        "__builtins__": __builtins__,
        "os": os,
        "re": re,
        "np": np,
        "pd": pd,
        "pytz": LtrPytzCompat,
        "Path": Path,
        "datetime": datetime,
        "timedelta": timedelta,
        "dtime": dtime,
        "load_workbook": load_workbook,
        "Alignment": Alignment,
        "get_column_letter": get_column_letter,
        "SWISS_TZ": ZoneInfo("Europe/Zurich"),
        "MAX_SINGLE_INTERVAL_HOURS": 24,
        "SERVICE_RESET_GAP_HOURS": 8.0,
        "WEEKLY_LIMIT_HOURS": 50.0,
        "SPAN_LIMIT_HOURS": 14.0,
        "REST_NORMAL_HOURS": 11.0,
        "REST_ABSOLUTE_MIN_HOURS": 8.0,
        "REQUIRE_AVG_FOR_REDUCED_REST": True,
        "RUN_OVER50H": True,
        "RUN_STREAK": True,
        "RUN_SPAN": True,
        "RUN_REST11": True,
        "RUN_BREAKS": True,
        "COUNT_TRANSPORT_AS_WORK_FOR_50H": True,
        "COUNT_TRANSPORT_AS_WORK_FOR_STREAK": True,
        "COUNT_TRANSPORT_AS_WORK_FOR_BREAKS": True,
        "COUNT_TRANSPORT_FOR_SERVICE_BOUNDARY": True,
        "PAUSE_CODES": {"16009", "95900"},
        "TRANSPORT_CODES": {"61800", "61010"},
        "EXCLUDE_PRESTATIONS": {"196", "60041"},
        "PSEUDO_NON_WORK_CODES": {"195"},
    }

    notebook = json.loads(notebook_path.read_text(encoding="utf-8"))
    for cell_index in range(2, 9):
        source = "".join(notebook["cells"][cell_index]["source"])
        exec(compile(source, f"{notebook_path.name}:cell{cell_index + 1}", "exec"), env)
    return env


def ltr_compute_summary(env: dict, df: pd.DataFrame, services_df: pd.DataFrame, calendar_slices_df: pd.DataFrame, all_infractions: pd.DataFrame, data_quality: pd.DataFrame) -> pd.DataFrame:
    raw_months = set(df["Mois"].dropna().astype(str).tolist()) if "Mois" in df.columns else set()
    service_months = set(services_df["service_month"].dropna().astype(str).tolist()) if not services_df.empty else set()
    calendar_months = set(calendar_slices_df["target_month"].dropna().astype(str).tolist()) if not calendar_slices_df.empty else set()
    inf_months = set(all_infractions["TARGET_MONTH"].dropna().astype(str).tolist()) if not all_infractions.empty else set()
    months = sorted(raw_months | service_months | calendar_months | inf_months)

    rows = []
    for month in months:
        row = {"TARGET_MONTH": month}
        for rule in [
            f"OVER_{int(env['WEEKLY_LIMIT_HOURS'])}H_WEEK",
            "STREAK_7DAYS",
            f"SPAN_OVER_{int(env['SPAN_LIMIT_HOURS'])}H",
            "REST_UNDER_11H",
            "PAUSE_INSUFF",
        ]:
            if all_infractions.empty:
                row[rule] = 0
            else:
                row[rule] = int(
                    (
                        all_infractions.get("TARGET_MONTH", pd.Series(dtype=str)).astype(str).eq(month)
                        & all_infractions.get("RULE", pd.Series(dtype=str)).astype(str).eq(rule)
                    ).sum()
                )
        row["TOTAL_INFRACTIONS"] = int(all_infractions["TARGET_MONTH"].astype(str).eq(month).sum()) if not all_infractions.empty else 0
        row["SERVICES_STARTED"] = int(services_df["service_month"].astype(str).eq(month).sum()) if not services_df.empty else 0
        row["CALENDAR_HOURS"] = float(calendar_slices_df.loc[calendar_slices_df["target_month"].astype(str).eq(month), "hours"].sum()) if not calendar_slices_df.empty else 0.0
        row["ORPHAN_PAUSE_ROWS"] = int(
            (
                data_quality.get("TARGET_MONTH", pd.Series(dtype=str)).astype(str).eq(month)
                & data_quality.get("QUALITY_TYPE", pd.Series(dtype=str)).astype(str).eq("ORPHAN_PAUSE_NO_SERVICE")
            ).sum()
        ) if not data_quality.empty else 0
        row["INVALID_INTERVAL_ROWS"] = int(
            (
                data_quality.get("TARGET_MONTH", pd.Series(dtype=str)).astype(str).eq(month)
                & data_quality.get("QUALITY_TYPE", pd.Series(dtype=str)).astype(str).eq("INVALID_INTERVAL")
            ).sum()
        ) if not data_quality.empty else 0
        rows.append(row)
    return pd.DataFrame(rows)


def ltr_write_workbook(env: dict, workbook_path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    orders = {
        "ALL_INFRACTIONS": ["TARGET_MONTH", "EVENT_DATE", "RULE", "SEVERITY", "Collaborateur", "No_collaborateur_codes", "Match_status", "DETAIL", "service_id", "service_date", "service_start", "service_end"],
        "SERVICES_AUDIT": ["service_month", "service_date", "service_start", "service_end", "continues_after_midnight", "continuation_row_count", "creates_worked_day", "Collaborateur", "No_collaborateur_codes", "Match_status", "amplitude_hours", "net_50h_minutes", "net_breaks_minutes", "pause_minutes_inside_service", "attached_calendar_dates", "service_id"],
        "CALENDAR_HOUR_SLICES": ["target_month", "calendar_date", "week_monday", "slice_start", "slice_end", "minutes", "hours", "continuation_from_previous_service", "service_date", "Collaborateur", "No_collaborateur_codes", "service_id", "note"],
        "DATA_QUALITY": ["TARGET_MONTH", "EVENT_DATE", "QUALITY_TYPE", "DETAIL", "Collaborateur", "No collaborateur", "No prestation", "Prestation", "start_dt_local", "end_dt_local", "interval_status", "service_id", "service_date", "continuation_from_previous_service"],
    }
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="w") as writer:
        for sheet_name, df_sheet in sheets.items():
            clean = env["excel_safe_no_tz"](env["prep_for_export"](df_sheet, drop_collaborateur_id=True))
            clean = env["reorder_columns"](clean, orders.get(sheet_name, []))
            clean = env["ensure_unique_columns"](clean)
            clean.to_excel(writer, index=False, sheet_name=sheet_name[:31])
    env["apply_swiss_formats_xlsx"](str(workbook_path))


def ltr_process(matched_upload, rda_upload, output_name: str) -> dict:
    output_root = ltr_unique_output_root(output_name)
    input_dir = output_root / "inputs"
    matched_path = ltr_save_upload(matched_upload, input_dir)
    rda_path = ltr_save_upload(rda_upload, input_dir)
    workbook_path = output_root / "multiple" / "LTR_CHECKS_MULTIPLE_ALL_MONTHS_HYBRID.xlsx"

    notebook_path = LTR_NOTEBOOK_PATH
    env = ltr_load_notebook_functions(notebook_path.stat().st_mtime)

    df_raw = env["load_and_normalize"](str(rda_path))
    matched = env["load_matched_collabs"](str(matched_path))
    df_raw = env["attach_collab_master"](df_raw, matched)
    df = env["add_interval_columns"](df_raw)
    services_df, df_tagged, orphan_pauses_df = env["build_services_and_tagged_rows"](df)
    calendar_slices_df = env["build_calendar_slices"](services_df)

    over50_detail, over50_all = env["check_over_50h"](calendar_slices_df)
    streak_detail, streak_all = env["check_streak_7days"](services_df)
    span_detail, span_all = env["check_span_over_14h"](services_df)
    rest_detail, rest_all, rest_review = env["check_rest_under_11h"](services_df)
    breaks_detail, breaks_all, breaks_audit = env["check_breaks"](services_df)
    data_quality = env["build_data_quality"](df_tagged, orphan_pauses_df)

    infraction_parts = [part for part in [over50_all, streak_all, span_all, rest_all, breaks_all] if part is not None and not part.empty]
    if infraction_parts:
        all_infractions = pd.concat(infraction_parts, ignore_index=True, sort=False)
        all_infractions = env["ensure_unique_columns"](all_infractions)
        sort_cols = [col for col in ["TARGET_MONTH", "EVENT_DATE", "Collaborateur", "RULE"] if col in all_infractions.columns]
        if sort_cols:
            all_infractions = all_infractions.sort_values(sort_cols, kind="stable").reset_index(drop=True)
    else:
        all_infractions = pd.DataFrame(columns=["TARGET_MONTH", "EVENT_DATE", "RULE", "SEVERITY", "DETAIL", "collab_uid", "Collaborateur"])

    summary_by_month = ltr_compute_summary(env, df, services_df, calendar_slices_df, all_infractions, data_quality)
    services_audit = services_df.drop(columns=[col for col in ["_net_50_intervals", "_net_breaks_intervals"] if col in services_df.columns]).copy()
    unmatched_mask = df["collab_key"].astype(str).str.startswith("UNMATCHED")
    ambig_mask = df["collab_match_status"].astype(str).eq("AMBIG_MATCH")
    unmatched_df = df[unmatched_mask | ambig_mask].copy()
    if unmatched_df.empty:
        unrecognized_summary = pd.DataFrame(columns=["Collaborateur", "No collaborateur", "collab_match_status", "collab_key", "Row Count"])
    else:
        unrecognized_summary = (
            unmatched_df.groupby(["Collaborateur", "No collaborateur", "collab_match_status", "collab_key"], dropna=False)
            .size()
            .reset_index(name="Row Count")
            .sort_values("Row Count", ascending=False)
            .reset_index(drop=True)
        )

    sheets = {
        "SUMMARY_BY_MONTH": summary_by_month,
        "ALL_INFRACTIONS": all_infractions,
        "OVER_50H_WEEK": over50_detail,
        "STREAK_7DAYS": streak_detail,
        "SPAN_OVER_14H": span_detail,
        "REST_UNDER_11H": rest_detail,
        "REST_REVIEW_ALLOWED": rest_review,
        "PAUSE_INSUFF": breaks_detail,
        "PAUSE_AUDIT_SERVICES": breaks_audit,
        "SERVICES_AUDIT": services_audit,
        "CALENDAR_HOUR_SLICES": calendar_slices_df,
        "DATA_QUALITY": data_quality,
    }
    ltr_write_workbook(env, workbook_path, sheets)

    return {
        "output_root": output_root,
        "workbook_path": workbook_path,
        "sheets": sheets,
        "summary_by_month": summary_by_month,
        "all_infractions": all_infractions,
        "rest_review": rest_review,
        "data_quality": data_quality,
        "services_audit": services_audit,
        "calendar_slices": calendar_slices_df,
        "breaks_audit": breaks_audit,
        "unrecognized_summary": unrecognized_summary,
        "metrics": {
            "raw_rows": len(df),
            "services": len(services_df),
            "calendar_slices": len(calendar_slices_df),
            "infractions": len(all_infractions),
            "affected_collaborators": all_infractions["collab_uid"].nunique() if "collab_uid" in all_infractions.columns and not all_infractions.empty else 0,
            "rest_review_rows": len(rest_review),
            "data_quality_rows": len(data_quality),
            "unrecognized_rows": int(unrecognized_summary["Row Count"].sum()) if not unrecognized_summary.empty else 0,
        },
    }


def ltr_filtered_df(df: pd.DataFrame, filters: dict) -> pd.DataFrame:
    out = df.copy()
    if out.empty:
        return out
    if filters.get("months") and "TARGET_MONTH" in out.columns:
        out = out[out["TARGET_MONTH"].astype(str).isin(filters["months"])]
    if filters.get("rules") and "RULE" in out.columns:
        out = out[out["RULE"].astype(str).isin(filters["rules"])]
    if filters.get("severities") and "SEVERITY" in out.columns:
        out = out[out["SEVERITY"].astype(str).isin(filters["severities"])]
    if filters.get("collaborators") and "Collaborateur" in out.columns:
        out = out[out["Collaborateur"].astype(str).isin(filters["collaborators"])]
    return out


def render_ltr_chart(title: str, df: pd.DataFrame, x_col: str, y_col: str) -> None:
    st.subheader(title)
    if df.empty or x_col not in df.columns or y_col not in df.columns:
        st.info("Aucune donnée de graphique.")
        return
    chart_df = df[[x_col, y_col]].copy()
    chart_df[x_col] = chart_df[x_col].astype(str)
    st.bar_chart(chart_df.set_index(x_col))


def render_ltr_dashboard(result: dict) -> None:
    metrics = result["metrics"]
    metric_cols = st.columns(6)
    metric_cols[0].metric("Infractions", f"{metrics['infractions']:,}")
    metric_cols[1].metric("Collaborateurs concernés", f"{metrics['affected_collaborators']:,}")
    metric_cols[2].metric("Services", f"{metrics['services']:,}")
    metric_cols[3].metric("Créneaux calendrier", f"{metrics['calendar_slices']:,}")
    metric_cols[4].metric("Qualité des données", f"{metrics['data_quality_rows']:,}")
    metric_cols[5].metric("Lignes non reconnues", f"{metrics['unrecognized_rows']:,}")

    if result["workbook_path"].exists():
        render_download_for_path(result["workbook_path"], "Télécharger le classeur LTR")

    all_infractions = result["all_infractions"]
    filter_cols = st.columns(4)
    months = sorted(all_infractions["TARGET_MONTH"].dropna().astype(str).unique().tolist()) if "TARGET_MONTH" in all_infractions.columns and not all_infractions.empty else []
    rules = sorted(all_infractions["RULE"].dropna().astype(str).unique().tolist()) if "RULE" in all_infractions.columns and not all_infractions.empty else []
    severities = sorted(all_infractions["SEVERITY"].dropna().astype(str).unique().tolist()) if "SEVERITY" in all_infractions.columns and not all_infractions.empty else []
    collaborators = sorted(all_infractions["Collaborateur"].dropna().astype(str).unique().tolist()) if "Collaborateur" in all_infractions.columns and not all_infractions.empty else []
    filters = {
        "months": filter_cols[0].multiselect("Mois", months),
        "rules": filter_cols[1].multiselect("Règle", rules),
        "severities": filter_cols[2].multiselect("Sévérité", severities),
        "collaborators": filter_cols[3].multiselect("Collaborateur", collaborators),
    }
    filtered = ltr_filtered_df(all_infractions, filters)

    chart_cols = st.columns(2)
    with chart_cols[0]:
        render_ltr_chart("Infractions par mois", result["summary_by_month"], "TARGET_MONTH", "TOTAL_INFRACTIONS")
    with chart_cols[1]:
        if filtered.empty or "RULE" not in filtered.columns:
            st.subheader("Infractions par règle")
            st.info("Aucune donnée de graphique.")
        else:
            by_rule = filtered.groupby("RULE").size().reset_index(name="Count")
            render_ltr_chart("Infractions par règle", by_rule, "RULE", "Count")

    chart_cols_2 = st.columns(2)
    with chart_cols_2[0]:
        if filtered.empty or "Collaborateur" not in filtered.columns:
            st.subheader("Principaux collaborateurs")
            st.info("Aucune donnée de graphique.")
        else:
            top_collabs = filtered.groupby("Collaborateur").size().sort_values(ascending=False).head(15).reset_index(name="Count")
            render_ltr_chart("Principaux collaborateurs", top_collabs, "Collaborateur", "Count")
    with chart_cols_2[1]:
        dq = result["data_quality"]
        if dq.empty or "QUALITY_TYPE" not in dq.columns:
            st.subheader("Types de qualité des données")
            st.info("Aucune donnée de graphique.")
        else:
            dq_counts = dq.groupby("QUALITY_TYPE").size().reset_index(name="Count")
            render_ltr_chart("Types de qualité des données", dq_counts, "QUALITY_TYPE", "Count")

    st.subheader("Infractions filtrées")
    st.dataframe(filtered, width="stretch", hide_index=True)
    if not filtered.empty:
        st.download_button(
            "Télécharger le CSV des infractions filtrées",
            filtered.to_csv(index=False, encoding="utf-8-sig"),
            file_name="ltr_filtered_infractions.csv",
            mime="text/csv",
        )

    tab_names = ["Résumé", "Revue des repos", "Qualité des données", "Services", "Audit des pauses", "Créneaux calendrier", "Non reconnus"]
    tabs = st.tabs(tab_names)
    with tabs[0]:
        st.dataframe(result["summary_by_month"], width="stretch", hide_index=True)
    with tabs[1]:
        st.dataframe(result["rest_review"], width="stretch", hide_index=True)
    with tabs[2]:
        data_quality = result["data_quality"]
        dq_types = sorted(data_quality["QUALITY_TYPE"].dropna().astype(str).unique().tolist()) if "QUALITY_TYPE" in data_quality.columns and not data_quality.empty else []
        selected_dq = st.multiselect("Type de qualité des données", dq_types)
        if selected_dq:
            data_quality = data_quality[data_quality["QUALITY_TYPE"].astype(str).isin(selected_dq)]
        st.dataframe(data_quality, width="stretch", hide_index=True)
    with tabs[3]:
        st.dataframe(result["services_audit"], width="stretch", hide_index=True)
    with tabs[4]:
        st.dataframe(result["breaks_audit"], width="stretch", hide_index=True)
    with tabs[5]:
        st.dataframe(result["calendar_slices"], width="stretch", hide_index=True)
    with tabs[6]:
        st.dataframe(result["unrecognized_summary"], width="stretch", hide_index=True)


def render_ltr_task() -> None:
    st.title("Contrôles LTR")
    st.caption("Exécute les contrôles LTR hybrides et crée le classeur Excel multi-feuilles avec un tableau de bord d'audit.")

    cols = st.columns(3)
    matched_file = cols[0].file_uploader("Classeur collaborateurs matchés", type=["xlsx", "xls"], key="ltr_matched")
    rda_file = cols[1].file_uploader("Fichier RDA fusionné", type=["xlsx", "xls", "csv"], key="ltr_rda")
    output_name = cols[2].text_input("Nom du dossier de sortie", value="")

    if st.button("Lancer les contrôles LTR", type="primary", disabled=matched_file is None or rda_file is None):
        progress = st.progress(0.0, text="Démarrage des contrôles LTR")
        try:
            progress.progress(0.1, text="Chargement des fichiers et de la logique notebook")
            result = ltr_process(matched_file, rda_file, output_name)
            progress.progress(1.0, text="Contrôles LTR terminés")
            st.session_state["latest_ltr_result"] = result
        except Exception as exc:
            progress.empty()
            st.exception(exc)
            return

    result = st.session_state.get("latest_ltr_result")
    if result:
        st.success(f"Classeur LTR créé : {result['workbook_path']}")
        render_ltr_dashboard(result)


# ============================================================
# Audit Webfleet-RDA — low-level helpers
# ============================================================

def audit_to_int_str(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    s = re.sub(r"\.0$", "", s)
    s = re.sub(r"\s+", "", s)
    try:
        return str(int(float(s)))
    except Exception:
        return s if s else None


def audit_extract_code_any(x):
    if pd.isna(x):
        return np.nan
    m = re.search(r"(\d{4,6})", str(x))
    return m.group(1) if m else np.nan


def audit_swiss_date(series_like):
    return pd.to_datetime(series_like, errors="coerce", dayfirst=True).dt.date


def audit_swiss_dt(series_like, tz_name=AUDIT_TZ_NAME):
    s = pd.to_datetime(series_like, errors="coerce", dayfirst=True)
    try:
        if s.dt.tz is None:
            return s.dt.tz_localize(tz_name, ambiguous="infer", nonexistent="shift_forward")
        return s.dt.tz_convert(tz_name)
    except Exception:
        def one(x):
            if pd.isna(x):
                return pd.NaT
            x = pd.Timestamp(x)
            if x.tzinfo is None:
                return x.tz_localize(tz_name, ambiguous="infer", nonexistent="shift_forward")
            return x.tz_convert(tz_name)
        return pd.Series(series_like).apply(one)


def audit_ensure_tz(series_like, tz_name=AUDIT_TZ_NAME):
    raw = pd.Series(series_like).copy()
    s = raw.astype(str).str.strip()
    has_tz = s.str.contains(r"(?:Z|[+-]\d{2}:?\d{2})$", regex=True, na=False)
    out = pd.Series(pd.NaT, index=raw.index, dtype=f"datetime64[ns, {tz_name}]")
    if has_tz.any():
        parsed = pd.to_datetime(raw.loc[has_tz], errors="coerce", utc=True)
        out.loc[has_tz] = parsed.dt.tz_convert(tz_name)
    if (~has_tz).any():
        parsed = pd.to_datetime(raw.loc[~has_tz], errors="coerce", dayfirst=True)
        out.loc[~has_tz] = parsed.dt.tz_localize(tz_name, ambiguous="infer", nonexistent="shift_forward")
    return out


def audit_align_to_zurich(s, tz_name=AUDIT_TZ_NAME):
    s2 = pd.to_datetime(s, errors="coerce")
    try:
        if s2.dt.tz is None:
            return s2.dt.tz_localize(tz_name, ambiguous="infer", nonexistent="shift_forward")
        return s2.dt.tz_convert(tz_name)
    except Exception:
        def fix_one(x):
            if pd.isna(x):
                return pd.NaT
            x = pd.Timestamp(x)
            try:
                if x.tzinfo is None:
                    return x.tz_localize(tz_name, ambiguous="infer", nonexistent="shift_forward")
                return x.tz_convert(tz_name)
            except Exception:
                return x
        return s.apply(fix_one)


def audit_to_local_naive(ts, tz_name=AUDIT_TZ_NAME):
    if pd.isna(ts):
        return pd.NaT
    ts = pd.Timestamp(ts)
    try:
        if ts.tzinfo is not None:
            return ts.tz_convert(tz_name).tz_localize(None)
    except Exception:
        try:
            return ts.tz_localize(None)
        except Exception:
            pass
    return ts


def audit_series_to_local_naive(s, tz_name=AUDIT_TZ_NAME):
    return pd.to_datetime(s.apply(lambda x: audit_to_local_naive(x, tz_name)), errors="coerce")


def audit_merge_blocks(grp_df, gap_min):
    df = grp_df[["start", "end"]].dropna().sort_values("start")
    if df.empty:
        return []
    merged = []
    cur_s, cur_e = df.iloc[0]["start"], df.iloc[0]["end"]
    for _, r in df.iloc[1:].iterrows():
        s, e = r["start"], r["end"]
        gap = (s - cur_e).total_seconds() / 60.0
        if gap <= gap_min:
            cur_e = max(cur_e, e)
        else:
            merged.append((cur_s, cur_e))
            cur_s, cur_e = s, e
    merged.append((cur_s, cur_e))
    return merged


def audit_trip_km_from_cols(df):
    cand_km = ["distance", "Distance", "distance_km", "mileage [km]", "km", "Mileage (km)", "Trip distance [km]"]
    cand_m = ["distance [m]", "Distance [m]", "Trip distance [m]", "distance_m", "mileage [m]", "Distance (m)", "Meters"]
    dcol = rda_pick_col(df, cand_km + cand_m)
    if dcol:
        vals = pd.to_numeric(df[dcol], errors="coerce")
        header = str(dcol).lower()
        looks_meter = ("[m]" in header) or ("(m)" in header) or ("meter" in header) or (dcol in cand_m)
        q95 = np.nanquantile(vals, 0.95) if np.isfinite(vals).any() else np.nan
        if looks_meter or (pd.notna(q95) and q95 > 800):
            return (vals / 1000.0).astype(float)
        return vals.astype(float)
    return pd.Series(np.nan, index=df.index, dtype=float)


def audit_series_speed_kmh(km, start, end):
    dur_h = (pd.to_datetime(end) - pd.to_datetime(start)).dt.total_seconds() / 3600.0
    with np.errstate(divide="ignore", invalid="ignore"):
        spd = km / dur_h
    spd = spd.replace([np.inf, -np.inf], np.nan)
    return spd


def audit_pick_best_sheet(file_bytes, required_groups, prefer_code=None):
    xf = pd.ExcelFile(BytesIO(file_bytes))
    best = None
    best_score = -1
    for name in xf.sheet_names:
        df = xf.parse(name)
        cols = set(df.columns.astype(str))
        score = sum(1 for grp in required_groups if any(c in cols for c in grp))
        if prefer_code is not None:
            try:
                hits = df.astype(str).apply(lambda c: c.str.contains(prefer_code, na=False)).sum().sum()
                if hits > 0:
                    score += 2
            except Exception:
                pass
        if score > best_score:
            best_score = score
            best = (name, df)
    return best


def audit_drop_tz_excel_safe(df, tz_name=AUDIT_TZ_NAME):
    from pandas.api import types as pdt
    out = df.copy()
    for c in out.columns:
        if pdt.is_datetime64tz_dtype(out[c]):
            s = out[c]
            try:
                s = s.dt.tz_convert(tz_name)
            except Exception:
                pass
            out[c] = s.dt.tz_localize(None)
    return out


def audit_fmt_hhmm(ts, tz_name=AUDIT_TZ_NAME):
    ts = audit_to_local_naive(ts, tz_name)
    if pd.isna(ts):
        return "-"
    return pd.Timestamp(ts).strftime("%H:%M")


def audit_fmt_hours(delta):
    if delta is None or pd.isna(delta):
        return "-"
    total_min = int(round(delta.total_seconds() / 60.0))
    if total_min < 0:
        return "-"
    hh = total_min // 60
    mm = total_min % 60
    return f"{hh}h{mm:02d}" if mm else f"{hh}h"


def audit_fmt_span(s, e):
    s = audit_to_local_naive(s)
    e = audit_to_local_naive(e)
    if pd.isna(s) or pd.isna(e) or e <= s:
        return "-"
    return f"{audit_fmt_hhmm(s)} → {audit_fmt_hhmm(e)} ({audit_fmt_hours(e - s)})"


def audit_uniq_flags(series):
    vals = series.fillna("").astype(str)
    s = set()
    for v in vals:
        if not v or v == "nan":
            continue
        for f in v.split(","):
            f = f.strip()
            if f:
                s.add(f)
    return ",".join(sorted(s))


def audit_safe_filename(x):
    s = ("" if pd.isna(x) else str(x)).strip()
    s = re.sub(r"[^\w\-. ]+", "_", s)
    s = re.sub(r"\s+", "_", s)
    return s[:180] if s else "collaborateur"


def audit_duration_mins(a, b):
    if pd.isna(a) or pd.isna(b):
        return np.nan
    return (pd.Timestamp(b) - pd.Timestamp(a)).total_seconds() / 60.0


# ============================================================
# Audit — planning date resolution helpers
# ============================================================

def _audit_as_local_timestamp(value, tz_name=AUDIT_TZ_NAME):
    if pd.isna(value):
        return pd.NaT
    try:
        ts = pd.Timestamp(value)
    except Exception:
        return pd.NaT
    if pd.isna(ts):
        return pd.NaT
    try:
        if ts.tzinfo is not None:
            return ts.tz_convert(tz_name).tz_localize(None)
    except Exception:
        try:
            return ts.tz_localize(None)
        except Exception:
            pass
    return ts


def _audit_valid_planning_date(ts):
    ts = _audit_as_local_timestamp(ts)
    if pd.isna(ts):
        return None
    if AUDIT_PLANNING_DATE_MIN_YEAR <= ts.year <= AUDIT_PLANNING_DATE_MAX_YEAR:
        return ts.date()
    return None


def _audit_excel_or_compact_date(value):
    try:
        v = float(value)
    except Exception:
        return None
    if not np.isfinite(v):
        return None
    if 20000 <= v <= 80000:
        return _audit_valid_planning_date(pd.Timestamp("1899-12-30") + pd.to_timedelta(v, unit="D"))
    if float(v).is_integer():
        text = str(int(v))
        if re.fullmatch(r"20\d{6}", text):
            return _audit_valid_planning_date(pd.to_datetime(text, format="%Y%m%d", errors="coerce"))
    return None


def _audit_try_timestamp_ymd(year, month, day):
    try:
        return pd.Timestamp(year=year, month=month, day=day)
    except ValueError:
        return pd.NaT


def _audit_planning_scalar_date_candidates(value):
    from datetime import date as _dt, datetime as _dtt, time as _tt
    candidates = []

    def add(val, src, pri):
        dt = _audit_valid_planning_date(val)
        if dt is not None:
            candidates.append((dt, src, pri))

    if pd.isna(value):
        return candidates

    if isinstance(value, (pd.Timestamp, _dtt)):
        add(value, "native_datetime", 500)
        return _audit_unique_date_candidates(candidates)

    if isinstance(value, _dt):
        add(pd.Timestamp(value), "native_date", 500)
        return _audit_unique_date_candidates(candidates)

    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        dt = _audit_excel_or_compact_date(value)
        if dt is not None:
            candidates.append((dt, "numeric_date", 480))
        return _audit_unique_date_candidates(candidates)

    text = str(value).strip()
    if not text or text.lower() in {"nan", "nat", "none", "null"}:
        return candidates

    dt = _audit_excel_or_compact_date(text)
    if dt is not None:
        candidates.append((dt, "numeric_text_date", 470))
        return _audit_unique_date_candidates(candidates)

    if re.match(r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}(?:\s|$)", text):
        add(pd.to_datetime(text, errors="coerce", yearfirst=True), "string_iso", 460)
        return _audit_unique_date_candidates(candidates)

    m = re.match(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})(?:\s.*)?$", text)
    if m:
        a = int(m.group(1))
        b = int(m.group(2))
        year = int(m.group(3))
        if year < 100:
            year += 2000
        if 1 <= a <= 31 and 1 <= b <= 31:
            if a > 12 and b <= 12:
                add(_audit_try_timestamp_ymd(year, b, a), "string_dayfirst_unambiguous", 430)
            elif b > 12 and a <= 12:
                add(_audit_try_timestamp_ymd(year, a, b), "string_monthfirst_unambiguous", 430)
            else:
                add(_audit_try_timestamp_ymd(year, b, a), "string_dayfirst_ambiguous", 300)
                add(_audit_try_timestamp_ymd(year, a, b), "string_monthfirst_ambiguous", 300)

    add(pd.to_datetime(text, errors="coerce", dayfirst=True), "string_dayfirst_fallback", 120)
    add(pd.to_datetime(text, errors="coerce", dayfirst=False), "string_monthfirst_fallback", 120)
    return _audit_unique_date_candidates(candidates)


def _audit_unique_date_candidates(candidates):
    best = {}
    for dt, source, priority in candidates:
        old = best.get(dt)
        if old is None or priority > old[1]:
            best[dt] = (source, priority)
    return [(dt, source, priority) for dt, (source, priority) in best.items()]


def _audit_planning_date_match_score(cid, dt, ref_by_collab, ref_all):
    if dt is None:
        return 0
    score = 0
    cid_text = None if pd.isna(cid) else str(cid)
    if cid_text and dt in ref_by_collab.get(cid_text, set()):
        score += 1000
    if dt in ref_all:
        score += 100
    return score


def _audit_candidate_for_order(value, order, ref_by_collab, ref_all, date_order_resolved):
    target = f"string_{order}"
    matches = [c for c in _audit_planning_scalar_date_candidates(value) if c[1].startswith(target)]
    if not matches:
        return None
    return sorted(matches, key=lambda x: x[2], reverse=True)[0][0]


def audit_build_reference_dates(rda_df, wf_df):
    ref_by_collab = {}
    ref_all = set()

    def add_ref(cid, value):
        if pd.isna(cid):
            return
        cands = _audit_planning_scalar_date_candidates(value)
        if not cands:
            return
        dt = cands[0][0]
        cid = str(cid)
        ref_by_collab.setdefault(cid, set()).add(dt)
        ref_all.add(dt)

    if not rda_df.empty and "collab_id" in rda_df.columns:
        for rr in rda_df.dropna(subset=["collab_id"]).itertuples(index=False):
            jour = getattr(rr, "jour", pd.NaT)
            start = getattr(rr, "start", pd.NaT)
            add_ref(rr.collab_id, jour if pd.notna(jour) else start)

    if not wf_df.empty and "collab_id" in wf_df.columns:
        for rr in wf_df.dropna(subset=["collab_id"]).itertuples(index=False):
            d = getattr(rr, "date", pd.NaT)
            start = getattr(rr, "start", pd.NaT)
            add_ref(rr.collab_id, d if pd.notna(d) else start)

    return ref_by_collab, ref_all


def audit_resolve_planning_date_order(raw_dates, collab_ids, ref_by_collab, ref_all):
    scores = {"dayfirst": 0, "monthfirst": 0}
    for value, cid in zip(raw_dates, collab_ids):
        if pd.isna(value):
            continue
        text = str(value).strip()
        if not re.match(r"^\d{1,2}[./-]\d{1,2}[./-]\d{2,4}(?:\s.*)?$", text):
            continue
        for order in ["dayfirst", "monthfirst"]:
            dt = _audit_candidate_for_order(value, order, ref_by_collab, ref_all, order)
            scores[order] += _audit_planning_date_match_score(cid, dt, ref_by_collab, ref_all)
    if scores["dayfirst"] > scores["monthfirst"]:
        return "dayfirst", scores
    if scores["monthfirst"] > scores["dayfirst"]:
        return "monthfirst", scores
    return AUDIT_PLANNING_DATE_TIE_BREAKER, scores


def audit_select_planning_date(value, cid, date_order_resolved, ref_by_collab, ref_all):
    candidates = _audit_planning_scalar_date_candidates(value)
    if not candidates:
        return pd.NaT, "unparsed"
    ranked = []
    for dt, source, priority in candidates:
        score = priority + _audit_planning_date_match_score(cid, dt, ref_by_collab, ref_all)
        if source.startswith(f"string_{date_order_resolved}"):
            score += 20
        ranked.append((score, dt, source))
    ranked.sort(key=lambda x: x[0], reverse=True)
    _, dt, source = ranked[0]
    return dt, source


def audit_planning_time_parts(value):
    from datetime import datetime as _dtt, time as _tt
    if pd.isna(value):
        return None
    if isinstance(value, (pd.Timestamp, _dtt)):
        ts = _audit_as_local_timestamp(value)
        if pd.notna(ts):
            return int(ts.hour), int(ts.minute), int(ts.second)
    if isinstance(value, _tt):
        return int(value.hour), int(value.minute), int(value.second)
    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        try:
            v = float(value)
        except Exception:
            return None
        if not np.isfinite(v):
            return None
        if 0 <= v < 1:
            total_seconds = int(round(v * 86400)) % 86400
        elif 1 <= v < 24:
            total_seconds = int(round(v * 3600)) % 86400
        elif 20000 <= v <= 80000:
            total_seconds = int(round((v % 1) * 86400)) % 86400
        else:
            return None
        return total_seconds // 3600, (total_seconds % 3600) // 60, total_seconds % 60
    text = str(value).strip()
    if not text or text.lower() in {"nan", "nat", "none", "null"}:
        return None
    m = re.search(r"(\d{1,2})[:hH](\d{2})(?::(\d{2}))?", text)
    if m:
        hh, mm, ss = int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)
        if 0 <= hh < 24 and 0 <= mm < 60 and 0 <= ss < 60:
            return hh, mm, ss
    ts = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.notna(ts):
        ts = _audit_as_local_timestamp(ts)
        return int(ts.hour), int(ts.minute), int(ts.second)
    return None


def audit_combine_plan_date_time(parsed_dates, time_ser, tz_name=AUDIT_TZ_NAME):
    from datetime import time as _tt
    values = []
    for d, time_value in zip(parsed_dates, time_ser):
        parts = audit_planning_time_parts(time_value)
        if pd.isna(d) or parts is None:
            values.append(pd.NaT)
            continue
        hh, mm, ss = parts
        values.append(pd.Timestamp.combine(d, _tt(hh, mm, ss)))
    out = pd.to_datetime(pd.Series(values, index=time_ser.index), errors="coerce")
    out = out.dt.tz_localize(tz_name, ambiguous="infer", nonexistent="shift_forward")
    return out


# ============================================================
# Audit — main processing function
# ============================================================

def audit_process(rda_file, wf_file, mapping_file, planning_file, progress_cb=None):
    def _prog(pct, msg):
        if progress_cb:
            progress_cb(pct, msg)

    _prog(0.05, "Chargement des fichiers...")

    RDA_REQUIRED = [
        ["Jour", "Date", "date"],
        ["Début", "Debut", "Heure Début", "Heure Debut", "Start", "start"],
        ["Fin", "Heure fin", "End", "end", "fin"],
        ["Durée", "Duree", "duration", "Minutes"],
        ["No collaborateur", "No Collaborateur", "Employee No", "no_collaborateur"],
    ]
    WF_REQUIRED = [
        ["tripid", "Trip ID", "TripId"],
        ["tripmode", "Trip Mode", "trip_mode"],
        ["start_time", "Start Time", "Start time"],
        ["end_time", "End Time", "End time"],
        ["driverno", "Driver No", "driver_no"],
    ]
    PLANNING_REQUIRED = [
        ["emp_nr"], ["date"], ["start"], ["end"], ["event_color"], ["client_absent"],
    ]

    rda_bytes = rda_file.read()
    wf_bytes = wf_file.read()
    map_bytes = mapping_file.read()
    plan_bytes = planning_file.read()

    rda_name = rda_file.name
    wf_name = wf_file.name

    def _strip_bom(df):
        df.columns = [str(c).lstrip("﻿").strip() for c in df.columns]
        return df

    if rda_name.lower().endswith(".csv"):
        RDA = _strip_bom(read_csv_flex(BytesIO(rda_bytes)))
    else:
        _, RDA = audit_pick_best_sheet(
            rda_bytes, RDA_REQUIRED,
            prefer_code=AUDIT_PRESTATION_61010_CODE if AUDIT_ENABLE_61010_FEATURE else None
        )
        RDA = _strip_bom(RDA)

    if wf_name.lower().endswith(".csv"):
        WF = _strip_bom(read_csv_flex(BytesIO(wf_bytes)))
    else:
        _, WF = audit_pick_best_sheet(wf_bytes, WF_REQUIRED)
        WF = _strip_bom(WF)

    _, PLANNING = audit_pick_best_sheet(plan_bytes, PLANNING_REQUIRED)
    PLANNING = _strip_bom(PLANNING)

    map_sheets = pd.read_excel(BytesIO(map_bytes), sheet_name=None)
    MAP = _strip_bom(map_sheets.get("Matched Collaborateurs", next(iter(map_sheets.values()))))

    _prog(0.12, "Normalisation RDA...")

    rda_cols = {
        "jour": rda_pick_col(RDA, ["Jour", "Date", "date"]),
        "debut": rda_pick_col(RDA, ["Début", "Debut", "Heure Début", "Heure Debut", "Start", "start"]),
        "fin": rda_pick_col(RDA, ["Fin", "Heure fin", "End", "end", "fin"]),
        "duree": rda_pick_col(RDA, ["Durée", "Duree", "duration", "Minutes"]),
        "collab_name": rda_pick_col(RDA, ["Collaborateur", "collaborateur", "Employee"]),
        "collab_no": rda_pick_col(RDA, ["No collaborateur", "No Collaborateur", "Employee No", "no_collaborateur"]),
    }
    missing = [k for k, v in rda_cols.items() if v is None and k in ["jour", "debut", "fin", "duree", "collab_no"]]
    if missing:
        raise KeyError(f"RDA colonnes manquantes : {missing}. Colonnes trouvées : {list(RDA.columns)}")

    rda = pd.DataFrame({
        "jour": audit_swiss_date(RDA[rda_cols["jour"]]),
        "start": audit_swiss_dt(RDA[rda_cols["debut"]]),
        "end": audit_swiss_dt(RDA[rda_cols["fin"]]),
        "duree_min": pd.to_numeric(RDA[rda_cols["duree"]], errors="coerce"),
        "collab_name": (RDA[rda_cols["collab_name"]].astype(str) if rda_cols["collab_name"] else ""),
        "collab_no_sarl": RDA[rda_cols["collab_no"]].apply(audit_to_int_str),
    })
    mask = rda["duree_min"].isna() & rda["start"].notna() & rda["end"].notna()
    rda.loc[mask, "duree_min"] = (rda.loc[mask, "end"] - rda.loc[mask, "start"]).dt.total_seconds() / 60.0

    best_col, best_score = None, -1
    for col in RDA.columns:
        ser = RDA[col].apply(audit_extract_code_any)
        hits_61010 = int((ser.astype(str) == AUDIT_PRESTATION_61010_CODE).sum())
        hits_any = int(ser.notna().sum())
        score = hits_61010 * 100000 + hits_any
        if score > best_score:
            best_score = score
            best_col = col
    rda["prestation_code"] = np.nan
    if best_col is not None:
        rda["prestation_code"] = RDA[best_col].apply(audit_extract_code_any)
    rda["rda_row_id"] = np.arange(len(rda), dtype=int)

    _prog(0.20, "Normalisation Webfleet...")

    wf_cols = {
        "tripid": rda_pick_col(WF, ["tripid", "Trip ID", "TripId"]),
        "tripmode": rda_pick_col(WF, ["tripmode", "Trip Mode", "trip_mode"]),
        "start": rda_pick_col(WF, ["start_time", "Start Time", "Start time"]),
        "end": rda_pick_col(WF, ["end_time", "End Time", "End time"]),
        "driverno": rda_pick_col(WF, ["driverno", "Driver No", "driver_no"]),
        "drivername": rda_pick_col(WF, ["drivername", "Driver Name", "driver_name"]),
    }
    missing_wf = [k for k, v in wf_cols.items() if v is None and k in ["tripid", "tripmode", "start", "end", "driverno"]]
    if missing_wf:
        raise KeyError(f"Webfleet colonnes manquantes : {missing_wf}. Colonnes trouvées : {list(WF.columns)}")

    wf = pd.DataFrame({
        "tripid": WF[wf_cols["tripid"]].astype(str),
        "tripmode": pd.to_numeric(WF[wf_cols["tripmode"]], errors="coerce").astype("Int64"),
        "start": audit_ensure_tz(WF[wf_cols["start"]]),
        "end": audit_ensure_tz(WF[wf_cols["end"]]),
        "driverno": WF[wf_cols["driverno"]].apply(audit_to_int_str),
        "drivername": (WF[wf_cols["drivername"]].astype(str) if wf_cols["drivername"] else ""),
    })
    wf["km"] = pd.to_numeric(audit_trip_km_from_cols(WF), errors="coerce").round(3)
    wf["duration_min"] = (pd.to_datetime(wf["end"]) - pd.to_datetime(wf["start"])).dt.total_seconds() / 60.0
    wf["speed_kmh"] = audit_series_speed_kmh(wf["km"], wf["start"], wf["end"])
    wf["date"] = wf["start"].dt.date

    _prog(0.28, "Chargement mapping collaborateurs...")

    collab_id_col = rda_pick_col(MAP, ["collaborateur-id", "collab_id", "Collaborateur_ID", "collaborateur_id"])
    if not collab_id_col:
        raise KeyError("Mapping : colonne collab_id introuvable (collaborateur-id/collab_id/Collaborateur_ID).")

    map_no_sarl_col = rda_pick_col(MAP, ["no-collaborateur-sarl-102", "collab_no_sarl", "No collaborateur", "No Collaborateur"])
    map_name_sarl_col = rda_pick_col(MAP, ["name-collaborateur", "collaborateur-sarl-102", "collab_sarl", "Collaborateur"])
    map_name_wf_col = rda_pick_col(MAP, ["collaborateur-webfleet", "collab_webfleet", "Driver Name", "name-collaborateur"])
    map_drv_main_col = rda_pick_col(MAP, ["no-collaborateur-wf", "no-collaborateur-webfleet", "driverno", "Driver No"])

    map_df = pd.DataFrame({
        "collab_id": MAP[collab_id_col].astype(str),
        "collab_no_sarl": (MAP[map_no_sarl_col].apply(audit_to_int_str) if map_no_sarl_col else None),
        "collab_name_sarl": (MAP[map_name_sarl_col].astype(str) if map_name_sarl_col else ""),
        "collab_name_wf": (MAP[map_name_wf_col].astype(str) if map_name_wf_col else ""),
        "driverno": (MAP[map_drv_main_col].apply(audit_to_int_str) if map_drv_main_col else None),
    }).dropna(subset=["collab_id"]).drop_duplicates()

    RDA_ID_CANDIDATES = [
        "no-collaborateur-sarl-102", "no-collaborateur-sa-101", "no-collaborateur-ne-103",
        "No collaborateur", "No Collaborateur", "collab_no_sarl",
    ]
    sarlno_to_id = {}
    for col in RDA_ID_CANDIDATES:
        if col in MAP.columns:
            tmp = MAP[[collab_id_col, col]].dropna()
            for _, rr in tmp.iterrows():
                rno = audit_to_int_str(rr[col])
                cid = rr[collab_id_col]
                if rno and pd.notna(cid):
                    sarlno_to_id.setdefault(str(rno), str(cid))
    rda["collab_id"] = rda["collab_no_sarl"].map(sarlno_to_id)

    UO_ID_CANDIDATES = [
        "no-collaborateur-wf", "no-collaborateur-webfleet", "no-collaborateur-sa-101",
        "no-collaborateur-sarl-102", "no-collaborateur-ne-103",
        "collab_no_webfleet", "UO ID", "UO ID 2", "UO ID 3", "driverno", "Driver No",
    ]
    driverno_to_id = {}
    for col in UO_ID_CANDIDATES:
        if col in MAP.columns:
            tmp = MAP[[collab_id_col, col]].dropna()
            for _, rr in tmp.iterrows():
                dno = audit_to_int_str(rr[col])
                cid = rr[collab_id_col]
                if dno and pd.notna(cid):
                    driverno_to_id.setdefault(str(dno), str(cid))
    wf["collab_id"] = wf["driverno"].map(driverno_to_id)

    _prog(0.35, "Normalisation planning...")

    plan_cols = {
        "emp_nr": rda_pick_col(PLANNING, ["emp_nr"]),
        "date": rda_pick_col(PLANNING, ["date"]),
        "start": rda_pick_col(PLANNING, ["start"]),
        "end": rda_pick_col(PLANNING, ["end"]),
        "duration": rda_pick_col(PLANNING, ["duration"]),
        "event_color": rda_pick_col(PLANNING, ["event_color"]),
        "client_absent": rda_pick_col(PLANNING, ["client_absent"]),
        "type": rda_pick_col(PLANNING, ["type"]),
        "note": rda_pick_col(PLANNING, ["note"]),
        "client_nr": rda_pick_col(PLANNING, ["client_nr"]),
    }
    missing_plan = [k for k, v in plan_cols.items() if v is None and k in ["emp_nr", "date", "start", "end", "event_color"]]
    if missing_plan:
        raise KeyError(f"Planning colonnes manquantes : {missing_plan}. Colonnes trouvées : {list(PLANNING.columns)}")

    planning_emp_nr = PLANNING[plan_cols["emp_nr"]].apply(audit_to_int_str)

    PLAN_EMP_ID_CANDIDATES = list(dict.fromkeys(RDA_ID_CANDIDATES + UO_ID_CANDIDATES + [collab_id_col, "emp_nr", "employee_nr", "Employee No"]))
    plan_emp_to_id = dict(sarlno_to_id)

    cols_list = list(MAP.columns)
    id_positions = [i for i, c in enumerate(cols_list) if c == collab_id_col]
    for col in PLAN_EMP_ID_CANDIDATES:
        if col in MAP.columns and col != collab_id_col:
            value_positions = [i for i, c in enumerate(cols_list) if c == col]
            for id_pos in id_positions:
                for value_pos in value_positions:
                    if id_pos == value_pos:
                        continue
                    tmp = MAP.iloc[:, [id_pos, value_pos]].copy()
                    tmp.columns = ["_cid", "_emp"]
                    for _, rr in tmp.dropna(subset=["_cid", "_emp"]).iterrows():
                        emp = audit_to_int_str(rr["_emp"])
                        cid = rr["_cid"]
                        if emp and pd.notna(cid):
                            plan_emp_to_id.setdefault(str(emp), str(cid))

    for cid_pos in id_positions:
        for cid in MAP.iloc[:, cid_pos].dropna().astype(str):
            plan_emp_to_id.setdefault(cid, cid)
            cid_norm = audit_to_int_str(cid)
            if cid_norm:
                plan_emp_to_id.setdefault(cid_norm, cid)

    planning_collab_id = planning_emp_nr.map(plan_emp_to_id)

    ref_by_collab, ref_all = audit_build_reference_dates(rda, wf)
    date_order_resolved, _ = audit_resolve_planning_date_order(
        PLANNING[plan_cols["date"]], planning_collab_id, ref_by_collab, ref_all
    )

    selected_plan_dates = [
        audit_select_planning_date(v, cid, date_order_resolved, ref_by_collab, ref_all)
        for v, cid in zip(PLANNING[plan_cols["date"]], planning_collab_id)
    ]
    planning_date_values = pd.Series([x[0] for x in selected_plan_dates], index=PLANNING.index)
    planning_date_sources = pd.Series([x[1] for x in selected_plan_dates], index=PLANNING.index)

    planning = pd.DataFrame({
        "emp_nr": planning_emp_nr,
        "collab_id": planning_collab_id,
        "date": planning_date_values,
        "date_parse_source": planning_date_sources,
        "raw_date": PLANNING[plan_cols["date"]],
        "raw_start": PLANNING[plan_cols["start"]],
        "raw_end": PLANNING[plan_cols["end"]],
        "start": audit_combine_plan_date_time(planning_date_values, PLANNING[plan_cols["start"]]),
        "end": audit_combine_plan_date_time(planning_date_values, PLANNING[plan_cols["end"]]),
        "duration_min": (pd.to_numeric(PLANNING[plan_cols["duration"]], errors="coerce") if plan_cols["duration"] else np.nan),
        "event_color": PLANNING[plan_cols["event_color"]].fillna("").astype(str).str.strip(),
        "client_absent": (
            PLANNING[plan_cols["client_absent"]].fillna("N").astype(str).str.strip().str.upper()
            if plan_cols["client_absent"] else "N"
        ),
        "type": (pd.to_numeric(PLANNING[plan_cols["type"]], errors="coerce") if plan_cols["type"] else np.nan),
        "note": (PLANNING[plan_cols["note"]].fillna("").astype(str) if plan_cols["note"] else ""),
        "client_nr": (PLANNING[plan_cols["client_nr"]].apply(audit_to_int_str) if plan_cols["client_nr"] else None),
    })

    overnight = planning["start"].notna() & planning["end"].notna() & (planning["end"] < planning["start"])
    planning.loc[overnight, "end"] = planning.loc[overnight, "end"] + pd.Timedelta(days=1)
    dur_missing = planning["duration_min"].isna() & planning["start"].notna() & planning["end"].notna()
    planning.loc[dur_missing, "duration_min"] = (planning.loc[dur_missing, "end"] - planning.loc[dur_missing, "start"]).dt.total_seconds() / 60.0

    planning["date_only"] = planning["date"]
    planning["event_color_key"] = planning["event_color"].fillna("").astype(str).str.strip().str.lower()

    drop_reasons = []
    for rr in planning.itertuples(index=False):
        reasons = []
        if pd.isna(getattr(rr, "collab_id", pd.NA)):
            reasons.append("unmapped_emp_nr")
        if pd.isna(getattr(rr, "date", pd.NaT)):
            reasons.append("bad_date")
        if pd.isna(getattr(rr, "start", pd.NaT)):
            reasons.append("bad_start_time")
        if pd.isna(getattr(rr, "end", pd.NaT)):
            reasons.append("bad_end_time")
        if (pd.notna(getattr(rr, "start", pd.NaT)) and pd.notna(getattr(rr, "end", pd.NaT))
                and getattr(rr, "end") <= getattr(rr, "start")):
            reasons.append("non_positive_span")
        drop_reasons.append(",".join(reasons))
    planning["_drop"] = drop_reasons
    planning = planning[planning["_drop"].str.len() == 0].copy()
    planning.drop(columns=["_drop"], inplace=True)
    planning["collab_id"] = planning["collab_id"].astype(str)
    planning["plot_color"] = np.where(
        planning["client_absent"].eq("Y"),
        "#000000",
        planning["event_color_key"].map(AUDIT_PLAN_COLOR_MAP).fillna("#bdbdbd"),
    )

    _prog(0.45, "Alignement timezone et construction rda_daily...")

    for df_obj in [rda, wf]:
        for col in ["start", "end"]:
            if col in df_obj.columns:
                df_obj[col] = audit_align_to_zurich(df_obj[col])
    wf["date"] = wf["start"].dt.date
    wf["duration_min"] = (pd.to_datetime(wf["end"]) - pd.to_datetime(wf["start"])).dt.total_seconds() / 60.0
    wf["speed_kmh"] = audit_series_speed_kmh(wf["km"], wf["start"], wf["end"])

    rows_daily = []
    for (cid, day), grp in rda.dropna(subset=["collab_id", "jour", "start", "end"]).groupby(["collab_id", "jour"]):
        grp = grp.copy().sort_values(["start", "end"])
        starts = grp["start"].dropna()
        ends = grp["end"].dropna()
        if starts.empty or ends.empty:
            continue
        first_start = starts.min()
        last_end = ends.max()
        total_min = pd.to_numeric(grp["duree_min"], errors="coerce")
        fallback = (grp["end"] - grp["start"]).dt.total_seconds() / 60.0
        total_min = float(total_min.fillna(fallback).sum())
        blocks_tight = audit_merge_blocks(grp, gap_min=AUDIT_GAP_MERGE_MIN)
        block_cnt = len(blocks_tight)
        duty_span_min = float((pd.to_datetime(last_end) - pd.to_datetime(first_start)).total_seconds() / 60.0)
        time_full = total_min >= AUDIT_FULL_DAY_MINUTES
        span_full = block_cnt >= AUDIT_MIN_BLOCKS_FOR_SPAN and duty_span_min >= AUDIT_FULL_SPAN_MINUTES
        day_class = "Full" if (time_full or span_full) else "Half"
        main_blocks = audit_merge_blocks(grp, gap_min=AUDIT_INTERNAL_BLOCK_GAP_MIN)
        ibs = pd.NaT
        ibe = pd.NaT
        if len(main_blocks) >= 2:
            raw_ibs = main_blocks[0][1]
            raw_ibe = main_blocks[1][0]
            ibs = raw_ibs + pd.Timedelta(minutes=AUDIT_INTERNAL_BUFFER_MIN)
            ibe = raw_ibe - pd.Timedelta(minutes=AUDIT_INTERNAL_BUFFER_MIN)
            if pd.notna(ibs) and pd.notna(ibe) and ibe <= ibs:
                ibs = pd.NaT
                ibe = pd.NaT
        rows_daily.append({
            "collab_id": str(cid),
            "date": day,
            "rda_first_start": first_start,
            "rda_last_end": last_end,
            "rda_total_min": total_min,
            "rda_block_count": int(block_cnt),
            "duty_span_min": duty_span_min,
            "day_class": day_class,
            "buffer_start": first_start - pd.Timedelta(minutes=AUDIT_PRE_SHIFT_BUFFER_MIN) if AUDIT_CHECK_PRE_SHIFT else pd.NaT,
            "buffer_end": last_end + pd.Timedelta(minutes=AUDIT_WORK_END_BUFFER_MIN),
            "internal_buf_start": ibs,
            "internal_buf_end": ibe,
        })
    rda_daily = pd.DataFrame(rows_daily)
    for col in ["rda_first_start", "rda_last_end", "buffer_start", "buffer_end", "internal_buf_start", "internal_buf_end"]:
        if col in rda_daily.columns:
            rda_daily[col] = audit_align_to_zurich(rda_daily[col])

    _prog(0.55, "Annotation des trajets Webfleet...")

    wf = wf.copy()
    wf["collab_id"] = wf["collab_id"].astype(str)
    merge_cols = ["rda_first_start", "rda_last_end", "buffer_start", "buffer_end", "internal_buf_start", "internal_buf_end", "day_class"]
    wf = wf.drop(columns=[c for c in merge_cols if c in wf.columns], errors="ignore")
    if not rda_daily.empty:
        wf = wf.merge(rda_daily[["collab_id", "date"] + merge_cols], how="left", on=["collab_id", "date"])
    else:
        for c in merge_cols:
            wf[c] = np.nan

    wf["offday"] = wf["rda_first_start"].isna()
    wf["within_service"] = (~wf["offday"]) & (wf["start"] <= wf["rda_last_end"]) & (wf["end"] >= wf["rda_first_start"])
    wf["after_buffer"] = (~wf["buffer_end"].isna()) & (wf["start"] >= wf["buffer_end"])
    if AUDIT_CHECK_PRE_SHIFT:
        wf["before_shift"] = (~wf["buffer_start"].isna()) & (wf["end"] <= wf["buffer_start"])
    else:
        wf["before_shift"] = False

    def _interval_overlap(a_s, a_e, b_s, b_e):
        if pd.isna(a_s) or pd.isna(a_e) or pd.isna(b_s) or pd.isna(b_e):
            return False
        ov_s = max(a_s, b_s)
        ov_e = min(a_e, b_e)
        ov_min = (ov_e - ov_s).total_seconds() / 60.0
        if ov_min <= 0:
            return False
        if AUDIT_INTERNAL_BUF_MIN_OVERLAP_MIN > 0 and ov_min < AUDIT_INTERNAL_BUF_MIN_OVERLAP_MIN:
            return False
        if AUDIT_INTERNAL_BUF_MIN_OVERLAP_RATIO > 0:
            trip_min = (a_e - a_s).total_seconds() / 60.0
            if trip_min > 0 and (ov_min / trip_min) < AUDIT_INTERNAL_BUF_MIN_OVERLAP_RATIO:
                return False
        return True

    wf["in_internal_buffer"] = [
        _interval_overlap(s, e, ibs, ibe)
        for s, e, ibs, ibe in zip(wf["start"], wf["end"], wf["internal_buf_start"], wf["internal_buf_end"])
    ]

    def _build_flags(row):
        km = float(row["km"]) if pd.notna(row["km"]) else 0.0
        f = []
        if row["within_service"] and row["tripmode"] == 1 and km >= AUDIT_KM_MIN_FOR_FLAG_CONTEXT:
            f.append("MODE1_DURING_SERVICE")
        if row["after_buffer"] and row["tripmode"] in [2, 3] and km >= AUDIT_KM_MIN_FOR_FLAG_CONTEXT:
            f.append("MODE23_AFTER_BUFFER")
        if row["offday"] and row["tripmode"] in [2, 3] and km >= AUDIT_KM_MIN_FOR_FLAG_CONTEXT:
            f.append("MODE23_ON_OFFDAY")
            f.append("NO_RDA_BUT_MODE23")
        if row["before_shift"] and row["tripmode"] in [2, 3] and km >= AUDIT_KM_MIN_FOR_FLAG_CONTEXT:
            f.append("PRE_SHIFT_MODE2")
        if row["in_internal_buffer"] and row["tripmode"] in [2, 3] and km >= AUDIT_KM_MIN_FOR_FLAG_CONTEXT:
            f.append("MODE23_IN_INTERNAL_BUFFER")
        if pd.isna(row["collab_id"]) or row["collab_id"] in ("None", "nan"):
            f.append("UNMAPPED_DRIVER")
        if pd.notna(row.get("speed_kmh", np.nan)) and row["speed_kmh"] > AUDIT_MAX_REASONABLE_SPEED_KMH:
            f.append("SPEED_EXCEEDS_MAX")
        has_rda_today = not bool(row["offday"])
        outside_service = (not bool(row["within_service"])) and (bool(row["after_buffer"]) or bool(row["before_shift"]))
        if has_rda_today and outside_service and row["tripmode"] in [2, 3] and km >= AUDIT_KM_MIN_FOR_FLAG_CONTEXT:
            f.append("MODE23_OUTSIDE_SERVICE_ON_RDA_DAY")
        return ",".join(sorted(set(f)))

    wf["flags"] = wf.apply(_build_flags, axis=1)
    wf["suspect_private_km"] = np.where(
        wf["tripmode"].isin([2, 3]) & (wf["after_buffer"] | wf["offday"] | wf["before_shift"] | wf["in_internal_buffer"]),
        wf["km"].fillna(0.0),
        0.0,
    )

    _prog(0.63, "Labels d'usage journaliers...")

    rda_dates = rda[["collab_id", "jour"]].dropna().copy()
    rda_dates["collab_id"] = rda_dates["collab_id"].astype(str)
    rda_dates["date"] = rda_dates["jour"]
    rda_dates = rda_dates[["collab_id", "date"]]
    wf_dates = wf[["collab_id"]].copy()
    wf_dates["collab_id"] = wf_dates["collab_id"].astype(str)
    wf_dates["date"] = wf["start"].dt.date
    all_days = pd.concat([rda_dates, wf_dates], ignore_index=True).dropna(subset=["collab_id", "date"]).drop_duplicates()
    all_days["date"] = pd.to_datetime(all_days["date"]).dt.date

    wf_day_agg = wf.groupby(["collab_id", "date"]).agg(
        wf_trip_count=("tripid", "count"),
        wf_km_total=("km", "sum"),
        wf_any_within_service=("within_service", lambda s: bool(pd.Series(s).fillna(False).any())),
        wf_any_after_buffer=("after_buffer", lambda s: bool(pd.Series(s).fillna(False).any())),
        wf_any_before_shift=("before_shift", lambda s: bool(pd.Series(s).fillna(False).any())),
    ).reset_index()

    if not rda_daily.empty:
        ad = (all_days
              .merge(rda_daily[["collab_id", "date", "rda_first_start", "rda_last_end", "day_class"]], how="left", on=["collab_id", "date"])
              .merge(wf_day_agg, how="left", on=["collab_id", "date"]))
    else:
        ad = all_days.merge(wf_day_agg, how="left", on=["collab_id", "date"])
        ad["rda_first_start"] = pd.NaT
        ad["rda_last_end"] = pd.NaT
        ad["day_class"] = None

    ad["wf_trip_count"] = ad["wf_trip_count"].fillna(0).astype(int)
    ad["wf_km_total"] = ad["wf_km_total"].fillna(0.0).astype(float)
    for c in ["wf_any_within_service", "wf_any_after_buffer", "wf_any_before_shift"]:
        ad[c] = ad[c].fillna(False).astype(bool)
    ad["has_rda"] = ad["rda_first_start"].notna()
    ad["has_wf"] = ad["wf_trip_count"] > 0
    ad["car_used_within_service"] = ad["has_rda"] & ad["wf_any_within_service"]
    ad["car_used_outside_service"] = ad["has_rda"] & (ad["wf_any_after_buffer"] | ad["wf_any_before_shift"])

    def _usage_label(row):
        if row["has_rda"]:
            if not row["has_wf"]:
                return "WORK_NO_CAR"
            if row["car_used_within_service"]:
                return "WORK_CAR_IN_SERVICE"
            return "WORK_CAR_ONLY_OUTSIDE"
        return "NO_WORK_CAR_USED" if row["has_wf"] else "OFF_NO_CAR"

    ad["usage_label"] = ad.apply(_usage_label, axis=1)
    daily_usage = ad[["collab_id", "date", "has_rda", "has_wf", "car_used_within_service", "car_used_outside_service", "usage_label", "day_class"]].copy()

    _prog(0.70, "Vérifications 61010...")

    prestation61010_checks = pd.DataFrame()
    prestation61010_day_counts = pd.DataFrame()

    if AUDIT_ENABLE_61010_FEATURE:
        r61010 = rda[
            (rda["prestation_code"].astype(str) == str(AUDIT_PRESTATION_61010_CODE)) &
            rda["collab_id"].notna() & rda["start"].notna() & rda["end"].notna()
        ].copy()
        r61010["collab_id"] = r61010["collab_id"].astype(str)
        r61010["start"] = audit_series_to_local_naive(r61010["start"])
        r61010["end"] = audit_series_to_local_naive(r61010["end"])
        r61010 = r61010[r61010["start"].notna() & r61010["end"].notna() & (r61010["end"] > r61010["start"])].copy()

        wf_ok = wf[wf["collab_id"].notna() & wf["start"].notna() & wf["end"].notna()].copy()
        wf_ok["collab_id"] = wf_ok["collab_id"].astype(str)
        wf_ok["start"] = audit_series_to_local_naive(wf_ok["start"])
        wf_ok["end"] = audit_series_to_local_naive(wf_ok["end"])
        wf_ok = wf_ok[wf_ok["start"].notna() & wf_ok["end"].notna() & (wf_ok["end"] > wf_ok["start"])].copy()
        wf_groups = {cid: g.sort_values("start").copy() for cid, g in wf_ok.groupby("collab_id")}

        def _gap_min(tr_s, tr_e, win_s, win_e):
            if pd.isna(tr_s) or pd.isna(tr_e) or pd.isna(win_s) or pd.isna(win_e):
                return np.nan
            if tr_e < win_s:
                return (win_s - tr_e).total_seconds() / 60.0
            if tr_s > win_e:
                return (tr_s - win_e).total_seconds() / 60.0
            return 0.0

        def _overlap_min(a_s, a_e, b_s, b_e):
            if pd.isna(a_s) or pd.isna(a_e) or pd.isna(b_s) or pd.isna(b_e):
                return np.nan
            ov_s = max(a_s, b_s)
            ov_e = min(a_e, b_e)
            return max(0.0, (ov_e - ov_s).total_seconds() / 60.0)

        rows_61010 = []
        for rr in r61010.itertuples(index=False):
            cid = str(rr.collab_id)
            st, en = rr.start, rr.end
            buf_s = st - pd.Timedelta(minutes=AUDIT_PRESTATION_61010_BUFFER_MIN)
            buf_e = en + pd.Timedelta(minutes=AUDIT_PRESTATION_61010_BUFFER_MIN)
            trips = wf_groups.get(cid, pd.DataFrame())
            has_inside = False
            nearest_gap = np.nan
            nearest_tripid = None
            nearest_tripmode = None
            nearest_trip_start = pd.NaT
            nearest_trip_end = pd.NaT
            nearest_overlap_buf_min = np.nan
            nearest_overlap_61010_min = np.nan
            start_delta = np.nan
            end_delta = np.nan
            rda_dur = max(0.0, (en - st).total_seconds() / 60.0)
            uncovered = np.nan
            total_dev = np.nan
            if not trips.empty:
                cand = trips[(trips["start"] <= buf_e + pd.Timedelta(hours=12)) & (trips["end"] >= buf_s - pd.Timedelta(hours=12))].copy()
                if not cand.empty:
                    inside = (cand["start"] >= buf_s) & (cand["end"] <= buf_e)
                    nearest_row = None
                    if inside.any():
                        has_inside = True
                        nearest_gap = 0.0
                        nearest_row = cand.loc[inside].sort_values("start").iloc[0]
                    else:
                        cand["gap_min"] = [_gap_min(ts, te, buf_s, buf_e) for ts, te in zip(cand["start"], cand["end"])]
                        cand = cand.dropna(subset=["gap_min"]).sort_values(["gap_min", "start"])
                        if not cand.empty:
                            nearest_row = cand.iloc[0]
                            nearest_gap = float(nearest_row.get("gap_min", np.nan))
                    if nearest_row is not None:
                        nearest_tripid = str(nearest_row.get("tripid", ""))
                        nearest_tripmode = nearest_row.get("tripmode", None)
                        nearest_trip_start = nearest_row.get("start", pd.NaT)
                        nearest_trip_end = nearest_row.get("end", pd.NaT)
                        nearest_overlap_buf_min = _overlap_min(nearest_trip_start, nearest_trip_end, buf_s, buf_e)
                        nearest_overlap_61010_min = _overlap_min(nearest_trip_start, nearest_trip_end, st, en)
                        if pd.notna(nearest_trip_start):
                            start_delta = (st - nearest_trip_start).total_seconds() / 60.0
                        if pd.notna(nearest_trip_end):
                            end_delta = (en - nearest_trip_end).total_seconds() / 60.0
                        ov_for_calc = nearest_overlap_61010_min if pd.notna(nearest_overlap_61010_min) else 0.0
                        uncovered = max(0.0, rda_dur - ov_for_calc)
                        if pd.notna(start_delta) and pd.notna(end_delta):
                            total_dev = abs(start_delta) + abs(end_delta)
            if pd.isna(uncovered):
                uncovered = rda_dur
            if pd.isna(total_dev):
                total_dev = rda_dur
            rows_61010.append({
                "rda_row_id": int(rr.rda_row_id), "collab_id": cid,
                "date": pd.to_datetime(st).date(), "start": st, "end": en,
                "buf_start": buf_s, "buf_end": buf_e,
                "has_wf_trip_in_buffer": bool(has_inside), "nearest_gap_min": nearest_gap,
                "nearest_tripid": nearest_tripid, "nearest_tripmode": nearest_tripmode,
                "nearest_trip_start": nearest_trip_start, "nearest_trip_end": nearest_trip_end,
                "rda_61010_duration_min": rda_dur,
                "nearest_overlap_buf_min": nearest_overlap_buf_min,
                "nearest_overlap_61010_min": nearest_overlap_61010_min,
                "start_delta_rda_vs_wf_min": start_delta, "end_delta_rda_vs_wf_min": end_delta,
                "uncovered_61010_min": uncovered, "total_deviation_min": total_dev,
            })

        prestation61010_checks = pd.DataFrame(rows_61010)
        if not prestation61010_checks.empty:
            prestation61010_day_counts = (
                prestation61010_checks.groupby(["collab_id", "date"]).agg(
                    prestation_61010_count=("rda_row_id", "count"),
                    prestation_61010_missing=("has_wf_trip_in_buffer", lambda s: int((~pd.Series(s).astype(bool)).sum())),
                ).reset_index()
            )
        merge_back_cols = ["has_wf_trip_in_buffer", "buf_start", "buf_end", "rda_61010_duration_min",
                           "nearest_overlap_buf_min", "nearest_overlap_61010_min",
                           "start_delta_rda_vs_wf_min", "end_delta_rda_vs_wf_min",
                           "uncovered_61010_min", "total_deviation_min"]
        rda = rda.drop(columns=[c for c in merge_back_cols if c in rda.columns], errors="ignore")
        if not prestation61010_checks.empty:
            rda = rda.merge(prestation61010_checks[["rda_row_id"] + merge_back_cols], how="left", on="rda_row_id")
        else:
            for c in merge_back_cols:
                rda[c] = np.nan

    _prog(0.80, "Agrégats et statistiques collaborateurs...")

    wf2 = wf.copy()
    wf2["collab_id"] = wf2["collab_id"].astype(str)
    wf2["km"] = pd.to_numeric(wf2["km"], errors="coerce").fillna(0.0)
    wf2["km_m1"] = np.where(wf2["tripmode"] == 1, wf2["km"], 0.0)
    wf2["km_m2"] = np.where(wf2["tripmode"] == 2, wf2["km"], 0.0)
    wf2["km_m3"] = np.where(wf2["tripmode"] == 3, wf2["km"], 0.0)
    wf2["trips_m1"] = (wf2["tripmode"] == 1).astype(int)
    wf2["trips_m2"] = (wf2["tripmode"] == 2).astype(int)
    wf2["trips_m3"] = (wf2["tripmode"] == 3).astype(int)

    agg_daily = wf2.groupby(["collab_id", "date"], dropna=False).agg(
        wf_trip_count=("tripid", "count"),
        trips_m1=("trips_m1", "sum"),
        trips_m2=("trips_m2", "sum"),
        trips_m3=("trips_m3", "sum"),
        km_mode1=("km_m1", "sum"),
        km_mode2=("km_m2", "sum"),
        km_mode3=("km_m3", "sum"),
        suspect_private_km=("suspect_private_km", "sum"),
        flags_concat=("flags", audit_uniq_flags),
    ).reset_index()
    if not rda_daily.empty:
        agg_daily = agg_daily.merge(rda_daily, how="left", on=["collab_id", "date"])
    agg_daily["private_km_total_for_day"] = (agg_daily.get("km_mode1", 0).fillna(0) + agg_daily["suspect_private_km"].fillna(0)).round(3)
    agg_daily = agg_daily.merge(daily_usage[["collab_id", "date", "usage_label", "has_rda", "has_wf", "car_used_within_service", "car_used_outside_service"]], how="left", on=["collab_id", "date"])

    name_cols = map_df[["collab_id", "collab_name_sarl", "collab_no_sarl", "collab_name_wf", "driverno"]].drop_duplicates()
    name_cols["collab_id"] = name_cols["collab_id"].astype(str)
    agg_daily = agg_daily.merge(name_cols, how="left", on="collab_id")

    collab_stats = wf2.groupby("collab_id", dropna=False).agg(
        trips_total=("tripid", "count"),
        trips_m1=("trips_m1", "sum"),
        trips_m2=("trips_m2", "sum"),
        trips_m3=("trips_m3", "sum"),
        km_total=("km", "sum"),
        km_m1=("km_m1", "sum"),
        km_m2=("km_m2", "sum"),
        km_m3=("km_m3", "sum"),
        suspect_private_km=("suspect_private_km", "sum"),
        flags_concat=("flags", audit_uniq_flags),
    ).reset_index().merge(name_cols, how="left", on="collab_id")

    if not rda_daily.empty:
        day_counts = rda_daily.pivot_table(index="collab_id", columns="day_class", values="date", aggfunc="count").fillna(0).astype(int)
        day_counts = day_counts.rename(columns={"Full": "days_full", "Half": "days_half"})
        for c in ["days_full", "days_half"]:
            if c not in day_counts.columns:
                day_counts[c] = 0
        collab_stats = collab_stats.merge(day_counts.reset_index(), how="left", on="collab_id")
    else:
        collab_stats["days_full"] = 0
        collab_stats["days_half"] = 0

    usage_counts_pivot = agg_daily.groupby("collab_id")["usage_label"].value_counts().unstack(fill_value=0)
    for col in ["WORK_NO_CAR", "NO_WORK_CAR_USED", "WORK_CAR_IN_SERVICE", "WORK_CAR_ONLY_OUTSIDE", "OFF_NO_CAR"]:
        if col not in usage_counts_pivot.columns:
            usage_counts_pivot[col] = 0
    collab_stats = collab_stats.merge(usage_counts_pivot.reset_index(), how="left", on="collab_id")
    for col in ["WORK_NO_CAR", "NO_WORK_CAR_USED", "WORK_CAR_IN_SERVICE", "WORK_CAR_ONLY_OUTSIDE", "OFF_NO_CAR"]:
        collab_stats[col] = collab_stats[col].fillna(0).astype(int)

    base = agg_daily.copy()
    base["car_used_within_service"] = base.get("car_used_within_service", pd.Series(False)).fillna(False).astype(bool)
    full_wcar = base[(base.get("day_class", "") == "Full") & base["car_used_within_service"]].groupby("collab_id")["date"].nunique()
    half_wcar = base[(base.get("day_class", "") == "Half") & base["car_used_within_service"]].groupby("collab_id")["date"].nunique()
    collab_stats["full_days_with_car"] = collab_stats["collab_id"].map(full_wcar).fillna(0).astype(int)
    collab_stats["half_days_with_car"] = collab_stats["collab_id"].map(half_wcar).fillna(0).astype(int)
    collab_stats["work_days_total"] = (collab_stats["days_full"].fillna(0).astype(int) + collab_stats["days_half"].fillna(0).astype(int))

    if AUDIT_ENABLE_61010_FEATURE and not prestation61010_checks.empty:
        tot = prestation61010_checks.groupby("collab_id").agg(
            prestation_61010_total=("rda_row_id", "count"),
            prestation_61010_missing_total=("has_wf_trip_in_buffer", lambda s: int((~pd.Series(s).astype(bool)).sum())),
        ).reset_index()
        collab_stats = collab_stats.merge(tot, how="left", on="collab_id")
    else:
        collab_stats["prestation_61010_total"] = 0
        collab_stats["prestation_61010_missing_total"] = 0
    collab_stats["prestation_61010_total"] = collab_stats["prestation_61010_total"].fillna(0).astype(int)
    collab_stats["prestation_61010_missing_total"] = collab_stats["prestation_61010_missing_total"].fillna(0).astype(int)

    flags_exp = wf[["collab_id", "flags", "km"]].copy()
    flags_exp["flag"] = flags_exp["flags"].fillna("").astype(str).str.split(",")
    flags_exp = flags_exp.explode("flag")
    flags_exp = flags_exp[flags_exp["flag"].astype(str).str.strip() != ""]
    flag_summary = (
        flags_exp.groupby("flag").agg(count=("collab_id", "count"), km_sum=("km", "sum"))
        .reset_index().sort_values(["count", "km_sum"], ascending=[False, False])
    )

    _prog(0.90, "Export Excel...")

    output_dir = APP_ROOT / AUDIT_OUTPUT_FOLDER
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = output_dir / f"audit_{timestamp}.xlsx"

    excel_bytes = BytesIO()
    with pd.ExcelWriter(excel_bytes, engine="openpyxl") as xw:
        audit_drop_tz_excel_safe(collab_stats).to_excel(xw, index=False, sheet_name="Collaborator_Summary")
        audit_drop_tz_excel_safe(agg_daily).to_excel(xw, index=False, sheet_name="Daily_Aggregates")
        audit_drop_tz_excel_safe(wf2).to_excel(xw, index=False, sheet_name="All_Trips")
        audit_drop_tz_excel_safe(rda).to_excel(xw, index=False, sheet_name="All_RDA_Entries")
        if not rda_daily.empty:
            audit_drop_tz_excel_safe(rda_daily).to_excel(xw, index=False, sheet_name="RDA_Daily")
        map_df.to_excel(xw, index=False, sheet_name="Mapping")
        flag_summary.to_excel(xw, index=False, sheet_name="Flag_Summary")
        audit_drop_tz_excel_safe(planning).to_excel(xw, index=False, sheet_name="Planning")
        if AUDIT_ENABLE_61010_FEATURE and not prestation61010_checks.empty:
            audit_drop_tz_excel_safe(prestation61010_checks).to_excel(xw, index=False, sheet_name="Prestation_61010_Checks")
        if AUDIT_ENABLE_61010_FEATURE and not prestation61010_day_counts.empty:
            prestation61010_day_counts.to_excel(xw, index=False, sheet_name="Prestation_61010_DayCounts")
    excel_bytes.seek(0)
    with open(excel_path, "wb") as f:
        f.write(excel_bytes.getvalue())
    excel_bytes.seek(0)

    total_trips = len(wf)
    flagged = int((wf["flags"].fillna("").str.strip() != "").sum())
    flagged_pct = round(100.0 * flagged / total_trips, 1) if total_trips > 0 else 0.0
    suspect_km = float(wf["suspect_private_km"].fillna(0).sum())

    _prog(1.0, "Audit terminé")

    return {
        "rda": rda,
        "wf": wf,
        "wf2": wf2,
        "map_df": map_df,
        "planning": planning,
        "rda_daily": rda_daily,
        "agg_daily": agg_daily,
        "collab_stats": collab_stats,
        "daily_usage": daily_usage,
        "flag_summary": flag_summary,
        "prestation61010_checks": prestation61010_checks,
        "prestation61010_day_counts": prestation61010_day_counts,
        "excel_path": excel_path,
        "excel_bytes": excel_bytes,
        "metrics": {
            "total_trips": total_trips,
            "total_collabs": wf["collab_id"].nunique(),
            "flagged_trip_pct": flagged_pct,
            "suspect_km_total": suspect_km,
        },
    }


# ============================================================
# Audit — chart data builder (no matplotlib)
# ============================================================

def audit_build_chart_data(result: dict):
    """Build events DataFrame and lookup dicts needed for Gantt charts. No matplotlib dependency."""
    tz_name = AUDIT_TZ_NAME
    wf = result["wf"]
    rda = result["rda"]
    planning = result["planning"]
    map_df = result["map_df"]

    def _to_ln(ts):
        return audit_to_local_naive(ts, tz_name)

    def _safe_text(x):
        return "" if pd.isna(x) else str(x)

    def _cmp_date_str(value, fallback_ts=None):
        if value is not None and not pd.isna(value):
            d = pd.to_datetime(value, errors="coerce", dayfirst=True)
            if pd.notna(d):
                return pd.Timestamp(d).strftime("%Y-%m-%d")
        if fallback_ts is not None and not pd.isna(fallback_ts):
            fb = _to_ln(fallback_ts)
            if pd.notna(fb):
                return pd.Timestamp(fb).strftime("%Y-%m-%d")
        return None

    def _cmp_rda_orig_color(code):
        code = str(code).strip() if pd.notna(code) else ""
        if code in ["16009", "95900"]:
            return "#FFD700"
        if code == str(AUDIT_PRESTATION_61010_CODE):
            return "#800080"
        return "#2ca02c"

    def _cmp_planning_color(row_dict):
        if str(row_dict.get("client_absent", "")).strip().upper() == "Y":
            return "#000000"
        key = str(row_dict.get("event_color_key", "")).strip().lower()
        return AUDIT_PLAN_COLOR_MAP.get(key, row_dict.get("plot_color", "#bdbdbd"))

    LANE_Y = {"WF": 0.0, "RDA_ORIG": 1.0, "Planning": 2.0}

    map_local = map_df.copy()
    map_local["collab_id"] = map_local["collab_id"].astype(str)
    all_ids = set()
    for df_obj in [wf, rda, planning]:
        if not df_obj.empty and "collab_id" in df_obj.columns:
            all_ids.update(df_obj["collab_id"].dropna().astype(str).unique().tolist())
    collab_labels = {}
    for cid in sorted(all_ids):
        label = None
        if not map_local.empty:
            row = map_local[map_local["collab_id"] == cid]
            if not row.empty:
                rr = row.iloc[0]
                nm = _safe_text(rr.get("collab_name_sarl", "")).strip() or _safe_text(rr.get("collab_name_wf", "")).strip()
                sa = _safe_text(rr.get("collab_no_sarl", "")).strip() or "-"
                wf_no = _safe_text(rr.get("driverno", "")).strip() or "-"
                if nm:
                    label = f"{nm} (ID collab-{cid}, SA:{sa}, WF:{wf_no})"
        collab_labels[cid] = label or f"Collaborateur {cid}"

    rows_ev = []
    wf_local = wf.dropna(subset=["collab_id", "start", "end"]).copy()
    wf_local["collab_id"] = wf_local["collab_id"].astype(str)
    wf_stack = {}
    for rr in wf_local.sort_values(["collab_id", "start", "end"]).itertuples(index=False):
        s = _to_ln(rr.start)
        e = _to_ln(rr.end)
        if pd.isna(s) or pd.isna(e) or e <= s:
            continue
        cid = str(rr.collab_id)
        date_str = s.strftime("%Y-%m-%d")
        day_key = (cid, date_str)
        wf_idx = wf_stack.get(day_key, 0) + 1
        wf_stack[day_key] = wf_idx
        is_suspect = float(getattr(rr, "suspect_private_km", 0.0) or 0.0) > 0
        try:
            tm = int(rr.tripmode)
        except Exception:
            tm = -1
        fill_c = "#d62728" if is_suspect else {1: "#6c757d", 2: "#ff7f0e", 3: "#1f77b4"}.get(tm, "#999999")
        km_txt = f"{float(rr.km):.1f}km" if hasattr(rr, "km") and pd.notna(rr.km) else ""
        rows_ev.append({
            "collab_id": cid, "collab_label": collab_labels.get(cid, cid), "date_str": date_str,
            "kind": "WF", "y": LANE_Y["WF"], "left": s, "right": e, "mid": s + (e - s) / 2, "height": 0.34,
            "fill_color": fill_c, "line_color": "#7f0000" if is_suspect else "#202020",
            "label_text": f"{int(round(audit_duration_mins(s, e)))}m", "label_y": -0.20, "label_color": "#222222",
            "km_label": km_txt, "km_label_y": -0.38 if ((wf_idx - 1) % 2 == 0) else -0.48,
            "km_label_color": "#d62728" if is_suspect else "#111111", "wf_index": wf_idx,
        })

    rda_local = rda.dropna(subset=["collab_id", "start", "end"]).copy()
    rda_local["collab_id"] = rda_local["collab_id"].astype(str)
    for rr in rda_local.itertuples(index=False):
        s = _to_ln(rr.start)
        e = _to_ln(rr.end)
        if pd.isna(s) or pd.isna(e) or e <= s:
            continue
        cid = str(rr.collab_id)
        jour_value = getattr(rr, "jour", pd.NaT)
        date_str = _cmp_date_str(jour_value, fallback_ts=s)
        if not date_str:
            continue
        code = getattr(rr, "prestation_code", np.nan)
        is_61010 = str(code) == str(AUDIT_PRESTATION_61010_CODE)
        rows_ev.append({
            "collab_id": cid, "collab_label": collab_labels.get(cid, cid), "date_str": date_str,
            "kind": "RDA_ORIG", "y": LANE_Y["RDA_ORIG"], "left": s, "right": e, "mid": s + (e - s) / 2, "height": 0.34,
            "fill_color": _cmp_rda_orig_color(code), "line_color": "#202020",
            "label_text": f"{int(round(audit_duration_mins(s, e)))}m", "label_y": 0.80,
            "label_color": "#b30000" if is_61010 else "#111111",
            "km_label": "", "km_label_y": np.nan, "km_label_color": "#111111", "wf_index": np.nan,
        })

    plan_local = planning.dropna(subset=["collab_id", "start", "end"]).copy()
    plan_local["collab_id"] = plan_local["collab_id"].astype(str)
    for rr in plan_local.itertuples(index=False):
        s = _to_ln(rr.start)
        e = _to_ln(rr.end)
        if pd.isna(s) or pd.isna(e) or e <= s:
            continue
        cid = str(rr.collab_id)
        rr_dict = rr._asdict()
        plan_date_value = rr_dict.get("date_only", rr_dict.get("date", pd.NaT))
        date_str = _cmp_date_str(plan_date_value, fallback_ts=s)
        if not date_str:
            continue
        rows_ev.append({
            "collab_id": cid, "collab_label": collab_labels.get(cid, cid), "date_str": date_str,
            "kind": "Planning", "y": LANE_Y["Planning"], "left": s, "right": e, "mid": s + (e - s) / 2, "height": 0.34,
            "fill_color": _cmp_planning_color(rr_dict), "line_color": "#4A3B33",
            "label_text": f"{int(round(audit_duration_mins(s, e)))}m", "label_y": 1.80, "label_color": "#111111",
            "km_label": "", "km_label_y": np.nan, "km_label_color": "#111111", "wf_index": np.nan,
        })

    if not rows_ev:
        return None

    events_cmp = pd.DataFrame(rows_ev).sort_values(["collab_label", "date_str", "y", "left", "right"]).reset_index(drop=True)

    def _build_rda_rest_lookup():
        lookup = {}
        rd = rda.copy()
        rd = rd[rd["collab_id"].notna() & rd["start"].notna() & rd["end"].notna()].copy()
        if rd.empty:
            return lookup
        rd["collab_id"] = rd["collab_id"].astype(str)
        rd["start"] = rd["start"].apply(_to_ln)
        rd["end"] = rd["end"].apply(_to_ln)
        rd = rd[rd["start"].notna() & rd["end"].notna() & (rd["end"] > rd["start"])].copy()
        if "jour" in rd.columns:
            rd["date_str"] = rd.apply(lambda r: _cmp_date_str(r.get("jour", pd.NaT), fallback_ts=r["start"]), axis=1)
        else:
            rd["date_str"] = rd["start"].apply(lambda x: pd.Timestamp(x).strftime("%Y-%m-%d"))
        rd = rd[rd["date_str"].notna()].copy()
        daily = (rd.groupby(["collab_id", "date_str"]).agg(day_first_rda_start=("start", "min"), day_last_rda_end=("end", "max")).reset_index().sort_values(["collab_id", "date_str"]))
        for cid, grp in daily.groupby("collab_id"):
            grp = grp.sort_values("date_str").reset_index(drop=True)
            for i, row in grp.iterrows():
                cur_s = row["day_first_rda_start"]
                cur_e = row["day_last_rda_end"]
                prev_e = grp.loc[i - 1, "day_last_rda_end"] if i > 0 else pd.NaT
                next_s = grp.loc[i + 1, "day_first_rda_start"] if i < len(grp) - 1 else pd.NaT
                prev_rest = max(0.0, audit_duration_mins(prev_e, cur_s)) if pd.notna(prev_e) and pd.notna(cur_s) else np.nan
                next_rest = max(0.0, audit_duration_mins(cur_e, next_s)) if pd.notna(cur_e) and pd.notna(next_s) else np.nan
                lookup[(str(cid), str(row["date_str"]))] = {
                    "cur_start": cur_s, "cur_end": cur_e,
                    "prev_end": prev_e, "next_start": next_s,
                    "prev_rest_min": prev_rest, "next_rest_min": next_rest,
                }
        return lookup

    rda_rest_lookup = _build_rda_rest_lookup()
    return {"events_cmp": events_cmp, "collab_labels": collab_labels, "rda_rest_lookup": rda_rest_lookup}


# ============================================================
# Audit — single-day figure builder (requires matplotlib)
# ============================================================

def audit_build_day_fig(data_ctx: dict, result: dict, cid: str, date_str: str):
    """Render the Gantt chart for one collaborator/day. Returns matplotlib Figure or None."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
    except ImportError:
        return None

    tz_name = AUDIT_TZ_NAME
    events_cmp = data_ctx["events_cmp"]
    rda_rest_lookup = data_ctx["rda_rest_lookup"]
    collab_labels = data_ctx["collab_labels"]
    agg_daily = result["agg_daily"]
    FIG_W, FIG_H = 19.0, 9.4

    def _to_ln(ts):
        return audit_to_local_naive(ts, tz_name)

    def _month_km_stats(c, d):
        out = {"month_label": "-", "km_private": 0.0, "km_professional": 0.0, "km_commute": 0.0,
               "km_total": 0.0, "km_suspect_private": 0.0, "km_private_plus_suspect": 0.0}
        if agg_daily.empty:
            return out
        base_day = pd.Timestamp(str(d))
        period = base_day.to_period("M")
        out["month_label"] = base_day.strftime("%B %Y")
        sub = agg_daily.copy()
        sub["collab_id"] = sub["collab_id"].astype(str)
        sub["date"] = pd.to_datetime(sub["date"], errors="coerce")
        sub = sub[(sub["collab_id"] == str(c)) & (sub["date"].dt.to_period("M") == period)]
        for key, col in [("km_private", "km_mode1"), ("km_professional", "km_mode2"), ("km_commute", "km_mode3"),
                         ("km_suspect_private", "suspect_private_km"), ("km_private_plus_suspect", "private_km_total_for_day")]:
            out[key] = float(sub[col].fillna(0).sum()) if col in sub.columns else 0.0
        out["km_total"] = out["km_private"] + out["km_professional"] + out["km_commute"]
        return out

    def _fmt_dt_short(ts):
        ts = _to_ln(ts)
        return "-" if pd.isna(ts) else pd.Timestamp(ts).strftime("%d/%m %H:%M")

    def _rest_texts(c, d):
        info = rda_rest_lookup.get((str(c), str(d)))
        if not info:
            return "REST BEFORE\n-", "REST AFTER\n-"
        pr = info["prev_rest_min"]
        nr = info["next_rest_min"]
        prev_text = ("REST BEFORE\n" + f"{_fmt_dt_short(info['prev_end'])} → {_fmt_dt_short(info['cur_start'])}\n{audit_fmt_hours(pd.Timedelta(minutes=float(pr)))}"
                     if pd.notna(pr) else "REST BEFORE\nAucun jour RDA précédent")
        next_text = ("REST AFTER\n" + f"{_fmt_dt_short(info['cur_end'])} → {_fmt_dt_short(info['next_start'])}\n{audit_fmt_hours(pd.Timedelta(minutes=float(nr)))}"
                     if pd.notna(nr) else "REST AFTER\nAucun jour RDA suivant")
        return prev_text, next_text

    def _wf_all_text(day_events):
        sub = day_events[day_events["kind"] == "WF"].sort_values(["left", "right"]).reset_index(drop=True)
        if sub.empty:
            return "ALL WF TRIPS\n-"
        lines = []
        for _, row in sub.iterrows():
            idx_txt = f"{int(row['wf_index']):02d}" if pd.notna(row.get("wf_index")) else "--"
            s = audit_fmt_hhmm(row["left"])
            e = audit_fmt_hhmm(row["right"])
            mins = audit_duration_mins(row["left"], row["right"])
            mins_txt = f"{int(round(mins))}m" if pd.notna(mins) else "-"
            km_txt = f" | {str(row.get('km_label', '')).strip()}" if str(row.get("km_label", "")).strip() else ""
            lines.append(f"{idx_txt}. {s} → {e} ({mins_txt}){km_txt}")
        return "ALL WF TRIPS\n" + "\n".join(lines)

    def _day_bounds(day_events, day_str, min_hours=16):
        if day_events.empty:
            base = pd.Timestamp(f"{day_str} 00:00:00")
            return base, base + pd.Timedelta(hours=min_hours)
        left = _to_ln(day_events["left"].min())
        right = _to_ln(day_events["right"].max())
        if pd.isna(left) or pd.isna(right) or right <= left:
            base = pd.Timestamp(f"{day_str} 00:00:00")
            return base, base + pd.Timedelta(hours=min_hours)
        if (right - left) < pd.Timedelta(hours=min_hours):
            right = left + pd.Timedelta(hours=min_hours)
        return left, right

    def _lane_span(day_events, kind):
        sub = day_events[day_events["kind"] == kind]
        if sub.empty:
            return pd.NaT, pd.NaT
        s = _to_ln(sub["left"].min())
        e = _to_ln(sub["right"].max())
        return (s, e) if (pd.notna(s) and pd.notna(e) and e > s) else (pd.NaT, pd.NaT)

    def _draw_markers(ax, day_events):
        sub = day_events[day_events["kind"] == "RDA_ORIG"]
        if sub.empty:
            return
        rs = _to_ln(sub["left"].min())
        re_ = _to_ln(sub["right"].max())
        if pd.isna(rs) or pd.isna(re_) or re_ <= rs:
            return
        ymin, ymax = -0.36, 2.36
        for x, lab in [(rs, f"Start {rs.strftime('%H:%M')}"), (re_, f"End {re_.strftime('%H:%M')}")]:
            xx = mdates.date2num(x)
            ax.vlines(xx, ymin=ymin, ymax=ymax, colors="#555555", linestyles="--", linewidth=2.0, alpha=0.90, zorder=1)
            ax.text(xx, 2.52, lab, ha="center", va="bottom", fontsize=11.5, color="#555555", zorder=6, clip_on=False)
        span_hours = max(0.0, (re_ - rs).total_seconds() / 3600.0)
        for h in range(1, int(math.floor(span_hours)) + 2):
            x = rs + pd.Timedelta(hours=h)
            xx = mdates.date2num(x)
            ax.vlines(xx, ymin=ymin, ymax=ymax, colors="#9a9a9a", linestyles=":", linewidth=1.2, alpha=0.80, zorder=1)
            ax.text(xx, 2.45, f"{h}h", ha="center", va="bottom", fontsize=10, color="#6f6f6f", zorder=6, clip_on=False)

    def _draw_day(ax, day_events):
        for _, rr in day_events.iterrows():
            left = _to_ln(rr["left"])
            right = _to_ln(rr["right"])
            if pd.isna(left) or pd.isna(right) or right <= left:
                continue
            y = float(rr["y"])
            h = float(rr["height"])
            ax.broken_barh(
                [(mdates.date2num(left), mdates.date2num(right) - mdates.date2num(left))],
                (y - h / 2.0, h),
                facecolors=rr["fill_color"], edgecolors=rr["line_color"], linewidth=1.0, alpha=0.95, zorder=3,
            )
            mid = _to_ln(rr["mid"])
            lab = rr["label_text"]
            if pd.notna(mid) and str(lab).strip():
                ax.text(mdates.date2num(mid), rr["label_y"], lab, ha="center", va="center", fontsize=8.5,
                        color=rr["label_color"], zorder=6, clip_on=False, bbox=dict(facecolor="white", edgecolor="none", alpha=0.55, pad=0.18))
            if str(rr["kind"]) == "WF":
                wf_idx = rr.get("wf_index", np.nan)
                if pd.notna(mid) and pd.notna(wf_idx):
                    ax.text(mdates.date2num(mid), float(rr["y"]) + 0.26, f"{int(wf_idx)}", ha="center", va="center",
                            fontsize=8.3, color="#000000", fontweight="bold", zorder=7, clip_on=False,
                            bbox=dict(facecolor="white", edgecolor="#222222", alpha=0.88, boxstyle="round,pad=0.18"))
            km_lab = str(rr.get("km_label", "") or "").strip()
            km_y = rr.get("km_label_y", np.nan)
            if pd.notna(mid) and km_lab and pd.notna(km_y):
                ax.text(mdates.date2num(mid), km_y, km_lab, ha="center", va="center", fontsize=8.2,
                        color=rr["km_label_color"], zorder=6, clip_on=False, bbox=dict(facecolor="white", edgecolor="none", alpha=0.55, pad=0.15))

    collab_label = collab_labels.get(str(cid), f"Collaborateur {cid}")
    day_events = events_cmp[
        (events_cmp["collab_id"].astype(str) == str(cid)) &
        (events_cmp["date_str"].astype(str) == str(date_str))
    ].copy()
    if day_events.empty:
        return None
    day_events = day_events.sort_values(["y", "left", "right"]).reset_index(drop=True)
    left_bound, right_bound = _day_bounds(day_events, date_str)
    wf_s, wf_e = _lane_span(day_events, "WF")
    ro_s, ro_e = _lane_span(day_events, "RDA_ORIG")
    pl_s, pl_e = _lane_span(day_events, "Planning")
    km_stats = _month_km_stats(cid, date_str)
    km_line = (f"Month KM ({km_stats['month_label']})    |    Private: {km_stats['km_private']:.1f}km"
               f"    |    Suspect: {km_stats['km_suspect_private']:.1f}km"
               f"    |    Private+Suspect: {km_stats['km_private_plus_suspect']:.1f}km"
               f"    |    Dom-travail: {km_stats['km_commute']:.1f}km"
               f"    |    Professionnel: {km_stats['km_professional']:.1f}km"
               f"    |    Total: {km_stats['km_total']:.1f}km")
    wf_all_text = _wf_all_text(day_events)
    wf_count = int((day_events["kind"] == "WF").sum())
    wf_fs = 8.4 if wf_count <= 12 else 7.3 if wf_count <= 25 else 6.4 if wf_count <= 40 else 5.5 if wf_count <= 60 else 4.8
    prev_rest_text, next_rest_text = _rest_texts(cid, date_str)
    subtitle = (f"Frame: {audit_fmt_hhmm(left_bound)} → {audit_fmt_hhmm(right_bound)}"
                f"    |    WF span: {audit_fmt_span(wf_s, wf_e)}"
                f"    |    RDA original: {audit_fmt_span(ro_s, ro_e)}"
                f"    |    Planning: {audit_fmt_span(pl_s, pl_e)}")
    fig, ax = plt.subplots(figsize=(FIG_W, FIG_H), constrained_layout=False)
    _draw_markers(ax, day_events)
    _draw_day(ax, day_events)
    ax.set_xlim(mdates.date2num(left_bound), mdates.date2num(right_bound))
    ax.set_ylim(-0.52, 2.72)
    ax.margins(x=0, y=0)
    ax.set_yticks([0.0, 1.0, 2.0])
    ax.set_yticklabels(["WF Trips", "RDA Original", "Planning"], fontsize=11.5)
    span_hours = max(1.0, (right_bound - left_bound).total_seconds() / 3600.0)
    major_interval = 1 if span_hours <= 14 else 2 if span_hours <= 22 else 3
    ax.xaxis.set_major_locator(mdates.HourLocator(interval=major_interval))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    ax.tick_params(axis="x", labelsize=11)
    ax.grid(axis="x", linestyle=":", linewidth=0.8, alpha=0.45)
    ax.grid(axis="y", visible=False)
    for sp in ax.spines.values():
        sp.set_linewidth(0.8)
        sp.set_alpha(0.70)
    fig.suptitle(f"{collab_label} | {date_str}", fontsize=14, y=0.985, fontweight="bold")
    fig.text(0.5, 0.948, km_line, ha="center", va="center", fontsize=10.5, fontweight="bold",
             bbox=dict(facecolor="#f2f2f2", edgecolor="#d0d0d0", boxstyle="round,pad=0.30"))
    fig.text(0.5, 0.915, subtitle, ha="center", va="center", fontsize=9.2, color="#333333")
    fig.text(0.105, 0.875, prev_rest_text, ha="left", va="top", fontsize=8.7, family="monospace", color="#222222",
             bbox=dict(facecolor="#ffffff", edgecolor="#cfcfcf", boxstyle="round,pad=0.35", alpha=0.96))
    fig.text(0.775, 0.875, next_rest_text, ha="right", va="top", fontsize=8.7, family="monospace", color="#222222",
             bbox=dict(facecolor="#ffffff", edgecolor="#cfcfcf", boxstyle="round,pad=0.35", alpha=0.96))
    fig.text(0.805, 0.775, wf_all_text, ha="left", va="top", fontsize=wf_fs, family="monospace", color="#222222",
             bbox=dict(facecolor="#ffffff", edgecolor="#cfcfcf", boxstyle="round,pad=0.45", alpha=0.96))
    fig.subplots_adjust(left=0.10, right=0.775, top=0.79, bottom=0.12)
    return fig


# ============================================================
# Audit — PDF generation (optional, requires matplotlib)
# ============================================================

def audit_generate_pdfs(result: dict, progress_cb=None):
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_pdf import PdfPages
    except ImportError:
        st.warning("matplotlib n'est pas installé. Installez-le avec : pip install matplotlib")
        return None

    data_ctx = audit_build_chart_data(result)
    if data_ctx is None:
        st.warning("Aucune donnée à exporter en PDF.")
        return None

    events_cmp = data_ctx["events_cmp"]
    pairs = events_cmp[["collab_id", "collab_label", "date_str"]].drop_duplicates().copy()
    pairs["collab_id"] = pairs["collab_id"].astype(str)
    pairs["date_str"] = pairs["date_str"].astype(str)
    pairs = pairs.sort_values(["collab_label", "date_str"]).reset_index(drop=True)

    collab_groups = list(pairs.groupby("collab_label", sort=True))
    total_collabs = len(collab_groups)

    def _prog(i, label, pages=None):
        if not progress_cb:
            return
        pct = min(1.0, i / total_collabs) if total_collabs > 0 else 1.0
        pages_txt = f" — {pages} page(s)" if pages is not None else ""
        progress_cb(pct, f"PDF {i}/{total_collabs} : {label}{pages_txt}")

    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for idx, (collab_label, grp) in enumerate(collab_groups):
            _prog(idx, collab_label)
            grp = grp.sort_values("date_str").reset_index(drop=True)
            cid_values = grp["collab_id"].dropna().astype(str).unique().tolist()
            cid = cid_values[0] if cid_values else "unknown"
            pdf_name = f"{audit_safe_filename(collab_label)}__collab_{cid}__wf_rda_planning.pdf"
            pdf_buf = BytesIO()
            pages_written = 0
            with PdfPages(pdf_buf) as pdf:
                for _, rr in grp.iterrows():
                    fig = audit_build_day_fig(data_ctx, result, str(rr["collab_id"]), str(rr["date_str"]))
                    if fig is None:
                        continue
                    pdf.savefig(fig, dpi=180, bbox_inches="tight")
                    plt.close(fig)
                    pages_written += 1
            _prog(idx + 1, collab_label, pages=pages_written)
            pdf_buf.seek(0)
            zf.writestr(pdf_name, pdf_buf.read())

    zip_buf.seek(0)
    return zip_buf


# ============================================================
# Audit — dashboard render
# ============================================================

def render_audit_dashboard(result: dict) -> None:
    metrics = result["metrics"]
    collab_stats = result["collab_stats"]
    flag_summary = result["flag_summary"]
    daily_usage = result["daily_usage"]
    agg_daily = result["agg_daily"]
    wf_df = result["wf"]

    # --- Metrics row ---
    m_cols = st.columns(4)
    m_cols[0].metric("Trajets WF total", f"{metrics['total_trips']:,}")
    m_cols[1].metric("Collaborateurs", f"{metrics['total_collabs']:,}")
    m_cols[2].metric("% Trajets flaggés", f"{metrics['flagged_trip_pct']:.1f}%")
    m_cols[3].metric("KM suspects total", f"{metrics['suspect_km_total']:,.1f} km")

    st.divider()

    # --- Downloads + PDF generation ---
    dl_col, pdf_col = st.columns([1, 2])
    with dl_col:
        if result["excel_path"].exists():
            with open(result["excel_path"], "rb") as f:
                st.download_button(
                    "Télécharger le rapport Excel", f,
                    file_name=result["excel_path"].name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
    with pdf_col:
        p1, p2 = st.columns(2)
        with p1:
            if st.button("Générer les PDFs Gantt", key="audit_gen_pdf", use_container_width=True):
                progress = st.progress(0.0, text="Initialisation des PDFs...")
                try:
                    zip_bytes = audit_generate_pdfs(result, progress_cb=progress.progress)
                    progress.empty()
                    if zip_bytes:
                        st.session_state["latest_audit_pdf_zip"] = zip_bytes
                except Exception as exc:
                    progress.empty()
                    st.exception(exc)
        with p2:
            pdf_zip = st.session_state.get("latest_audit_pdf_zip")
            if pdf_zip:
                pdf_zip.seek(0)
                st.download_button(
                    "Télécharger les PDFs (zip)", pdf_zip,
                    file_name="audit_pdfs_gantt.zip", mime="application/zip",
                    use_container_width=True,
                )

    # --- In-UI Gantt viewer ---
    chart_data = st.session_state.get("latest_audit_chart_data")
    if chart_data is None:
        chart_data = audit_build_chart_data(result)
        if chart_data is not None:
            st.session_state["latest_audit_chart_data"] = chart_data

    if chart_data is not None:
        events_cmp = chart_data["events_cmp"]
        collab_options = (
            events_cmp[["collab_id", "collab_label"]]
            .drop_duplicates()
            .sort_values("collab_label")
            .reset_index(drop=True)
        )
        collab_display = [str(r["collab_label"]) for _, r in collab_options.iterrows()]
        collab_id_map = {str(r["collab_label"]): str(r["collab_id"]) for _, r in collab_options.iterrows()}

        with st.expander("Visualiser un graphique Gantt", expanded=False):
            vc1, vc2, vc3 = st.columns([3, 2, 1])
            with vc1:
                sel_collab_label = st.selectbox("Collaborateur", collab_display, key="audit_chart_collab")
            sel_cid = collab_id_map.get(sel_collab_label or "", "")
            dates_for_collab = sorted(
                events_cmp[events_cmp["collab_id"].astype(str) == sel_cid]["date_str"].unique().tolist()
            ) if sel_cid else []
            with vc2:
                sel_date = st.selectbox("Date", dates_for_collab, key="audit_chart_date")
            with vc3:
                st.write("")
                st.write("")
                show_chart_btn = st.button("Afficher", key="audit_show_chart", use_container_width=True)
            if show_chart_btn and sel_cid and sel_date:
                with st.spinner("Génération du graphique..."):
                    try:
                        import matplotlib.pyplot as plt
                        fig = audit_build_day_fig(chart_data, result, sel_cid, sel_date)
                        if fig:
                            img_buf = BytesIO()
                            fig.savefig(img_buf, format="png", dpi=120, bbox_inches="tight")
                            plt.close(fig)
                            img_buf.seek(0)
                            st.session_state["audit_gantt_img"] = {
                                "key": (sel_cid, sel_date),
                                "data": img_buf.getvalue(),
                            }
                        else:
                            st.session_state["audit_gantt_img"] = None
                            st.info("Aucune donnée pour ce collaborateur et cette date.")
                    except Exception as exc:
                        st.exception(exc)

            gantt_img = st.session_state.get("audit_gantt_img")
            if gantt_img and gantt_img.get("key") == (sel_cid, sel_date):
                st.image(gantt_img["data"], use_container_width=True)
            elif gantt_img and gantt_img.get("key") != (sel_cid, sel_date):
                st.caption("Sélection modifiée — cliquez « Afficher » pour régénérer.")

    st.divider()

    # --- Charts section (dataframe-based to avoid sticky Vega-Lite tooltips) ---
    with st.expander("Graphiques de synthèse", expanded=True):
        chart_cols = st.columns(2)
        with chart_cols[0]:
            st.subheader("Répartition usage quotidien")
            if not daily_usage.empty and "usage_label" in daily_usage.columns:
                usage_counts = daily_usage["usage_label"].value_counts().reset_index()
                usage_counts.columns = ["Usage", "Jours"]
                usage_max = int(usage_counts["Jours"].max()) if not usage_counts.empty else 1
                st.dataframe(
                    usage_counts,
                    column_config={"Jours": st.column_config.ProgressColumn("Jours", min_value=0, max_value=usage_max, format="%d")},
                    hide_index=True, use_container_width=True,
                )
            else:
                st.info("Aucune donnée d'usage.")

        with chart_cols[1]:
            st.subheader("Répartition des flags")
            if not flag_summary.empty and "flag" in flag_summary.columns:
                flag_df = flag_summary[["flag", "count"]].copy()
                flag_df.columns = ["Flag", "Occurrences"]
                flag_max = int(flag_df["Occurrences"].max()) if not flag_df.empty else 1
                st.dataframe(
                    flag_df,
                    column_config={"Occurrences": st.column_config.ProgressColumn("Occurrences", min_value=0, max_value=flag_max, format="%d")},
                    hide_index=True, use_container_width=True,
                )
            else:
                st.info("Aucun flag détecté.")

        chart_cols2 = st.columns(2)
        with chart_cols2[0]:
            st.subheader("KM par mode (top 15 collaborateurs)")
            if not collab_stats.empty:
                top15 = collab_stats.nlargest(15, "km_total")[["collab_id", "km_m1", "km_m2", "km_m3", "km_total"]].copy()
                top15.columns = ["Collaborateur", "Privé (m1)", "Pro (m2)", "Dom-travail (m3)", "Total"]
                km_max = int(top15["Total"].max()) if not top15.empty else 1
                st.dataframe(
                    top15,
                    column_config={
                        "Privé (m1)": st.column_config.ProgressColumn("Privé (m1)", min_value=0, max_value=km_max, format="%.1f km"),
                        "Pro (m2)": st.column_config.ProgressColumn("Pro (m2)", min_value=0, max_value=km_max, format="%.1f km"),
                        "Dom-travail (m3)": st.column_config.ProgressColumn("Dom-travail (m3)", min_value=0, max_value=km_max, format="%.1f km"),
                        "Total": st.column_config.NumberColumn("Total km", format="%.1f km"),
                    },
                    hide_index=True, use_container_width=True,
                )
            else:
                st.info("Aucune donnée.")

        with chart_cols2[1]:
            st.subheader("Top 15 collaborateurs — KM suspects")
            if not collab_stats.empty and "suspect_private_km" in collab_stats.columns:
                top_susp = collab_stats.nlargest(15, "suspect_private_km")[["collab_id", "suspect_private_km"]].copy()
                top_susp.columns = ["Collaborateur", "KM suspects"]
                susp_max = float(top_susp["KM suspects"].max()) if not top_susp.empty else 1.0
                st.dataframe(
                    top_susp,
                    column_config={"KM suspects": st.column_config.ProgressColumn("KM suspects", min_value=0, max_value=susp_max, format="%.1f km")},
                    hide_index=True, use_container_width=True,
                )
            else:
                st.info("Aucune donnée.")

        if not agg_daily.empty and "date" in agg_daily.columns and "suspect_private_km" in agg_daily.columns:
            st.subheader("Évolution KM suspects par jour")
            trend = agg_daily.copy()
            trend["date"] = pd.to_datetime(trend["date"], errors="coerce")
            trend = (trend.dropna(subset=["date"])
                     .groupby("date")["suspect_private_km"].sum()
                     .reset_index().sort_values("date"))
            trend.columns = ["Date", "KM suspects"]
            trend_max = float(trend["KM suspects"].max()) if not trend.empty else 1.0
            st.dataframe(
                trend,
                column_config={
                    "Date": st.column_config.DateColumn("Date"),
                    "KM suspects": st.column_config.ProgressColumn("KM suspects", min_value=0, max_value=trend_max, format="%.1f km"),
                },
                hide_index=True, use_container_width=True,
            )

    # --- Data tabs ---
    tab_names = ["Résumé Collaborateurs", "Agrégats Journaliers", "Tous les Trajets", "Entrées RDA", "Planning"]
    tabs = st.tabs(tab_names)

    with tabs[0]:
        f0c1, f0c2 = st.columns([3, 1])
        with f0c1:
            search_collab = st.text_input("Rechercher", key="audit_collab_search", placeholder="ID ou nom...")
        with f0c2:
            only_suspect_cs = st.checkbox("Suspects uniquement", key="audit_collab_only_suspect")
        cs_display = collab_stats.copy()
        if search_collab.strip():
            mask = cs_display.apply(lambda r: search_collab.lower() in str(r).lower(), axis=1)
            cs_display = cs_display[mask]
        if only_suspect_cs and "suspect_private_km" in cs_display.columns:
            cs_display = cs_display[cs_display["suspect_private_km"] > 0]
        st.caption(f"{len(cs_display):,} collaborateurs")
        st.dataframe(cs_display, use_container_width=True, hide_index=True)

    with tabs[1]:
        agg = agg_daily.copy()
        f1c1, f1c2, f1c3 = st.columns(3)
        with f1c1:
            agg_collabs = sorted(agg["collab_id"].dropna().astype(str).unique().tolist()) if not agg.empty and "collab_id" in agg.columns else []
            sel_agg_collabs = st.multiselect("Collaborateur(s)", agg_collabs, key="audit_agg_collab_filter")
        with f1c2:
            date_range_agg = None
            if not agg.empty and "date" in agg.columns:
                agg_dates = pd.to_datetime(agg["date"], errors="coerce").dropna()
                if not agg_dates.empty:
                    min_d, max_d = agg_dates.min().date(), agg_dates.max().date()
                    date_range_agg = st.date_input("Période", value=(min_d, max_d), min_value=min_d, max_value=max_d, key="audit_agg_date_range")
        with f1c3:
            only_suspect_agg = st.checkbox("Suspects uniquement", key="audit_agg_only_suspect")
        if sel_agg_collabs:
            agg = agg[agg["collab_id"].astype(str).isin(sel_agg_collabs)]
        if date_range_agg and len(date_range_agg) == 2:
            agg_dc = pd.to_datetime(agg["date"], errors="coerce")
            agg = agg[(agg_dc.dt.date >= date_range_agg[0]) & (agg_dc.dt.date <= date_range_agg[1])]
        if only_suspect_agg and "suspect_private_km" in agg.columns:
            agg = agg[agg["suspect_private_km"] > 0]
        st.caption(f"{len(agg):,} entrées")
        st.dataframe(agg, use_container_width=True, hide_index=True)

    with tabs[2]:
        wf_filtered = wf_df.copy()
        f2c1, f2c2, f2c3 = st.columns(3)
        with f2c1:
            if "flags" in wf_df.columns:
                all_flags = sorted(set(
                    f.strip()
                    for flags_str in wf_df["flags"].fillna("").astype(str)
                    for f in flags_str.split(",")
                    if f.strip()
                ))
            else:
                all_flags = []
            sel_flags = st.multiselect("Flag(s)", all_flags, key="audit_trip_flag_filter")
        with f2c2:
            trip_collabs = sorted(wf_df["collab_id"].dropna().astype(str).unique().tolist()) if "collab_id" in wf_df.columns else []
            sel_trip_collabs = st.multiselect("Collaborateur(s)", trip_collabs, key="audit_trip_collab_filter")
        with f2c3:
            only_suspect_trips = st.checkbox("Suspects uniquement", key="audit_trip_only_suspect")
        if sel_flags:
            wf_filtered = wf_filtered[wf_filtered["flags"].fillna("").apply(lambda x: any(f in x.split(",") for f in sel_flags))]
        if sel_trip_collabs:
            wf_filtered = wf_filtered[wf_filtered["collab_id"].astype(str).isin(sel_trip_collabs)]
        if only_suspect_trips and "suspect_private_km" in wf_filtered.columns:
            wf_filtered = wf_filtered[wf_filtered["suspect_private_km"] > 0]
        st.caption(f"{len(wf_filtered):,} trajets")
        st.dataframe(audit_drop_tz_excel_safe(wf_filtered), use_container_width=True, hide_index=True)

    with tabs[3]:
        rda_f = result["rda"].copy()
        rda_collabs = sorted(rda_f["collab_id"].dropna().astype(str).unique().tolist()) if "collab_id" in rda_f.columns else []
        sel_rda_collabs = st.multiselect("Collaborateur(s)", rda_collabs, key="audit_rda_collab_filter")
        if sel_rda_collabs:
            rda_f = rda_f[rda_f["collab_id"].astype(str).isin(sel_rda_collabs)]
        st.caption(f"{len(rda_f):,} entrées")
        st.dataframe(audit_drop_tz_excel_safe(rda_f), use_container_width=True, hide_index=True)

    with tabs[4]:
        plan_f = result["planning"].copy()
        plan_collabs = sorted(plan_f["collab_id"].dropna().astype(str).unique().tolist()) if "collab_id" in plan_f.columns else []
        sel_plan_collabs = st.multiselect("Collaborateur(s)", plan_collabs, key="audit_plan_collab_filter")
        if sel_plan_collabs:
            plan_f = plan_f[plan_f["collab_id"].astype(str).isin(sel_plan_collabs)]
        st.caption(f"{len(plan_f):,} entrées")
        st.dataframe(audit_drop_tz_excel_safe(plan_f), use_container_width=True, hide_index=True)


# ============================================================
# Audit — task entry point
# ============================================================

def render_audit_task() -> None:
    st.title("Audit Webfleet-RDA")
    st.caption("Croise les données Webfleet, RDA et planning pour détecter les usages suspects du véhicule de fonction.")

    upload_cols = st.columns(4)
    rda_file = upload_cols[0].file_uploader("Fichier RDA", type=["xlsx", "xls", "csv"], key="audit_rda")
    wf_file = upload_cols[1].file_uploader("Fichier Webfleet", type=["xlsx", "xls", "csv"], key="audit_wf")
    mapping_file = upload_cols[2].file_uploader("Fichier Mapping", type=["xlsx", "xls"], key="audit_map")
    planning_file = upload_cols[3].file_uploader("Fichier Planning", type=["xlsx", "xls"], key="audit_plan")

    all_uploaded = all([rda_file, wf_file, mapping_file, planning_file])

    if st.button("Lancer l'audit", type="primary", disabled=not all_uploaded):
        progress = st.progress(0.0, text="Démarrage de l'audit...")
        try:
            result = audit_process(rda_file, wf_file, mapping_file, planning_file, progress_cb=progress.progress)
            progress.progress(1.0, text="Audit terminé")
            st.session_state["latest_audit_result"] = result
            st.session_state.pop("latest_audit_pdf_zip", None)
            st.session_state.pop("latest_audit_chart_data", None)
            st.session_state.pop("audit_gantt_img", None)
        except Exception as exc:
            progress.empty()
            st.exception(exc)
            return

    result = st.session_state.get("latest_audit_result")
    if result:
        st.success(f"Rapport Excel créé : {result['excel_path']}")
        render_audit_dashboard(result)


def render_placeholder_task(title: str) -> None:
    st.title(title)
    st.info("Cette tâche est préparée pour la prochaine passe d'intégration. Le notebook a été revu ; sa logique sera déplacée dans des fonctions appelables sans modifier les entrées ni les sorties.")


def inject_app_css() -> None:
    colors = {
        "teal": "#31b6a7",
        "teal_dark": "#188f83",
        "orange": "#f59b45",
        "yellow": "#ffc400",
        "danger": "#e45f4f",
    }
    st.markdown(
        f"""
        <style>
        :root {{
            --ha-bg: var(--background-color);
            --ha-surface: var(--secondary-background-color);
            --ha-surface-2: var(--background-color);
            --ha-text: var(--text-color);
            --ha-muted: color-mix(in srgb, var(--text-color) 64%, transparent);
            --ha-border: color-mix(in srgb, var(--text-color) 16%, transparent);
            --ha-teal: {colors["teal"]};
            --ha-teal-dark: {colors["teal_dark"]};
            --ha-orange: {colors["orange"]};
            --ha-yellow: {colors["yellow"]};
            --ha-danger: {colors["danger"]};
        }}
        .stApp {{
            background: var(--ha-bg);
            color: var(--ha-text);
        }}
        [data-testid="stHeader"] {{
            background: color-mix(in srgb, var(--ha-bg) 86%, transparent);
        }}
        .block-container {{
            padding-top: 1.35rem;
        }}
        h1, h2, h3, h4, h5, h6, p, label, span {{
            color: var(--ha-text);
        }}
        h1 {{
            letter-spacing: 0;
            font-weight: 800;
        }}
        h2, h3 {{
            font-weight: 750;
        }}
        [data-testid="stMarkdownContainer"] p,
        [data-testid="stCaptionContainer"] {{
            color: var(--ha-muted);
        }}
        div[data-testid="stMetric"],
        div[data-testid="stExpander"],
        [data-testid="stDataFrame"],
        [data-testid="stTabs"] {{
            border-color: var(--ha-border);
        }}
        div[data-testid="stMetric"] {{
            background: var(--ha-surface);
            border: 1px solid var(--ha-border);
            border-radius: 8px;
            padding: 0.75rem 0.9rem;
        }}
        div[data-testid="stMetric"] label {{
            color: var(--ha-muted);
            font-weight: 700;
        }}
        div[data-testid="stMetricValue"] {{
            color: var(--ha-text);
        }}
        div[data-testid="stExpander"] {{
            background: var(--ha-surface);
            border-radius: 8px;
        }}
        .stButton > button,
        .stDownloadButton > button {{
            background: var(--ha-teal);
            color: white;
            border: 1px solid var(--ha-teal);
            border-radius: 6px;
            font-weight: 800;
            letter-spacing: 0;
            min-height: 2.4rem;
            box-shadow: 0 8px 18px rgba(49, 182, 167, 0.18);
        }}
        .stButton > button:hover,
        .stDownloadButton > button:hover {{
            background: var(--ha-teal-dark);
            border-color: var(--ha-teal-dark);
            color: white;
        }}
        .stButton > button[kind="primary"] {{
            background: var(--ha-orange);
            border-color: var(--ha-orange);
            box-shadow: 0 8px 18px rgba(245, 155, 69, 0.20);
        }}
        .stTextInput input,
        .stNumberInput input,
        .stDateInput input,
        .stTextArea textarea,
        div[data-baseweb="select"] > div,
        div[data-baseweb="base-input"] {{
            background: var(--ha-surface);
            color: var(--ha-text);
            border-color: var(--ha-border);
            border-radius: 6px;
        }}
        .stRadio,
        .stCheckbox,
        .stToggle {{
            color: var(--ha-text);
        }}
        div[role="radiogroup"] label[data-baseweb="radio"] > div:first-child {{
            border-color: var(--ha-teal);
        }}
        .stTabs [data-baseweb="tab-list"] {{
            gap: 0.25rem;
            border-bottom: 1px solid var(--ha-border);
        }}
        .stTabs [data-baseweb="tab"] {{
            color: var(--ha-muted);
            border-radius: 6px 6px 0 0;
            font-weight: 700;
        }}
        .stTabs [aria-selected="true"] {{
            color: var(--ha-teal);
            border-bottom: 3px solid var(--ha-orange);
        }}
        .task-panel {{
            position: sticky;
            top: 1rem;
            border: 1px solid var(--ha-border);
            border-radius: 8px;
            padding: 0.85rem 0.9rem;
            background: var(--ha-surface);
            box-shadow: 0 10px 26px rgba(0, 0, 0, 0.08);
        }}
        .task-panel h3 {{
            font-size: 1rem;
            margin: 0 0 0.5rem 0;
            color: var(--ha-text);
        }}
        .brand-strip {{
            height: 4px;
            border-radius: 999px;
            background: linear-gradient(90deg, var(--ha-teal), var(--ha-yellow), var(--ha-orange));
            margin: 0.35rem 0 0.9rem;
        }}
        .brand-logo-mark {{
            display: inline-flex;
            align-items: center;
            gap: 0.35rem;
            font-weight: 900;
            color: var(--ha-text);
            margin-bottom: 0.35rem;
        }}
        .brand-logo-mark::before {{
            content: "";
            display: inline-block;
            width: 0.55rem;
            height: 1.65rem;
            background: var(--ha-teal);
            border-radius: 2px;
            box-shadow: 0.72rem 0 0 var(--ha-orange);
            margin-right: 0.75rem;
        }}
        .element-container:has(.stAlert) {{
            color: var(--ha-text);
        }}
        [data-testid="stDataFrame"] {{
            background: var(--ha-surface);
            border-radius: 8px;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def main() -> None:
    st.set_page_config(page_title="Outils Webfleet", layout="wide")
    inject_app_css()

    main_col, task_col = st.columns([5, 1.35], gap="large")

    with task_col:
        st.markdown('<div class="task-panel"><div class="brand-logo-mark">HOME ASSISTANCE</div><div class="brand-strip"></div><h3>Tâche</h3></div>', unsafe_allow_html=True)
        task_labels = list(TASKS.values())
        current_label = st.radio(
            "Sélectionner une tâche",
            task_labels,
            label_visibility="collapsed",
            key="selected_task_label",
        )

    selected_key = next(key for key, label in TASKS.items() if label == current_label)

    with main_col:
        if selected_key == "webfleet":
            render_webfleet_task()
        elif selected_key == "merge":
            render_merge_task()
        elif selected_key == "rda":
            render_rda_task()
        elif selected_key == "ltr":
            render_ltr_task()
        elif selected_key == "audit":
            render_audit_task()


if __name__ == "__main__":
    main()

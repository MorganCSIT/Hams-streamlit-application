from app_config import *
from ui_common import read_any_flex, render_blocking_run_warning, render_download_or_placeholder, safe_folder_name

def ltr_unique_output_root(output_name: str) -> Path:
    safe_name = safe_folder_name(output_name) if output_name.strip() else f"LTR_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    root = get_session_output_root(LTR_OUTPUT_FOLDER) / safe_name
    if root.exists():
        root = root.with_name(f"{root.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    root.mkdir(parents=True, exist_ok=True)
    return root


def ltr_workbook_path(output_root: Path) -> Path:
    return output_root / "multiple" / f"{output_root.name}.xlsx"


def ltr_week_workbook_path(output_root: Path, start_date: date, end_date: date) -> Path:
    return output_root / "week" / f"{output_root.name}_week_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx"


def ltr_package_path(output_root: Path, start_date: date, end_date: date) -> Path:
    return output_root / f"{output_root.name}_full_and_week_{start_date.isoformat()}_to_{end_date.isoformat()}.zip"


def ltr_save_upload(uploaded_file, target_dir: Path) -> Path:
    target_dir.mkdir(parents=True, exist_ok=True)
    path = target_dir / uploaded_file.name
    path.write_bytes(uploaded_file.getvalue())
    return path


def ltr_previous_full_week(today: date | None = None) -> tuple[date, date]:
    today = today or date.today()
    this_monday = today - timedelta(days=today.weekday())
    previous_sunday = this_monday - timedelta(days=1)
    previous_monday = previous_sunday - timedelta(days=6)
    return previous_monday, previous_sunday


class LtrPytzCompat:
    @staticmethod
    def timezone(name: str):
        return ZoneInfo(name)


@st.cache_resource(show_spinner=False)
def ltr_load_notebook_functions(notebook_mtime: float) -> dict:
    _ = notebook_mtime
    notebook_path = LTR_NOTEBOOK_PATH
    if not notebook_path.exists():
        raise FileNotFoundError(f"LTR notebook not found: {notebook_path}")

    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    env = {
        "__builtins__": __builtins__,
        "os": os,
        "re": re,
        "np": np,
        "pd": pd,
        "pytz": LtrPytzCompat,
        "Path": Path,
        "datetime": datetime,
        "timedelta": timedelta,
        "dtime": dtime,
        "load_workbook": load_workbook,
        "Alignment": Alignment,
        "get_column_letter": get_column_letter,
        "SWISS_TZ": ZoneInfo("Europe/Zurich"),
        "MAX_SINGLE_INTERVAL_HOURS": 24,
        "SERVICE_RESET_GAP_HOURS": 8.0,
        "WEEKLY_LIMIT_HOURS": 50.0,
        "SPAN_LIMIT_HOURS": 14.0,
        "REST_NORMAL_HOURS": 11.0,
        "REST_ABSOLUTE_MIN_HOURS": 8.0,
        "REQUIRE_AVG_FOR_REDUCED_REST": True,
        "RUN_OVER50H": True,
        "RUN_STREAK": True,
        "RUN_SPAN": True,
        "RUN_REST11": True,
        "RUN_BREAKS": True,
        "COUNT_TRANSPORT_AS_WORK_FOR_50H": True,
        "COUNT_TRANSPORT_AS_WORK_FOR_STREAK": True,
        "COUNT_TRANSPORT_AS_WORK_FOR_BREAKS": True,
        "COUNT_TRANSPORT_FOR_SERVICE_BOUNDARY": True,
        "PAUSE_CODES": {"16009", "95900"},
        "TRANSPORT_CODES": {"61800", "61010"},
        "EXCLUDE_PRESTATIONS": {"196", "60041"},
        "PSEUDO_NON_WORK_CODES": {"195"},
    }

    notebook = json.loads(notebook_path.read_text(encoding="utf-8"))
    for cell_index in range(2, 9):
        source = "".join(notebook["cells"][cell_index]["source"])
        exec(compile(source, f"{notebook_path.name}:cell{cell_index + 1}", "exec"), env)
    return env


def ltr_notebook_mtime() -> float:
    notebook_path = LTR_NOTEBOOK_PATH
    if not notebook_path.exists():
        raise FileNotFoundError(f"LTR notebook not found: {notebook_path}")
    return notebook_path.stat().st_mtime


def ltr_compute_summary(env: dict, df: pd.DataFrame, services_df: pd.DataFrame, calendar_slices_df: pd.DataFrame, all_infractions: pd.DataFrame, data_quality: pd.DataFrame) -> pd.DataFrame:
    raw_months = set(df["Mois"].dropna().astype(str).tolist()) if "Mois" in df.columns else set()
    service_months = set(services_df["service_month"].dropna().astype(str).tolist()) if not services_df.empty else set()
    calendar_months = set(calendar_slices_df["target_month"].dropna().astype(str).tolist()) if not calendar_slices_df.empty else set()
    inf_months = set(all_infractions["TARGET_MONTH"].dropna().astype(str).tolist()) if not all_infractions.empty else set()
    months = sorted(raw_months | service_months | calendar_months | inf_months)

    rows = []
    for month in months:
        row = {"TARGET_MONTH": month}
        for rule in [
            f"OVER_{int(env['WEEKLY_LIMIT_HOURS'])}H_WEEK",
            "STREAK_7DAYS",
            f"SPAN_OVER_{int(env['SPAN_LIMIT_HOURS'])}H",
            "REST_UNDER_11H",
            "PAUSE_INSUFF",
        ]:
            if all_infractions.empty:
                row[rule] = 0
            else:
                row[rule] = int(
                    (
                        all_infractions.get("TARGET_MONTH", pd.Series(dtype=str)).astype(str).eq(month)
                        & all_infractions.get("RULE", pd.Series(dtype=str)).astype(str).eq(rule)
                    ).sum()
                )
        row["TOTAL_INFRACTIONS"] = int(all_infractions["TARGET_MONTH"].astype(str).eq(month).sum()) if not all_infractions.empty else 0
        row["SERVICES_STARTED"] = int(services_df["service_month"].astype(str).eq(month).sum()) if not services_df.empty else 0
        row["CALENDAR_HOURS"] = float(calendar_slices_df.loc[calendar_slices_df["target_month"].astype(str).eq(month), "hours"].sum()) if not calendar_slices_df.empty else 0.0
        row["ORPHAN_PAUSE_ROWS"] = int(
            (
                data_quality.get("TARGET_MONTH", pd.Series(dtype=str)).astype(str).eq(month)
                & data_quality.get("QUALITY_TYPE", pd.Series(dtype=str)).astype(str).eq("ORPHAN_PAUSE_NO_SERVICE")
            ).sum()
        ) if not data_quality.empty else 0
        row["INVALID_INTERVAL_ROWS"] = int(
            (
                data_quality.get("TARGET_MONTH", pd.Series(dtype=str)).astype(str).eq(month)
                & data_quality.get("QUALITY_TYPE", pd.Series(dtype=str)).astype(str).eq("INVALID_INTERVAL")
            ).sum()
        ) if not data_quality.empty else 0
        rows.append(row)
    return pd.DataFrame(rows)


def ltr_pick_first_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    return next((col for col in candidates if col in df.columns), None)


def ltr_pick_client_id_column(df: pd.DataFrame) -> str | None:
    explicit_candidates = [
        "N° du client",
        "No client",
        "ID client",
        "Client ID",
        "Numero client",
        "Numéro client",
        "KD-Nr",
        "KD_Nr",
    ]
    for candidate in explicit_candidates:
        if candidate in df.columns:
            return candidate

    scored_columns = []
    for col in df.columns:
        normalized = str(col).strip().lower()
        normalized_ascii = re.sub(r"[^a-z0-9]+", "", normalized)
        if normalized in {"kd-nr", "kd_nr"} or normalized_ascii in {"kdnr", "idclient", "clientid"}:
            return col
        if "client" in normalized_ascii and (
            normalized_ascii.startswith("n") or "no" in normalized_ascii or "id" in normalized_ascii
        ):
            values = df[col].apply(lambda value: ltr_norm_with_env({}, value))
            numeric_count = values.astype(str).str.fullmatch(r"\d+").sum()
            scored_columns.append((numeric_count, col))

    if scored_columns:
        numeric_count, col = max(scored_columns, key=lambda item: item[0])
        if numeric_count > 0:
            return col

    if "Client" in df.columns:
        values = df["Client"].apply(lambda value: ltr_norm_with_env({}, value))
        if values.astype(str).str.fullmatch(r"\d+").sum() > 0:
            return "Client"

    return None


def ltr_norm_with_env(env: dict, value) -> str:
    norm_id = env.get("_norm_id")
    if norm_id is not None:
        return norm_id(value)
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"", "nan", "none", "nat", "<na>"}:
        return ""
    text = re.sub(r"\.0+$", "", text)
    digits = re.sub(r"\D+", "", text)
    return digits if digits else text


def ltr_clean_report_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"", "nan", "none", "nat", "<na>"} else text


def ltr_branch_hint(env: dict, df: pd.DataFrame) -> pd.Series:
    find_oe = env.get("_find_oe_column")
    branch_from_oe = env.get("_branch_hint_from_oe")
    oe_col = find_oe(df.columns) if find_oe is not None else None
    if oe_col is not None and branch_from_oe is not None:
        return df[oe_col].apply(branch_from_oe).fillna("").astype(str)
    return pd.Series([""] * len(df), index=df.index, dtype=object)


def ltr_unmatched_collaborator_report(unrecognized_summary: pd.DataFrame) -> pd.DataFrame:
    columns = ["Entity Type", "Name", "ID", "Match Status", "Row Count"]
    if unrecognized_summary is None or unrecognized_summary.empty:
        return pd.DataFrame(columns=columns)
    report = pd.DataFrame({
        "Entity Type": "Collaborateur",
        "Name": unrecognized_summary.get("Collaborateur", pd.Series(dtype=object)).map(ltr_clean_report_text),
        "ID": unrecognized_summary.get("No collaborateur", pd.Series(dtype=object)).map(ltr_clean_report_text),
        "Match Status": unrecognized_summary.get("collab_match_status", pd.Series(dtype=object)).map(ltr_clean_report_text),
        "Row Count": pd.to_numeric(unrecognized_summary.get("Row Count", 0), errors="coerce").fillna(0).astype(int),
    })
    report["Match Status"] = report["Match Status"].replace({"": "UNMATCHED"})
    return report[columns]


def ltr_load_matched_clients(path: Path) -> pd.DataFrame:
    try:
        clients = pd.read_excel(path, sheet_name="Matched Clients", dtype=object)
    except Exception:
        return pd.DataFrame()
    clients.columns = clients.columns.astype(str).str.strip().str.lstrip("\ufeff")
    normalized_cols = {str(col).strip().lower() for col in clients.columns}
    if "client-id" not in normalized_cols or not any(col.startswith("no-client-") for col in normalized_cols):
        return pd.DataFrame()
    return clients


def ltr_client_match_lookup(env: dict, matched_clients: pd.DataFrame) -> dict:
    if matched_clients is None or matched_clients.empty:
        return {"known_ids": set()}

    client_no_cols = [col for col in matched_clients.columns if str(col).strip().lower().startswith("no-client-")]
    known_ids = set()
    for col in client_no_cols:
        known_ids.update(
            value
            for value in matched_clients[col].apply(lambda item: ltr_norm_with_env(env, item)).tolist()
            if re.fullmatch(r"\d+", value or "")
        )

    return {"known_ids": known_ids}


def ltr_unmatched_client_report(env: dict, df: pd.DataFrame, matched_path: Path) -> pd.DataFrame:
    columns = ["Entity Type", "Name", "ID", "Match Status", "Row Count"]
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)

    id_col = ltr_pick_client_id_column(df)
    if id_col is None:
        return pd.DataFrame(columns=columns)

    name_col = ltr_pick_first_column(df, ["Client", "Klient", "Nom client", "client_name"])
    if name_col == id_col:
        name_col = None
    client_ids = df[id_col].apply(lambda value: ltr_norm_with_env(env, value))
    client_names = df[name_col].map(ltr_clean_report_text) if name_col is not None else pd.Series([""] * len(df), index=df.index)
    lookup = ltr_client_match_lookup(env, ltr_load_matched_clients(matched_path))
    known_ids = lookup["known_ids"]
    if not known_ids:
        return pd.DataFrame(columns=columns)

    work = pd.DataFrame({
        "Entity Type": "Client",
        "Name": client_names,
        "ID": client_ids,
        "Match Status": "UNMATCHED",
    })
    work = work[
        work["ID"].astype(str).str.fullmatch(r"\d+")
        & ~work["ID"].astype(str).isin(known_ids)
    ].copy()
    if work.empty:
        return pd.DataFrame(columns=columns)
    return (
        work.groupby(["Entity Type", "Name", "ID", "Match Status"], dropna=False)
        .size()
        .reset_index(name="Row Count")
        .sort_values(["Entity Type", "Row Count", "Name", "ID"], ascending=[True, False, True, True])
        .reset_index(drop=True)
    )[columns]


def ltr_unmatched_mapping_report(env: dict, df: pd.DataFrame, matched_path: Path, unrecognized_summary: pd.DataFrame) -> pd.DataFrame:
    columns = ["Entity Type", "Name", "ID", "Match Status", "Row Count"]
    parts = [
        ltr_unmatched_collaborator_report(unrecognized_summary),
        ltr_unmatched_client_report(env, df, matched_path),
    ]
    parts = [part for part in parts if part is not None and not part.empty]
    if not parts:
        return pd.DataFrame(columns=columns)
    report = pd.concat(parts, ignore_index=True, sort=False)
    report["Row Count"] = pd.to_numeric(report["Row Count"], errors="coerce").fillna(0).astype(int)
    return report.sort_values(["Entity Type", "Row Count", "Name", "ID"], ascending=[True, False, True, True]).reset_index(drop=True)


def ltr_write_workbook(env: dict, workbook_path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    orders = {
        "ALL_INFRACTIONS": ["TARGET_MONTH", "EVENT_DATE", "RULE", "SEVERITY", "Collaborateur", "No_collaborateur_codes", "Match_status", "DETAIL", "service_id", "service_date", "service_start", "service_end"],
        "SERVICES_AUDIT": ["service_month", "service_date", "service_start", "service_end", "continues_after_midnight", "continuation_row_count", "creates_worked_day", "Collaborateur", "No_collaborateur_codes", "Match_status", "amplitude_hours", "net_50h_minutes", "net_breaks_minutes", "pause_minutes_inside_service", "attached_calendar_dates", "service_id"],
        "CALENDAR_HOUR_SLICES": ["target_month", "calendar_date", "week_monday", "slice_start", "slice_end", "minutes", "hours", "continuation_from_previous_service", "service_date", "Collaborateur", "No_collaborateur_codes", "service_id", "note"],
        "DATA_QUALITY": ["TARGET_MONTH", "EVENT_DATE", "QUALITY_TYPE", "DETAIL", "Collaborateur", "No collaborateur", "No prestation", "Prestation", "start_dt_local", "end_dt_local", "interval_status", "service_id", "service_date", "continuation_from_previous_service"],
        "UNMATCHED_MAPPING": ["Entity Type", "Name", "ID", "Match Status", "Row Count"],
    }
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="w") as writer:
        for sheet_name, df_sheet in sheets.items():
            if sheet_name == "UNMATCHED_MAPPING":
                clean = env["excel_safe_no_tz"](df_sheet)
            else:
                clean = env["excel_safe_no_tz"](env["prep_for_export"](df_sheet, drop_collaborateur_id=True))
            clean = env["reorder_columns"](clean, orders.get(sheet_name, []))
            clean = env["ensure_unique_columns"](clean)
            clean.to_excel(writer, index=False, sheet_name=sheet_name[:31])
    env["apply_swiss_formats_xlsx"](str(workbook_path))


def ltr_date_filtered_sheet(df: pd.DataFrame, start_date: date, end_date: date) -> pd.DataFrame:
    if df is None or df.empty:
        return df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()

    date_candidates = [
        "EVENT_DATE",
        "event_date",
        "service_date",
        "calendar_date",
        "week_monday",
        "TARGET_DATE",
        "start_dt_local",
        "end_dt_local",
        "service_start",
        "service_end",
        "slice_start",
        "slice_end",
    ]
    date_col = next((col for col in date_candidates if col in df.columns), None)
    if date_col is None:
        return df.copy()

    dates = pd.to_datetime(df[date_col], errors="coerce", dayfirst=True).dt.date
    mask = dates.apply(lambda value: pd.notna(value) and start_date <= value <= end_date)
    return df.loc[mask].copy()


def ltr_week_label_from_date(value) -> str:
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return ""
    day = pd.Timestamp(parsed).date()
    monday = day - timedelta(days=day.weekday())
    sunday = monday + timedelta(days=6)
    return f"{monday.isoformat()} to {sunday.isoformat()}"


def ltr_week_labels_for_df(df: pd.DataFrame, date_cols: list[str]) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype=object)
    date_col = next((col for col in date_cols if col in df.columns), None)
    if date_col is None:
        return pd.Series([""] * len(df), index=df.index, dtype=object)
    return df[date_col].apply(ltr_week_label_from_date)


def ltr_weekly_summary_sheet(filtered_infractions: pd.DataFrame, start_date: date, end_date: date) -> pd.DataFrame:
    rows = [
        {"Metric": "Date début", "Value": start_date.isoformat()},
        {"Metric": "Date fin", "Value": end_date.isoformat()},
        {"Metric": "Infractions", "Value": int(len(filtered_infractions))},
    ]
    if filtered_infractions is not None and not filtered_infractions.empty:
        if "Collaborateur" in filtered_infractions.columns:
            rows.append({"Metric": "Collaborateurs concernés", "Value": int(filtered_infractions["Collaborateur"].nunique())})
        if "RULE" in filtered_infractions.columns:
            for rule, count in filtered_infractions.groupby("RULE", dropna=False).size().sort_values(ascending=False).items():
                rows.append({"Metric": f"Règle: {rule}", "Value": int(count)})
    return pd.DataFrame(rows)


def ltr_build_week_workbook(result: dict, start_date: date, end_date: date) -> Path:
    output_root = Path(result["output_root"])
    workbook_path = ltr_week_workbook_path(output_root, start_date, end_date)
    filtered_sheets = {}
    for sheet_name, df_sheet in result.get("sheets", {}).items():
        if sheet_name == "SUMMARY_BY_MONTH":
            continue
        filtered_sheets[sheet_name] = ltr_date_filtered_sheet(df_sheet, start_date, end_date)

    weekly_infractions = filtered_sheets.get("ALL_INFRACTIONS", pd.DataFrame())
    filtered_sheets = {"SUMMARY_SELECTED_WEEK": ltr_weekly_summary_sheet(weekly_infractions, start_date, end_date), **filtered_sheets}

    env = ltr_load_notebook_functions(ltr_notebook_mtime())
    ltr_write_workbook(env, workbook_path, filtered_sheets)
    return workbook_path


def ltr_zip_full_and_week(result: dict, week_workbook_path: Path, start_date: date, end_date: date) -> Path:
    output_root = Path(result["output_root"])
    package_path = ltr_package_path(output_root, start_date, end_date)
    package_path.parent.mkdir(parents=True, exist_ok=True)
    full_workbook_path = Path(result["workbook_path"])
    with zipfile.ZipFile(package_path, "w", zipfile.ZIP_DEFLATED) as archive:
        if full_workbook_path.is_file():
            archive.write(full_workbook_path, arcname=f"Full_Range/{full_workbook_path.name}")
        if week_workbook_path.is_file():
            archive.write(week_workbook_path, arcname=f"Selected_Week/{week_workbook_path.name}")
    return package_path


def ltr_process(matched_upload, rda_upload, output_name: str) -> dict:
    output_root = ltr_unique_output_root(output_name)
    input_dir = output_root / "inputs"
    matched_path = ltr_save_upload(matched_upload, input_dir)
    rda_path = ltr_save_upload(rda_upload, input_dir)
    workbook_path = ltr_workbook_path(output_root)

    env = ltr_load_notebook_functions(ltr_notebook_mtime())

    df_raw = env["load_and_normalize"](str(rda_path))
    matched = env["load_matched_collabs"](str(matched_path))
    df_raw = env["attach_collab_master"](df_raw, matched)
    df = env["add_interval_columns"](df_raw)
    services_df, df_tagged, orphan_pauses_df = env["build_services_and_tagged_rows"](df)
    calendar_slices_df = env["build_calendar_slices"](services_df)

    over50_detail, over50_all = env["check_over_50h"](calendar_slices_df)
    streak_detail, streak_all = env["check_streak_7days"](services_df)
    span_detail, span_all = env["check_span_over_14h"](services_df)
    rest_detail, rest_all, rest_review = env["check_rest_under_11h"](services_df)
    breaks_detail, breaks_all, breaks_audit = env["check_breaks"](services_df)
    data_quality = env["build_data_quality"](df_tagged, orphan_pauses_df)

    infraction_parts = [part for part in [over50_all, streak_all, span_all, rest_all, breaks_all] if part is not None and not part.empty]
    if infraction_parts:
        all_infractions = pd.concat(infraction_parts, ignore_index=True, sort=False)
        all_infractions = env["ensure_unique_columns"](all_infractions)
        sort_cols = [col for col in ["TARGET_MONTH", "EVENT_DATE", "Collaborateur", "RULE"] if col in all_infractions.columns]
        if sort_cols:
            all_infractions = all_infractions.sort_values(sort_cols, kind="stable").reset_index(drop=True)
    else:
        all_infractions = pd.DataFrame(columns=["TARGET_MONTH", "EVENT_DATE", "RULE", "SEVERITY", "DETAIL", "collab_uid", "Collaborateur"])

    summary_by_month = ltr_compute_summary(env, df, services_df, calendar_slices_df, all_infractions, data_quality)
    services_audit = services_df.drop(columns=[col for col in ["_net_50_intervals", "_net_breaks_intervals"] if col in services_df.columns]).copy()
    unmatched_mask = df["collab_key"].astype(str).str.startswith("UNMATCHED")
    ambig_mask = df["collab_match_status"].astype(str).eq("AMBIG_MATCH")
    unmatched_df = df[unmatched_mask | ambig_mask].copy()
    if unmatched_df.empty:
        unrecognized_summary = pd.DataFrame(columns=["Collaborateur", "No collaborateur", "collab_match_status", "collab_key", "Row Count"])
        unrecognized_rows = pd.DataFrame()
    else:
        unrecognized_rows = unmatched_df.copy()
        if "start_dt_local" in unrecognized_rows.columns:
            unrecognized_rows["TARGET_MONTH"] = pd.to_datetime(unrecognized_rows["start_dt_local"], errors="coerce").dt.strftime("%Y-%m")
        unrecognized_summary = (
            unmatched_df.groupby(["Collaborateur", "No collaborateur", "collab_match_status", "collab_key"], dropna=False)
            .size()
            .reset_index(name="Row Count")
            .sort_values("Row Count", ascending=False)
            .reset_index(drop=True)
        )

    unmatched_mapping = ltr_unmatched_mapping_report(env, df, matched_path, unrecognized_summary)

    sheets = {
        "SUMMARY_BY_MONTH": summary_by_month,
        "ALL_INFRACTIONS": all_infractions,
        "OVER_50H_WEEK": over50_detail,
        "STREAK_7DAYS": streak_detail,
        "SPAN_OVER_14H": span_detail,
        "REST_UNDER_11H": rest_detail,
        "REST_REVIEW_ALLOWED": rest_review,
        "PAUSE_INSUFF": breaks_detail,
        "PAUSE_AUDIT_SERVICES": breaks_audit,
        "SERVICES_AUDIT": services_audit,
        "CALENDAR_HOUR_SLICES": calendar_slices_df,
        "DATA_QUALITY": data_quality,
        "UNMATCHED_MAPPING": unmatched_mapping,
    }
    ltr_write_workbook(env, workbook_path, sheets)

    return {
        "output_root": output_root,
        "workbook_path": workbook_path,
        "sheets": sheets,
        "summary_by_month": summary_by_month,
        "all_infractions": all_infractions,
        "rest_review": rest_review,
        "data_quality": data_quality,
        "services_audit": services_audit,
        "calendar_slices": calendar_slices_df,
        "breaks_audit": breaks_audit,
        "unrecognized_summary": unrecognized_summary,
        "unrecognized_rows": unrecognized_rows,
        "unmatched_mapping": unmatched_mapping,
        "metrics": {
            "raw_rows": len(df),
            "services": len(services_df),
            "calendar_slices": len(calendar_slices_df),
            "infractions": len(all_infractions),
            "affected_collaborators": all_infractions["collab_uid"].nunique() if "collab_uid" in all_infractions.columns and not all_infractions.empty else 0,
            "rest_review_rows": len(rest_review),
            "data_quality_rows": len(data_quality),
            "unrecognized_rows": int(unmatched_mapping["Row Count"].sum()) if not unmatched_mapping.empty else 0,
        },
    }


def ltr_filtered_df(df: pd.DataFrame, filters: dict) -> pd.DataFrame:
    out = df.copy()
    if out.empty:
        return out
    if filters.get("months") and "TARGET_MONTH" in out.columns:
        out = out[out["TARGET_MONTH"].astype(str).isin(filters["months"])]
    if filters.get("weeks") and "EVENT_DATE" in out.columns:
        week_labels = ltr_week_labels_for_df(out, ["EVENT_DATE"])
        out = out[week_labels.isin(filters["weeks"])]
    if filters.get("rules") and "RULE" in out.columns:
        out = out[out["RULE"].astype(str).isin(filters["rules"])]
    if filters.get("severities") and "SEVERITY" in out.columns:
        out = out[out["SEVERITY"].astype(str).isin(filters["severities"])]
    if filters.get("collaborators") and "Collaborateur" in out.columns:
        out = out[out["Collaborateur"].astype(str).isin(filters["collaborators"])]
    return out


def ltr_filtered_support_df(df: pd.DataFrame, filters: dict, month_cols: list[str]) -> pd.DataFrame:
    out = df.copy()
    if out.empty:
        return out
    if filters.get("months"):
        for col in month_cols:
            if col in out.columns:
                out = out[out[col].astype(str).isin(filters["months"])]
                break
    if filters.get("weeks"):
        week_labels = ltr_week_labels_for_df(
            out,
            ["EVENT_DATE", "service_date", "calendar_date", "week_monday", "start_dt_local", "service_start", "slice_start"],
        )
        if not week_labels.empty and week_labels.astype(str).str.len().gt(0).any():
            out = out[week_labels.isin(filters["weeks"])]
    if filters.get("collaborators") and "Collaborateur" in out.columns:
        out = out[out["Collaborateur"].astype(str).isin(filters["collaborators"])]
    return out


def ltr_summary_from_filtered_infractions(filtered: pd.DataFrame) -> pd.DataFrame:
    if filtered.empty or "TARGET_MONTH" not in filtered.columns:
        return pd.DataFrame(columns=["TARGET_MONTH", "TOTAL_INFRACTIONS"])

    summary = filtered.groupby("TARGET_MONTH", dropna=False).size().reset_index(name="TOTAL_INFRACTIONS")
    for rule in sorted(filtered["RULE"].dropna().astype(str).unique().tolist()) if "RULE" in filtered.columns else []:
        counts = (
            filtered[filtered["RULE"].astype(str).eq(rule)]
            .groupby("TARGET_MONTH", dropna=False)
            .size()
            .rename(rule)
        )
        summary = summary.merge(counts, left_on="TARGET_MONTH", right_index=True, how="left")
    return summary.fillna(0)


def ltr_unrecognized_summary(unrecognized_rows: pd.DataFrame) -> pd.DataFrame:
    columns = ["Collaborateur", "No collaborateur", "collab_match_status", "collab_key", "Row Count"]
    if unrecognized_rows.empty:
        return pd.DataFrame(columns=columns)
    return (
        unrecognized_rows.groupby(["Collaborateur", "No collaborateur", "collab_match_status", "collab_key"], dropna=False)
        .size()
        .reset_index(name="Row Count")
        .sort_values("Row Count", ascending=False)
        .reset_index(drop=True)
    )


def render_ltr_chart(title: str, df: pd.DataFrame, x_col: str, y_col: str) -> None:
    st.subheader(title)
    if df.empty or x_col not in df.columns or y_col not in df.columns:
        st.info("Aucune donnée de graphique.")
        return
    chart_df = df[[x_col, y_col]].copy()
    chart_df = chart_df.rename(columns={x_col: "Catégorie", y_col: "Nombre"})
    chart_df["Catégorie"] = chart_df["Catégorie"].astype(str)

    try:
        import altair as alt

        chart = (
            alt.Chart(chart_df)
            .mark_bar()
            .encode(
                x=alt.X("Catégorie:N", title=None, sort=None),
                y=alt.Y("Nombre:Q", title=None),
            )
            .properties(title=None)
        )
        st.altair_chart(chart, use_container_width=True)
    except Exception:
        st.bar_chart(chart_df.set_index("Catégorie"))


def render_ltr_dashboard(result: dict) -> None:
    all_infractions = result["all_infractions"]
    unmatched_mapping = result.get("unmatched_mapping", pd.DataFrame())
    filter_cols = st.columns(4)
    months = sorted(all_infractions["TARGET_MONTH"].dropna().astype(str).unique().tolist()) if "TARGET_MONTH" in all_infractions.columns and not all_infractions.empty else []
    weeks = []
    if "EVENT_DATE" in all_infractions.columns and not all_infractions.empty:
        weeks = sorted(label for label in ltr_week_labels_for_df(all_infractions, ["EVENT_DATE"]).dropna().astype(str).unique().tolist() if label)
    rules = sorted(all_infractions["RULE"].dropna().astype(str).unique().tolist()) if "RULE" in all_infractions.columns and not all_infractions.empty else []
    collaborators = sorted(all_infractions["Collaborateur"].dropna().astype(str).unique().tolist()) if "Collaborateur" in all_infractions.columns and not all_infractions.empty else []
    filters = {
        "months": filter_cols[0].multiselect("Mois", months),
        "weeks": filter_cols[1].multiselect("Semaines", weeks),
        "rules": filter_cols[2].multiselect("Règle", rules),
        "collaborators": filter_cols[3].multiselect("Collaborateur", collaborators),
    }
    filtered = ltr_filtered_df(all_infractions, filters)
    filtered_services = ltr_filtered_support_df(result["services_audit"], filters, ["service_month", "TARGET_MONTH"])
    filtered_calendar = ltr_filtered_support_df(result["calendar_slices"], filters, ["target_month", "TARGET_MONTH"])
    filtered_rest_review = ltr_filtered_support_df(result["rest_review"], filters, ["TARGET_MONTH"])
    unrecognized_rows = result.get("unrecognized_rows", pd.DataFrame())
    if unrecognized_rows.empty and not result["unrecognized_summary"].empty:
        filtered_unrecognized_summary = ltr_filtered_support_df(result["unrecognized_summary"], filters, [])
        filtered_unrecognized_count = int(filtered_unrecognized_summary["Row Count"].sum()) if "Row Count" in filtered_unrecognized_summary.columns else 0
    else:
        filtered_unrecognized_rows = ltr_filtered_support_df(unrecognized_rows, filters, ["TARGET_MONTH"])
        filtered_unrecognized_summary = ltr_unrecognized_summary(filtered_unrecognized_rows)
        filtered_unrecognized_count = len(filtered_unrecognized_rows)
    unmatched_mapping_count = int(unmatched_mapping["Row Count"].sum()) if not unmatched_mapping.empty and "Row Count" in unmatched_mapping.columns else filtered_unrecognized_count
    filtered_summary = ltr_summary_from_filtered_infractions(filtered)

    metric_cols = st.columns(5)
    metric_cols[0].metric("Infractions", f"{len(filtered):,}")
    affected = filtered["collab_uid"].nunique() if "collab_uid" in filtered.columns and not filtered.empty else 0
    metric_cols[1].metric("Collaborateurs concernés", f"{affected:,}")
    metric_cols[2].metric("Services", f"{len(filtered_services):,}")
    metric_cols[3].metric("Créneaux calendrier", f"{len(filtered_calendar):,}")
    metric_cols[4].metric("Mapping non reconnu", f"{unmatched_mapping_count:,}")

    chart_cols = st.columns(2)
    with chart_cols[0]:
        render_ltr_chart("Infractions par mois", filtered_summary, "TARGET_MONTH", "TOTAL_INFRACTIONS")
    with chart_cols[1]:
        if filtered.empty or "RULE" not in filtered.columns:
            st.subheader("Infractions par règle")
            st.info("Aucune donnée de graphique.")
        else:
            by_rule = filtered.groupby("RULE").size().reset_index(name="Count")
            render_ltr_chart("Infractions par règle", by_rule, "RULE", "Count")

    if filtered.empty or "Collaborateur" not in filtered.columns:
        st.subheader("Principaux collaborateurs")
        st.info("Aucune donnée de graphique.")
    else:
        top_collabs = filtered.groupby("Collaborateur").size().sort_values(ascending=False).head(15).reset_index(name="Count")
        render_ltr_chart("Principaux collaborateurs", top_collabs, "Collaborateur", "Count")

    st.subheader("Infractions filtrées")
    st.dataframe(filtered, width="stretch", hide_index=True)
    if not filtered.empty:
        st.download_button(
            "Télécharger le CSV des infractions filtrées",
            filtered.to_csv(index=False, encoding="utf-8-sig"),
            file_name="ltr_filtered_infractions.csv",
            mime="text/csv",
        )

    tab_names = ["Résumé", "Revue des repos", "Services", "Audit des pauses", "Créneaux calendrier", "Non reconnus"]
    tabs = st.tabs(tab_names)
    with tabs[0]:
        st.dataframe(filtered_summary, width="stretch", hide_index=True)
    with tabs[1]:
        st.dataframe(filtered_rest_review, width="stretch", hide_index=True)
    with tabs[2]:
        st.dataframe(filtered_services, width="stretch", hide_index=True)
    with tabs[3]:
        st.dataframe(ltr_filtered_support_df(result["breaks_audit"], filters, ["TARGET_MONTH"]), width="stretch", hide_index=True)
    with tabs[4]:
        st.dataframe(filtered_calendar, width="stretch", hide_index=True)
    with tabs[5]:
        collaborator_mapping = unmatched_mapping[
            unmatched_mapping.get("Entity Type", pd.Series(dtype=object)).astype(str).eq("Collaborateur")
        ].copy() if not unmatched_mapping.empty else filtered_unrecognized_summary
        client_mapping = unmatched_mapping[
            unmatched_mapping.get("Entity Type", pd.Series(dtype=object)).astype(str).eq("Client")
        ].copy() if not unmatched_mapping.empty else pd.DataFrame(columns=["Entity Type", "Name", "ID", "Match Status", "Row Count"])

        st.subheader("Collaborateurs non reconnus")
        st.dataframe(collaborator_mapping, width="stretch", hide_index=True)

        st.subheader("Clients non reconnus")
        st.dataframe(client_mapping, width="stretch", hide_index=True)


def render_ltr_task() -> None:
    st.title("Contrôles LTR")
    st.caption("Exécute les contrôles LTR hybrides et crée le classeur Excel multi-feuilles avec un tableau de bord d'audit.")

    cols = st.columns(3)
    matched_file = cols[0].file_uploader("Fichier mapping", type=["xlsx", "xls"], key="ltr_matched")
    rda_file = cols[1].file_uploader("Fichier RDA", type=["xlsx", "xls", "csv"], key="ltr_rda")
    output_name = cols[2].text_input("Nom du dossier de sortie", value="")

    option_cols = st.columns([1.2, 2.8])
    choose_week_output = option_cols[0].checkbox("Choisir des dates spécifiques", key="ltr_choose_week_dates")
    week_date_range = None
    if choose_week_output:
        previous_monday, previous_sunday = ltr_previous_full_week()
        week_date_range = option_cols[1].date_input(
            "Dates à inclure dans le classeur semaine",
            value=(previous_monday, previous_sunday),
            key="ltr_week_date_range",
        )

    action_cols = st.columns([2, 1])
    with action_cols[0]:
        invalid_week_dates = (
            choose_week_output
            and (not isinstance(week_date_range, (tuple, list)) or len(week_date_range) != 2 or week_date_range[0] > week_date_range[1])
        )
        run_ltr = st.button(
            "Lancer les contrôles LTR",
            type="primary",
            disabled=matched_file is None or rda_file is None or invalid_week_dates,
            width="stretch",
        )

    if invalid_week_dates:
        st.error("La plage de dates doit contenir une date de début et une date de fin valides.")

    if run_ltr:
        render_blocking_run_warning()
        progress = st.progress(0.0, text="Démarrage des contrôles LTR")
        try:
            progress.progress(0.1, text="Chargement des fichiers et de la logique notebook")
            result = ltr_process(matched_file, rda_file, output_name)
            if choose_week_output:
                week_start, week_end = week_date_range
                progress.progress(0.82, text="Création du classeur LTR semaine")
                week_workbook_path = ltr_build_week_workbook(result, week_start, week_end)
                package_path = ltr_zip_full_and_week(result, week_workbook_path, week_start, week_end)
                result["week_workbook_path"] = week_workbook_path
                result["package_path"] = package_path
                result["week_date_range"] = (week_start.isoformat(), week_end.isoformat())
            else:
                result.pop("week_workbook_path", None)
                result.pop("package_path", None)
                result.pop("week_date_range", None)
            progress.progress(1.0, text="Contrôles LTR terminés")
            st.session_state["latest_ltr_result"] = result
        except Exception as exc:
            progress.empty()
            st.exception(exc)
            return

    result = st.session_state.get("latest_ltr_result")
    workbook_path = result["workbook_path"] if result else None
    package_path = result.get("package_path") if result else None
    with action_cols[1]:
        if package_path:
            render_download_or_placeholder(package_path, "Télécharger full + semaine ZIP", key="ltr_week_package")
        else:
            render_download_or_placeholder(workbook_path, "Télécharger le classeur LTR", key="ltr_main_workbook")

    if result:
        if result.get("package_path"):
            st.success("Classeur LTR complet et classeur semaine créés dans un ZIP.")
            st.caption(
                "Semaine sélectionnée: "
                f"{result.get('week_date_range', ('', ''))[0]} à {result.get('week_date_range', ('', ''))[1]}"
            )
        else:
            st.success("Classeur LTR créé et disponible au téléchargement.")
        render_ltr_dashboard(result)


# ============================================================
# Audit Webfleet-RDA — low-level helpers
# ============================================================
